

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.utils import timezone
from datetime import datetime

def gerar_excel_generic(queryset, tipo_relatorio, data_inicio=None, data_fim=None):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"relatorio_{tipo_relatorio}_{timezone.now().strftime('%d_%m_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    
    # Tenta buscar o perfil da empresa através do usuário da primeira conta
    perfil = queryset.first().user.company_profile if queryset.exists() else None
    nome_exibir = perfil.nome_empresa if perfil else "Relatório Financeiro"
    
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Empresa: {nome_exibir}"
    ws['A1'].font = Font(bold=True, size=12)
    if data_inicio and data_fim:
        try:
            d_ini = datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
            d_fim = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')
            ws.merge_cells('A2:D2')
            ws['A2'] = f"Período: {d_ini} à {d_fim}"
        except:
            pass
    ws.append([]) # Linha em branco para separar do cabeçalho
    

    # AJUSTE CIRÚRGICO: Cabeçalhos condicionais
    if tipo_relatorio == 'pagar':
        headers = ['Nome', 'Descrição', 'Vencimento', 'Valor', 'Status', 'Categoria', 'Área-DRE', 'Centro de Custo', 'Banco', 'Forma Pag.']
    else: # receber
        # Removido 'Centro de Custo' desta lista
        headers = ['Nome', 'Descrição', 'Vencimento', 'Valor', 'Status', 'Categoria', 'Área-DRE', 'Banco', 'Forma Pag.']

    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for conta in queryset:
        status_text = "Pago" if tipo_relatorio == 'pagar' and conta.is_paid else \
                      "Recebido" if tipo_relatorio != 'pagar' and conta.is_received else "Aberto"

        banco = str(conta.bank_account) if conta.bank_account else "-"
        categoria = str(conta.category.name) if conta.category else "-"
        forma_pag = conta.get_payment_method_display() if hasattr(conta, 'get_payment_method_display') else conta.payment_method
        dre_area = conta.get_dre_area_display() if hasattr(conta, 'get_dre_area_display') else "-"
        
        # AJUSTE CIRÚRGICO: Montagem da linha dinâmica
        row = [
            conta.name,
            conta.description,
            conta.due_date,
            conta.amount,
            status_text,
            categoria,
            dre_area,
        ]

        # Adiciona Centro de Custo APENAS se não for 'receber'
        if tipo_relatorio != 'receber':
            c_custo = str(conta.centro_custo.nome) if hasattr(conta, 'centro_custo') and conta.centro_custo else "-"
            row.append(c_custo)

        row.extend([banco, forma_pag])
        ws.append(row)
        # Formata a coluna 'Valor' (coluna D, índice 4) para padrão contábil brasileiro
        cell_valor = ws.cell(row=ws.max_row, column=4)
        cell_valor.number_format = '#,##0.00' 
        # ------------------------------------------------

    wb.save(response)
    return response

def gerar_pdf_generic(queryset, tipo_relatorio, data_inicio=None, data_fim=None):
    response = HttpResponse(content_type='application/pdf')
    filename = f"relatorio_{tipo_relatorio}_{timezone.now().strftime('%d_%m_%Y')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"' 

    titulo_texto = "Contas a Pagar" if tipo_relatorio == 'pagar' else "Contas a Receber"

    doc = SimpleDocTemplate(
        response, 
        pagesize=landscape(A4), 
        title=f"Relatório de {titulo_texto}" 
    )
    
    elements = []
    # AJUSTE VITAL: Definir os estilos antes de qualquer uso
    styles = getSampleStyleSheet() 

    # Criamos um estilo exclusivo para o cabeçalho (Alinhado à esquerda, Fonte 12, com Espaçamento)
    estilo_info = styles['Normal'].__class__('EstiloCabecalho', styles['Normal'])
    estilo_info.fontSize = 12
    estilo_info.leading = 16  # Espaçamento entre linhas para evitar sobreposição
    estilo_info.alignment = 0 # 0 = Alinhado à Esquerda

    # Busca o perfil da empresa
    perfil = queryset.first().user.company_profile if queryset.exists() else None
    nome_exibir = perfil.nome_empresa if perfil else "Relatório Financeiro"
    
    # 1. Nome da Empresa (Negrito)
    elements.append(Paragraph(f"<b>Empresa:</b> {nome_exibir}", estilo_info))
    
    # 2. Período (Negrito)
    if data_inicio and data_fim:
        try:
            d_ini = datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
            d_fim = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')
            elements.append(Paragraph(f"<b>Período:</b> {d_ini} à {d_fim}", estilo_info))
        except Exception:
            elements.append(Paragraph(f"<b>Período:</b> {data_inicio} à {data_fim}", estilo_info))
    
    # 3. Título do Relatório (Negrito)
    elements.append(Paragraph(f"<b>Relatório:</b> {titulo_texto}", estilo_info))

    elements.append(Spacer(1, 20)) # Espaço generoso antes da tabela

    # AJUSTE CIRÚRGICO: Cabeçalhos e Larguras condicionais
    if tipo_relatorio == 'receber':
        headers = ['Nome', 'Desc.', 'Venc.', 'Valor', 'Status', 'Cat.', 'DRE', 'Banco', 'Forma']
        # Nome aumentado para 160 (antes era menor no cálculo implícito)
        col_widths = [180, 100, 60, 80, 50, 70, 70, 80, 60]
    else:
        headers = ['Nome', 'Desc.', 'Venc.', 'Valor', 'Status', 'Cat.', 'DRE', 'C. Custo', 'Banco', 'Forma']
        # Coluna Nome aumentada para 145 e Desc. para 115 (Total 780 pontos)
        col_widths = [180, 100, 55, 75, 45, 65, 65, 65, 75, 55]

    data = [headers]
    total = 0

    for conta in queryset:
        status_text = "Pago" if tipo_relatorio == 'pagar' and conta.is_paid else \
                      "Recebido" if tipo_relatorio != 'pagar' and conta.is_received else "Aberto"
        
        categoria = str(conta.category.name) if conta.category else "-"
        banco = str(conta.bank_account) if conta.bank_account else "-"
        forma_pag = conta.get_payment_method_display() if hasattr(conta, 'get_payment_method_display') else conta.payment_method
        dre_area = conta.get_dre_area_display() if hasattr(conta, 'get_dre_area_display') else "-"

        # Estilo para o nome aceitar quebra de linha (sem cortes)
        style_nome = styles["Normal"]
        style_nome.fontSize = 7
        style_nome.leading = 8 

        nome_paragrafo = Paragraph(conta.name, style_nome) # Aqui o nome fica completo
        desc_curta = conta.description[:25] + '..' if len(conta.description) > 25 else conta.description
        
        row_pdf = [
            nome_paragrafo,
            desc_curta,
            conta.due_date.strftime('%d/%m/%Y'),
            f"R$ {conta.amount:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
            status_text,
            categoria[:12],
            dre_area[:12],
        ]

        if tipo_relatorio != 'receber':
            c_custo = str(conta.centro_custo.nome) if hasattr(conta, 'centro_custo') and conta.centro_custo else "-"
            row_pdf.append(c_custo[:12])

        row_pdf.extend([banco[:12], forma_pag])
        data.append(row_pdf)
        total += conta.amount

    # AJUSTE CIRÚRGICO: Linha de Total com número de colunas correto
    if tipo_relatorio == 'receber':
        data.append(['', '', 'TOTAL:', f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), '', '', '', '', '']) # 9 colunas
    else:
        data.append(['', '', 'TOTAL:', f"R$ {total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'), '', '', '', '', '', '']) # 10 colunas

    table = Table(data, colWidths=col_widths)
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('ALIGN', (2, -1), (2, -1), 'RIGHT'), 
    ])
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    return response

# --- AS FUNÇÕES ABAIXO NÃO SOFRERAM ALTERAÇÕES PARA GARANTIR INTEGRIDADE ---

def gerar_excel_orcamento(dados_orcamento, ano):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"relatorio_orcamento_{ano}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb = openpyxl.Workbook()
    ws_despesa = wb.active
    ws_despesa.title = "Despesas"
    headers = ['Categoria']
    meses = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    for mes in meses:
        headers.extend([f'{mes} (Prev)', f'{mes} (Real)'])
    headers.extend(['Total Prev.', 'Total Real.'])
    ws_despesa.append(headers)
    for cell in ws_despesa[1]: cell.font = Font(bold=True)
    for cat_name, data in dados_orcamento['despesas'].items():
        row = [cat_name]
        for i in range(12):
            row.append(data['orcado_mensal'][i])
            row.append(data['realizado_mensal'][i])
        row.append(data['total_orcado_ano'])
        row.append(data['total_realizado_ano'])
        ws_despesa.append(row)
    ws_receita = wb.create_sheet(title="Receitas")
    ws_receita.append(headers)
    for cell in ws_receita[1]: cell.font = Font(bold=True)
    for cat_name, data in dados_orcamento['receitas'].items():
        row = [cat_name]
        for i in range(12):
            row.append(data['orcado_mensal'][i])
            row.append(data['realizado_mensal'][i])
        row.append(data['total_orcado_ano'])
        row.append(data['total_realizado_ano'])
        ws_receita.append(row)
    wb.save(response)
    return response

def gerar_pdf_orcamento(dados_orcamento, ano):
    response = HttpResponse(content_type='application/pdf')
    filename = f"relatorio_orcamento_{ano}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), title=f"Relatório Orçamentário {ano}")
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph(f"Relatório Orçamentário Anual - {ano}", styles['Title']))
    elements.append(Spacer(1, 12))
    headers = ['Tipo', 'Categoria', 'Total Planejado', 'Total Realizado', 'Diferença', 'Status']
    data = [headers]
    for cat_name, info in dados_orcamento['receitas'].items():
        prev = info['total_orcado_ano']
        real = info['total_realizado_ano']
        diff = real - prev
        status = "Acima da Meta" if diff >= 0 else "Abaixo da Meta"
        data.append(["Receita", cat_name[:25], f"R$ {prev:,.2f}", f"R$ {real:,.2f}", f"R$ {diff:,.2f}", status])
    data.append(['', '', '', '', '', ''])
    for cat_name, info in dados_orcamento['despesas'].items():
        prev = info['total_orcado_ano']
        real = info['total_realizado_ano']
        diff = prev - real
        status = "Dentro do Orçamento" if diff >= 0 else "Estourou Orçamento"
        data.append(["Despesa", cat_name[:25], f"R$ {prev:,.2f}", f"R$ {real:,.2f}", f"R$ {diff:,.2f}", status])
    col_widths = [60, 200, 100, 100, 100, 120]
    table = Table(data, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    return response

def gerar_excel_dre(dados_dre, periodo_str):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"relatorio_dre_{timezone.now().strftime('%d_%m_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DRE Gerencial"
    ws.append([f"DRE Gerencial - Período: {periodo_str}"])
    ws.append([])
    headers = ['Estrutura', 'Valor (R$)', 'Análise Vertical (%)']
    ws.append(headers)
    for cell in ws[3]: cell.font = Font(bold=True)
    ordem_dre = [
        ('Receita Bruta', 'receita_bruta'), ('(-) Impostos', 'impostos'), ('(=) Receita Líquida', 'receita_liquida'),
        ('(-) Custos (CMV/CSP)', 'custos'), ('(=) Lucro Bruto', 'lucro_bruto'), ('(-) Despesas Operacionais', 'despesas_operacionais'),
        ('(=) EBITDA', 'ebitda'), ('(-) Depreciação/Amortização', 'depreciacao'), ('(=) EBIT', 'ebit'),
        ('(-) Resultado Financeiro/Não Op.', 'nao_operacionais'), ('(=) LAIR', 'lair'), ('(-) IRPJ/CSLL', 'tributacao'),
        ('(=) Lucro Líquido', 'lucro_liquido'), ('(-) Distribuição de Lucros', 'distribuicao_lucro'), ('(=) Resultado Final', 'resultado_final'),
    ]
    receita_liquida = dados_dre.get('receita_liquida', 0)
    for label, key in ordem_dre:
        valor = dados_dre.get(key, 0)
        av = (valor / receita_liquida * 100) if receita_liquida and receita_liquida > 0 else 0
        ws.append([label, valor, f"{av:.2f}%"])
    for row in ws.iter_rows(min_row=4, max_col=2, max_row=19):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"R$ "#,##0.00'
    wb.save(response)
    return response

def gerar_pdf_dre(dados_dre, periodo_str):
    response = HttpResponse(content_type='application/pdf')
    filename = f"relatorio_dre_{timezone.now().strftime('%d_%m_%Y')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    doc = SimpleDocTemplate(response, pagesize=A4, title="Relatório DRE")
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph(f"DRE Gerencial - {periodo_str}", styles['Title']))
    elements.append(Spacer(1, 20))
    data = [['Estrutura', 'Valor', 'A.V. %']]
    ordem_dre = [
        ('Receita Bruta', 'receita_bruta'), ('(-) Impostos', 'impostos'), ('(=) Receita Líquida', 'receita_liquida'),
        ('(-) Custos (CMV/CSP)', 'custos'), ('(=) Lucro Bruto', 'lucro_bruto'), ('(-) Despesas Operacionais', 'despesas_operacionais'),
        ('(=) EBITDA', 'ebitda'), ('(-) Depreciação', 'depreciacao'), ('(=) EBIT', 'ebit'),
        ('(-) Não Operacionais', 'nao_operacionais'), ('(=) LAIR', 'lair'), ('(-) IRPJ/CSLL', 'tributacao'),
        ('(=) Lucro Líquido', 'lucro_liquido'), ('(-) Distribuição', 'distribuicao_lucro'), ('(=) Resultado Final', 'resultado_final'),
    ]
    receita_liquida = dados_dre.get('receita_liquida', 0)
    table_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]
    for index, (label, key) in enumerate(ordem_dre):
        valor = dados_dre.get(key, 0)
        av = (valor / receita_liquida * 100) if receita_liquida and receita_liquida > 0 else 0
        if label.startswith('(=)'):
            row_idx = index + 1
            table_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightgrey))
        data.append([label, f"R$ {valor:,.2f}", f"{av:.2f}%"])
    table = Table(data, colWidths=[250, 120, 80])
    table.setStyle(TableStyle(table_styles))
    elements.append(table)
    doc.build(elements)
    return response

def gerar_excel_fluxo_caixa(dados_fc, periodo_str):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"relatorio_fluxo_caixa_{timezone.now().strftime('%d_%m_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fluxo de Caixa"
    ws.append([f"Fluxo de Caixa Realizado - {periodo_str}"])
    ws.append([])
    headers = ['Indicador'] + dados_fc['labels'] + ['TOTAL']
    ws.append(headers)
    for cell in ws[3]: cell.font = Font(bold=True)
    ws.append(['Entradas'] + dados_fc['entradas'] + [sum(dados_fc['entradas'])])
    ws.append(['Saídas'] + dados_fc['saidas'] + [sum(dados_fc['saidas'])])
    ws.append(['Saldo (Geração de Caixa)'] + dados_fc['geracao_caixa'] + [sum(dados_fc['geracao_caixa'])])
    wb.save(response)
    return response

def gerar_pdf_fluxo_caixa(dados_fc, periodo_str):
    response = HttpResponse(content_type='application/pdf')
    filename = f"relatorio_fluxo_caixa_{timezone.now().strftime('%d_%m_%Y')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), title="Relatório Fluxo de Caixa")
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph(f"Fluxo de Caixa Realizado - {periodo_str}", styles['Title']))
    elements.append(Spacer(1, 20))
    headers = ['Mês', 'Entradas', 'Saídas', 'Saldo']
    data = [headers]
    total_ent = total_sai = total_sal = 0
    for i, mes in enumerate(dados_fc['labels']):
        ent, sai, sal = dados_fc['entradas'][i], dados_fc['saidas'][i], dados_fc['geracao_caixa'][i]
        total_ent += ent; total_sai += sai; total_sal += sal
        data.append([mes, f"R$ {ent:,.2f}", f"R$ {sai:,.2f}", f"R$ {sal:,.2f}"])
    data.append(['TOTAL', f"R$ {total_ent:,.2f}", f"R$ {total_sai:,.2f}", f"R$ {total_sal:,.2f}"])
    table = Table(data, colWidths=[100, 150, 150, 150])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black), ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    doc.build(elements)
    return response