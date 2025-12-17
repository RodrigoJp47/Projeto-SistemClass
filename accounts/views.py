

# --- 1. Bibliotecas Padrão do Python ---
import io
import json
import locale
import logging
import os
import re
import xml.etree.ElementTree as ET
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta
from decimal import Decimal
from io import StringIO
from urllib.parse import urlparse
import requests
import time

import stripe

from django.views.decorators.csrf import csrf_exempt # Vamos usar depois no webhook


# --- 2. Bibliotecas de Terceiros (Pip) ---
import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from ofxparse import OfxParser
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from requests_oauthlib import OAuth2Session
from django.conf import settings

# --- 3. Importações do Django ---
from django.contrib import messages
from django.contrib.auth import (
    authenticate, login, logout, update_session_auth_hash
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (
    Sum, Count, Q, F, Min, ExpressionWrapper, fields, Avg
)
from django.db.models.functions import TruncMonth
from django.db.models.signals import post_save
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date

# --- 4. Importações Locais do Projeto (.forms, .models, .decorators) ---
from .decorators import (
    subscription_required, check_employee_permission, owner_required,
    module_access_required
)
from .forms import (
    PayableAccountForm, ReceivableAccountForm, OFXImportForm,
    MetaFaturamentoForm, CustomUserCreationForm, ContractForm,
    EmployeeCreationForm, CompanyProfileForm, CompanyDocumentForm,
    ProdutoServicoForm, VendedorForm, ClienteForm, OrcamentoForm
)
from .models import (
    PayableAccount, ReceivableAccount, Category, DRE_AREAS, BankAccount,
    OFXImport, MetaFaturamento, Subscription, Contract, Cliente,
    BPOClientLink, CompanyUserLink, create_subscription_for_new_user,
    CompanyProfile, CompanyDocument, Venda, ItemVenda, ProdutoServico,
    Vendedor, PagamentoVenda, PAYMENT_METHODS, Cidade,
    ContaAzulCredentials, Precificacao, Orcamento, OrcamentoVenda,
    ItemOrcamento, ClassificacaoAutomatica
)
from .utils_exports import gerar_pdf_generic, gerar_excel_generic
from .utils_inter import buscar_extrato_inter, buscar_saldo_inter 

from .forms import AsaasCredentialsForm 
from .models import AsaasCredentials
from .utils_asaas import buscar_extrato_asaas, buscar_saldo_asaas

from .forms import OmieCredentialsForm
from .models import OmieCredentials

from .utils_omie import sincronizar_omie_completo

# Imports para o Nibo
from .forms import NiboCredentialsForm
from .models import NiboCredentials
from .utils_nibo import sincronizar_nibo_completo

from .forms import TinyCredentialsForm
from .models import TinyCredentials
from .utils_tiny import sincronizar_tiny_completo


from .utils_inter import buscar_extrato_inter
from .forms import InterCredentialsForm
from .models import InterCredentials
from .forms import MercadoPagoCredentialsForm # <--- Adicionar
from .models import MercadoPagoCredentials    # <--- Adicionar
# Adicione ao topo do accounts/views.py, junto com as outras importações locais
from .utils_mercadopago import buscar_extrato_mercadopago, buscar_saldo_mercadopago


MESES_ABREVIADOS = {
    1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
    7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
}



# Em accounts/views.py

def aprender_classificacao(user, nome, categoria, dre_area, tipo, bank_account=None):
    """
    Salva a regra de classificação APENAS SE ELA AINDA NÃO EXISTIR.
    Isso protege o histórico contra edições pontuais (exceções).
    """
    from .models import ClassificacaoAutomatica 

    if nome and categoria and dre_area:
        # AQUI ESTÁ A MUDANÇA: de update_or_create para get_or_create
        ClassificacaoAutomatica.objects.get_or_create(
            user=user,
            termo__iexact=nome.strip(),
            tipo=tipo,
            defaults={
                'termo': nome.strip(),
                'categoria': categoria,
                'dre_area': dre_area,
                'bank_account': bank_account
            }
        )

def prever_classificacao(user, descricao, tipo):
    """
    Tenta encontrar uma categoria, DRE e Banco baseada na descrição.
    Retorna uma tupla (categoria, dre_area, bank_account) ou (None, None, None).
    """
    # Importação dentro da função para evitar erro circular, se necessário
    from .models import ClassificacaoAutomatica 

    regras = ClassificacaoAutomatica.objects.filter(user=user, tipo=tipo)
    descricao_lower = descricao.lower()
    
    for regra in regras:
        if regra.termo.lower() in descricao_lower:
            # RETORNA 3 VALORES AGORA
            return regra.categoria, regra.dre_area, regra.bank_account
            
    return None, None, None # Retorna 3 Nones

@login_required
@subscription_required
@check_employee_permission('can_access_home')
def home(request):
    if not request.is_managing: 
        try:
            if request.active_subscription.user_type == 'BPO':
                return redirect('bpo_dashboard')
        except (Subscription.DoesNotExist, AttributeError):
            pass 
    
    today = timezone.now().date()
    
    # --- 1. FILTROS DE DATA ---
    period = request.GET.get('period', 'current_month')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
    else:
        end_date = today
        if period == 'current_month':
            start_date = end_date.replace(day=1)
        elif period == '30':
            start_date = end_date - timedelta(days=30)
        elif period == '90':
            start_date = end_date - timedelta(days=90)
        elif period == '180':
            start_date = end_date - timedelta(days=180)
        else:
            start_date = end_date.replace(day=1)

    # --- 2. DADOS DO TOPO (Separados por Receber e Pagar) ---
    
    # RECEBER
    all_receivables = ReceivableAccount.objects.filter(user=request.user, is_received=False)
    total_areceber_aberto = all_receivables.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_areceber_atrasado = all_receivables.filter(due_date__lt=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_areceber_hoje = all_receivables.filter(due_date=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # PAGAR
    all_payables = PayableAccount.objects.filter(user=request.user, is_paid=False)
    total_apagar_aberto = all_payables.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_apagar_atrasado = all_payables.filter(due_date__lt=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_apagar_hoje = all_payables.filter(due_date=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # --- 3. DADOS DO MEIO (Financeiro no Período Filtrado) ---
    
    total_receivable = ReceivableAccount.objects.filter(
        user=request.user, is_received=True, due_date__gte=start_date, due_date__lte=end_date
    ).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or 0
    
    total_payable = PayableAccount.objects.filter(
        user=request.user, is_paid=True, due_date__gte=start_date, due_date__lte=end_date
    ).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or 0
    
    balance = total_receivable - total_payable

    # Saldo Bancário
    bank_accounts_balances = BankAccount.objects.filter(user=request.user).order_by('bank_name')
    saldo_consolidado = Decimal('0.00')

    for account in bank_accounts_balances:
        received = ReceivableAccount.objects.filter(
            user=request.user, bank_account=account, is_received=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        paid = PayableAccount.objects.filter(
            user=request.user, bank_account=account, is_paid=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        account.current_balance = account.initial_balance + received - paid
        saldo_consolidado += account.current_balance

    # --- 4. DADOS DO FUNDO (Gráfico) ---
    payable_by_date = PayableAccount.objects.filter(
        user=request.user, is_paid=True, due_date__gte=start_date, due_date__lte=end_date
    ).exclude(dre_area='NAO_CONSTAR').values('due_date').annotate(total=Sum('amount')).order_by('due_date')
    
    receivable_by_date = ReceivableAccount.objects.filter(
        user=request.user, is_received=True, due_date__gte=start_date, due_date__lte=end_date
    ).exclude(dre_area='NAO_CONSTAR').values('due_date').annotate(total=Sum('amount')).order_by('due_date')
    
    chart_labels, chart_payable, chart_receivable, chart_balance = [], [], [], []
    current_date = start_date
    while current_date <= end_date:
        chart_labels.append(current_date.strftime('%d/%m/%Y'))
        payable_total = next((item['total'] for item in payable_by_date if item['due_date'] == current_date), 0)
        receivable_total = next((item['total'] for item in receivable_by_date if item['due_date'] == current_date), 0)
        chart_payable.append(float(payable_total))
        chart_receivable.append(float(receivable_total))
        chart_balance.append(float(receivable_total - payable_total))
        current_date += timedelta(days=1)
    
    company_name = None
    try:
        if request.user.company_profile.nome_empresa:
            company_name = request.user.company_profile.nome_empresa
    except: pass

    context = {
        # Topo - Receber
        'total_areceber_aberto': total_areceber_aberto,
        'total_areceber_atrasado': total_areceber_atrasado,
        'total_areceber_hoje': total_areceber_hoje,
        
        # Topo - Pagar
        'total_apagar_aberto': total_apagar_aberto,
        'total_apagar_atrasado': total_apagar_atrasado,
        'total_apagar_hoje': total_apagar_hoje,
        
        # Meio
        'total_payable': total_payable,
        'total_receivable': total_receivable,
        'balance': balance,
        'bank_accounts_balances': bank_accounts_balances,
        'saldo_consolidado': saldo_consolidado,
        
        # Fundo
        'chart_labels': chart_labels,
        'chart_payable': chart_payable,
        'chart_receivable': chart_receivable,
        'chart_balance': chart_balance,
        
        'period': period,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        'company_name': company_name,
    }
    return render(request, 'accounts/home.html', context)

# Em accounts/views.py

@login_required
@owner_required
@subscription_required
def company_profile_view(request):
    # 1. Pega ou cria o perfil
    profile, created = CompanyProfile.objects.get_or_create(user=request.user)
    documents = CompanyDocument.objects.filter(user=request.user)

    # Inicializa os formulários com valores padrão (GET)
    # Eles serão sobrescritos dentro do bloco POST se necessário
    profile_form = CompanyProfileForm(instance=profile)
    doc_form = CompanyDocumentForm()

    if request.method == 'POST':
        print("\n=== O QUE O NAVEGADOR MANDOU? ===")
        print(f"Cidade (POST): '{request.POST.get('cidade')}'")
        print(f"IBGE (POST): '{request.POST.get('codigo_municipio')}'")
        print("=====================================\n")
        # --- Lógica do Perfil da Empresa ---
        if 'save_profile' in request.POST:
            profile_form = CompanyProfileForm(request.POST, request.FILES, instance=profile)

            if profile_form.is_valid():
                # 1. Captura os arquivos ANTES de salvar (para garantir leitura correta)
                arquivo_certificado = request.FILES.get('certificado_digital')
                senha_cert = profile_form.cleaned_data.get('senha_certificado')
                
                # Lê o binário do arquivo para memória se ele foi enviado
                conteudo_certificado = None
                if arquivo_certificado and senha_cert:
                    conteudo_certificado = arquivo_certificado.read()

                # 2. Salva no Banco de Dados Local
                profile = profile_form.save(commit=False)
                # Se a Inscrição Estadual estiver vazia, salva "0000000000000" no banco
                if not profile.inscricao_estadual:
                    profile.inscricao_estadual = "0000000000000"
                profile.save()

                # --- ESPIÃO DE DADOS (ADICIONE ISSO) ---
                print("\n=== DEBUG DADOS DO PERFIL ===")
                print(f"Cidade no Banco: '{profile.cidade}'")
                print(f"IBGE no Banco: '{profile.codigo_municipio}'")
                print(f"Estado no Banco: '{profile.estado}'")
                print(f"Tem espaços em branco? Cidade len={len(profile.cidade or '')}")
                print("=============================\n")
                # ---------------------------------------

                # 3. INTEGRAÇÃO COM FOCUS NFE
                # Só tenta integrar se tiver os dados mínimos
                # 3. INTEGRAÇÃO COM FOCUS NFE (LÓGICA INTELIGENTE SAAS)
                if profile.cnpj:
                    try:
                        # Ambiente Focus (Dinâmico REALMENTE baseado no settings)
                        BASE_URL = settings.FOCUS_API_URL
                        
                        if settings.DEBUG:
                            API_TOKEN = settings.NFE_TOKEN_HOMOLOGACAO
                            print("⚠️ MODO HOMOLOGAÇÃO ATIVADO")
                        else:
                            API_TOKEN = settings.NFE_TOKEN_PRODUCAO
                            print("🚀 MODO PRODUÇÃO ATIVADO")

                        # --- DEBUG DE CONEXÃO ---
                        token_masked = f"{API_TOKEN[:4]}...{API_TOKEN[-4:]}" if API_TOKEN else "None"
                        print(f"\n[DEBUG API] URL Base: {BASE_URL}")
                        print(f"[DEBUG API] Token Usado: {token_masked}")
                        print(f"[DEBUG API] Ambiente DEBUG: {settings.DEBUG}")
                        # ------------------------

                        # Limpeza de Dados
                        cnpj_limpo = re.sub(r'\D', '', profile.cnpj or '')
                        cep_limpo = re.sub(r'\D', '', profile.cep or '')
                        telefone_limpo = re.sub(r'\D', '', profile.telefone_contato or '')
                        
                        # URL base de empresas
                        url_base_empresas = f"{BASE_URL.rstrip('/')}/v2/empresas"
                        
                        # Monta o Payload com FALBACK (Se falhar no banco, pega do POST)
                        cidade_envio = profile.cidade
                        if not cidade_envio:
                            cidade_envio = request.POST.get('cidade', '') # Pega direto do navegador

                        cod_mun_envio = profile.codigo_municipio
                        if not cod_mun_envio:
                            cod_mun_envio = request.POST.get('codigo_municipio', '')

                        # Lógica Corrigida para Inscrição Estadual
                        raw_ie = getattr(profile, 'inscricao_estadual', '')
                        ie_limpa = str(raw_ie).replace(".", "").replace("-", "").replace("/", "").strip() if raw_ie else ""
                        
                        print(f"\n[DEBUG] Inscrição Estadual - Raw: '{raw_ie}' | Limpa: '{ie_limpa}'")
                        
                        # Se depois de limpar estiver vazio, força ZEROS (Bypass Focus)
                        if not ie_limpa:
                            ie_limpa = "0000000000000"
                            print("[DEBUG] Forçando ZEROS (0000000000000)")

                        payload_empresa = {
                            "nome": profile.nome_empresa,
                            "nome_fantasia": profile.nome_fantasia or profile.nome_empresa,
                            "cnpj": cnpj_limpo,
                            
                            "regime_tributario": profile.regime_tributario, 
                            "bairro": profile.bairro,
                            "cep": cep_limpo,
                            
                            # AQUI ESTÁ A CORREÇÃO BLINDADA:
                            "municipio": cidade_envio, 
                            "codigo_municipio": cod_mun_envio,
                            
                            "uf": profile.estado,
                            "logradouro": profile.endereco,
                            "numero": profile.numero or "S/N",
                            
                            # Lógica corrigida: Usa a variável tratada acima
                            "inscricao_estadual": ie_limpa,
                            "inscricao_municipal": (profile.inscricao_municipal or "").replace(".", "").replace("-", "").replace("/", "").strip(),
                            
                            "telefone": re.sub(r'\D', '', profile.telefone_contato or ''),
                            "email": profile.email_contato,
                            # Se for ISENTO, não pode emitir NFe (Produtos)
                            "habilita_nfe": False if ie_limpa == "ISENTO" else True,
                            "habilita_nfse": True,
                            "optante_simples_nacional": True if profile.regime_tributario == '1' else False,
                            "incentivador_cultural": profile.incentivador_cultural,
                            
                            # --- NOVOS CAMPOS DE NUMERAÇÃO ---
                            "proximo_numero_nfe_producao": profile.proximo_numero_nfe,
                            "serie_nfe_producao": profile.serie_nfe,
                            "proximo_numero_nfse_producao": profile.proximo_numero_nfse,
                            "serie_nfse_producao": profile.serie_nfse,
                        }

                        # --- LÓGICA DE LOGO (BASE64) ---
                        if profile.arquivo_logo:
                            try:
                                # Abre o arquivo do disco/S3, lê e converte
                                profile.arquivo_logo.open('rb')
                                logo_content = profile.arquivo_logo.read()
                                import base64
                                logo_base64 = base64.b64encode(logo_content).decode('utf-8')
                                payload_empresa["arquivo_logo_base64"] = logo_base64
                                print(f"[DEBUG] Logo convertido para Base64. Tamanho: {len(logo_base64)}")
                            except Exception as e_logo:
                                print(f"[ERRO] Falha ao ler logo: {e_logo}")
                        # -------------------------------
                        
                        print(f"\n[DEBUG FINAL] Payload Inscricao Estadual: '{payload_empresa['inscricao_estadual']}'")
                        print(f"[DEBUG FINAL] Profile Inscricao Estadual (Memory): '{profile.inscricao_estadual}'")
                        print(f"[DEBUG FINAL] Payload Completo: {payload_empresa}")

                        # Tenta enviar o certificado JÁ NA CRIAÇÃO (Tentativa 1)
                        cert_enviado_no_post = False
                        if conteudo_certificado and senha_cert:
                            import base64
                            arquivo_base64 = base64.b64encode(conteudo_certificado).decode('utf-8')
                            payload_empresa["arquivo_certificado_base64"] = arquivo_base64
                            payload_empresa["senha_certificado"] = senha_cert
                            cert_enviado_no_post = True

                        print(f"--- TENTATIVA 1: CRIAR EMPRESA {cnpj_limpo} ---")
                        # Tenta CRIAR (POST)
                        response = requests.post(url_base_empresas, json=payload_empresa, auth=(API_TOKEN, ""))
                        
                        # --- DEBUG DO ERRO DE CRIAÇÃO ---
                        if response.status_code not in [200, 201]:
                            print(f"❌ FALHA AO CRIAR: {response.status_code}")
                            print(f"MOTIVO: {response.text}")
                        # --------------------------------
                        
                        # Se der erro de "Já existe" (422 ou 409) ou "404" (Token restrito tentando POST geral)
                        if response.status_code in [422, 409, 403]: 
                            
                            print(f"--- TENTATIVA 2: ATUALIZAR EMPRESA (PUT) ---")
                            # Tenta ATUALIZAR (PUT) na URL específica
                            url_put = f"{url_base_empresas}/{cnpj_limpo}"
                            
                            # No PUT, não mandamos o CNPJ no corpo (já está na URL) para evitar alguns erros
                            payload_put = payload_empresa.copy()
                            
                            response = requests.put(url_put, json=payload_put, auth=(API_TOKEN, ""))

                        # --- TRATAMENTO ESPECIAL PARA TOKEN DE EMPRESA (404 NO ENDPOINT /empresas) ---
                        if response.status_code == 404 and "Endpoint não encontrado" in response.text:
                            print("⚠️ DETECTADO TOKEN DE EMPRESA (SEM ACESSO A /empresas)")
                            print("✅ Permitindo salvar perfil localmente para emissão direta.")
                            messages.success(request, 'Perfil salvo! Token de Empresa detectado (Sincronização cadastral ignorada).')
                            # Finge que deu certo para não bloquear o fluxo
                            response.status_code = 200 
                        
                        # --- RESULTADO FINAL ---
                        print(f"Status Final: {response.status_code}")
                        print(f"Resposta: {response.text}")

                        if response.status_code in [200, 201, 204]:
                            if response.status_code != 404: # Só mostra msg padrão se não foi o bypass acima
                                messages.success(request, 'Empresa configurada e sincronizada com sucesso na Focus!')

                            # --- LÓGICA FORÇADA DE UPLOAD DO CERTIFICADO ---
                            # O usuário relatou que o envio junto com o POST falha silenciosamente.
                            # Portanto, SEMPRE faremos o upload separado para garantir.
                            
                            precisa_enviar_separado = False
                            if conteudo_certificado and senha_cert:
                                # FORÇAR ENVIO SEPARADO SEMPRE
                                precisa_enviar_separado = True
                                print("⚠️ Forçando envio separado do certificado para garantir vínculo.")

                            if precisa_enviar_separado:
                                try:
                                    print(f"--- TENTATIVA 3: UPLOAD CERTIFICADO (SEPARADO) ---")
                                    
                                    # PAUSA PARA PROPAGAÇÃO (CORREÇÃO DO 404)
                                    print("⏳ Aguardando 2s para propagação no banco da Focus...")
                                    time.sleep(2)

                                    import base64
                                    arquivo_base64 = base64.b64encode(conteudo_certificado).decode('utf-8')
                                    
                                    # --- CORREÇÃO FINAL: PAYLOAD EXCLUSIVO + CNPJ ---
                                    # Enviar APENAS o certificado e o CNPJ (para validação cruzada)
                                    # TENTATIVA COM CHAVES PADRÃO: 'certificado' e 'senha_certificado'
                                    print(f"Tamanho do arquivo Base64: {len(arquivo_base64)} caracteres")
                                    
                                    payload_certificado_exclusivo = {
                                        "cnpj": cnpj_limpo,
                                        "arquivo_certificado_base64": arquivo_base64,
                                        "senha_certificado": senha_cert
                                    }
                                    
                                    # --- CORREÇÃO DO 404: USAR ID SE DISPONÍVEL ---
                                    # Se a empresa acabou de ser criada, o ID já existe e é imediato.
                                    # O CNPJ pode demorar para indexar.
                                    company_id_focus = None
                                    try:
                                        resp_json_create = response.json()
                                        company_id_focus = resp_json_create.get('id')
                                    except:
                                        pass

                                    if company_id_focus:
                                        print(f"✅ Usando ID da Focus para upload: {company_id_focus}")
                                        url_cert = f"{url_base_empresas}/{company_id_focus}"
                                    else:
                                        print(f"⚠️ ID não encontrado, usando CNPJ: {cnpj_limpo}")
                                        url_cert = f"{url_base_empresas}/{cnpj_limpo}"
                                    
                                    print(f"--- ENVIANDO PUT CERTIFICADO ---")
                                    resp_cert = requests.put(url_cert, json=payload_certificado_exclusivo, auth=(API_TOKEN, ""))
                                    print(f"RESPOSTA CERTIFICADO: {resp_cert.status_code} - {resp_cert.text}")
                                    
                                    if resp_cert.status_code in [200, 201, 204]:
                                        messages.success(request, 'Certificado Digital enviado com sucesso!')
                                    elif resp_cert.status_code == 404 and "Endpoint não encontrado" in resp_cert.text:
                                         print("⚠️ Token de empresa não permite upload de certificado via API.")
                                         messages.info(request, 'Certificado salvo localmente. (Upload via API não permitido para este Token).')
                                    else:
                                        print(f"Erro Certificado: {resp_cert.text}")
                                        messages.warning(request, f'Empresa salva, mas erro ao enviar certificado: {resp_cert.text}')
                                except Exception as e_cert:
                                    print(f"Erro Exception Certificado: {e_cert}")
                                    messages.warning(request, f'Empresa salva, mas falha interna no certificado: {str(e_cert)}')
                            # -----------------------------------------------------

                        else:
                            # Tenta ler o erro de forma legível
                            try:
                                resp_json = response.json()
                                msg_erro = resp_json.get('mensagem', str(resp_json))
                                if 'erros' in resp_json:
                                    msg_erro = str(resp_json['erros'])
                            except:
                                msg_erro = response.text
                            
                            messages.error(request, f'Erro na integração Focus: {msg_erro}')

                    except Exception as e:
                        messages.warning(request, f'Perfil salvo localmente, mas houve erro na integração: {str(e)}')
                else:
                    messages.success(request, 'Perfil salvo localmente.')

                return redirect('company_profile')

        # ... (resto do código de documentos/anexar continua igual) ...
        elif 'upload_document' in request.POST:
             # ... (código existente)
             pass
        elif 'delete_document' in request.POST:
             # ... (código existente)
             pass

    # else:
    #     # Não é necessário inicializar aqui pois já foi feito no início
    #     pass

    context = {
        'profile_form': profile_form,
        'doc_form': doc_form,
        'documents': documents,
    }
    return render(request, 'accounts/company_profile.html', context)

@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_painel_financeiro')
def dre_view(request):
    # --- Lógica para capturar o regime de apuração ---
    regime = request.GET.get('regime', 'caixa')

    # --- Lógica de Filtros de Período ---
    period = request.GET.get('period', '90')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        end_date = datetime.today().date()
        days_map = {'30': 30, '90': 90, '180': 180}
        start_date = end_date - relativedelta(days=days_map.get(period, 90))

    # --- Lógica para Período Anterior e Rótulos (sem alterações) ---
    period_duration = end_date - start_date
    end_date_anterior = start_date - timedelta(days=1)
    start_date_anterior = end_date_anterior - period_duration
    month_labels = []
    current_month = start_date.replace(day=1)
    while current_month <= end_date:
        month_labels.append(current_month.strftime('%b/%Y'))
        current_month += relativedelta(months=1)

    # --- Estrutura da DRE (sem alterações) ---
    dre_structure = OrderedDict([
        ('RECEITAS_BRUTAS', {'label': '(=) RECEITAS BRUTAS', 'is_total': False}),
        ('DEDUCAO_RECEITA_BRUTA', {'label': '(-) DEDUÇÃO DA RECEITA BRUTA', 'is_total': False}),
        ('RECEITAS_LIQUIDAS', {'label': '(=) RECEITAS LÍQUIDAS', 'is_total': True}),
        ('CUSTOS_CSP_CMV', {'label': '(-) CUSTOS - CSP/CMV', 'is_total': False}),
        ('LUCRO_BRUTO_MC', {'label': '(=) LUCRO BRUTO / MC', 'is_total': True}),
        ('DESPESAS_OPERACIONAIS', {'label': '(-) DESPESAS OPERACIONAIS (SG&A)', 'is_total': False}),
        ('EBITDA', {'label': '(=) Lucro Liquido Operacional (EBITDA)', 'is_total': True}),
        ('DEPRECIACAO', {'label': '(-) DEPRECIAÇÃO E AMORTIZAÇÃO', 'is_total': False}), # <-- NOVA LINHA
        ('EBIT', {'label': '(=) EBIT (Lucro Antes dos Juros e Impostos)', 'is_total': True}), # <-- NOVA LINHA
        ('NAO_OPERACIONAIS', {'label': '(+/-) RESULTADO NÃO OPERACIONAL', 'is_total': True}),
        ('LUCRO_ANTES_IR', {'label': '(=) LUCRO ANTES DO IMPOSTO DE RENDA (LAIR)', 'is_total': True}), # <-- NOVA LINHA
        ('TRIBUTACAO', {'label': '(-) IRPJ E CSLL (TRIBUTAÇÃO)', 'is_total': False}), # <-- NOVA LINHA
        ('LUCRO_LIQUIDO_FINAL', {'label': '(=) LUCRO LÍQUIDO DO EXERCÍCIO', 'is_total': True}),
        ('DISTRIBUICAO_LUCRO_SOCIOS', {'label': '(-) DISTRIBUIÇÃO DE LUCRO SÓCIOS', 'is_total': False}),
        ('APORTES_RETIRADAS', {'label': '(+/-) Aportes / Retiradas dos Sócios', 'is_total': False}),
        ('RESULTADO_FINAL', {'label': '(=) RESULTADO FINAL', 'is_total': True}),
    ])

    # --- Lógica de Apuração Condicional (Caixa vs. Competência) ---
    monthly_totals = {key: [Decimal('0.0')] * len(month_labels) for key in dre_structure.keys()}
    monthly_details = {key: [[] for _ in range(len(month_labels))] for key in dre_structure.keys()}
    
    # PERÍODO ATUAL
    receivables_query = ReceivableAccount.objects.filter(
        user=request.user, 
        dre_area__in=['BRUTA', 'OUTRAS_RECEITAS', 'APORTE_SOCIOS'], # <--- Agora aceita os dois
        due_date__range=[start_date, end_date]
    )
    payables_query = PayableAccount.objects.filter(user=request.user, due_date__range=[start_date, end_date]).exclude(dre_area='NAO_CONSTAR')
    if regime == 'caixa':
        receivables_query = receivables_query.filter(is_received=True)
        payables_query = payables_query.filter(is_paid=True)

    # PERÍODO ANTERIOR
    receivables_anterior_query = ReceivableAccount.objects.filter(
        user=request.user, 
        dre_area__in=['BRUTA', 'OUTRAS_RECEITAS', 'APORTE_SOCIOS'], # <--- Agora busca vendas e aportes
        due_date__range=[start_date_anterior, end_date_anterior]
    )
    payables_anterior_query = PayableAccount.objects.filter(user=request.user, due_date__range=[start_date_anterior, end_date_anterior]).exclude(dre_area='NAO_CONSTAR')
    if regime == 'caixa':
        receivables_anterior_query = receivables_anterior_query.filter(is_received=True)
        payables_anterior_query = payables_anterior_query.filter(is_paid=True)

    # Em dre_view, cole este novo bloco de código

    # --- NOVO: Busca detalhes por nome para expansão na tabela (CORRIGIDO) ---
    details_by_name = {key: [] for key in dre_structure.keys()}

    # 1. Detalhes de Contas a Receber (Separando Bruta, Aporte e Outras)
    # Adicionamos 'dre_area' no values para poder filtrar
    receivables_details_by_name = receivables_query.values('dre_area', 'name').annotate(
        total=Sum('amount')
    ).order_by('-total')

    for item in receivables_details_by_name:
        if item['dre_area'] == 'BRUTA':
            details_by_name['RECEITAS_BRUTAS'].append({'name': item['name'], 'total': item['total']})
        
        elif item['dre_area'] == 'APORTE_SOCIOS':
            # Joga para a linha correta de Aportes
            details_by_name['APORTES_RETIRADAS'].append({'name': f"(+) {item['name']}", 'total': item['total']})
            
        elif item['dre_area'] == 'OUTRAS_RECEITAS':
            # Outras receitas abatem as despesas não operacionais
            details_by_name['NAO_OPERACIONAIS'].append({'name': f"(+) {item['name']}", 'total': item['total']})

    # 2. Detalhes de Contas a Pagar (Incluindo Retirada de Sócios)
    payables_details_by_name = payables_query.values('dre_area', 'name').annotate(
        total=Sum('amount')
    ).order_by('dre_area', '-total')

    dre_area_mapping = {
        'DEDUCAO': 'DEDUCAO_RECEITA_BRUTA', 
        'CUSTOS': 'CUSTOS_CSP_CMV', 
        'OPERACIONAL': 'DESPESAS_OPERACIONAIS', 
        'NAO_OPERACIONAL': 'NAO_OPERACIONAIS', 
        'DISTRIBUICAO': 'DISTRIBUICAO_LUCRO_SOCIOS', 
        'DEPRECIACAO': 'DEPRECIACAO', 
        'TRIBUTACAO': 'TRIBUTACAO',
        # ADICIONADO: Mapeamento para Retirada de Sócios aparecer na lista
        'RETIRADA_SOCIOS': 'APORTES_RETIRADAS' 
    }
    
    for item in payables_details_by_name:
        dre_key = dre_area_mapping.get(item['dre_area'])
        if dre_key:
            # Se for retirada, mostramos visualmente com (-)
            prefix = "(-) " if item['dre_area'] == 'RETIRADA_SOCIOS' else ""
            details_by_name[dre_key].append({'name': f"{prefix}{item['name']}", 'total': item['total']})
    # --- FIM DO NOVO BLOCO ---    

    # --- Cálculos e Processamento (sem alterações na lógica interna) ---
    # ... (O restante da função, com os loops e cálculos, permanece o mesmo da versão que já te passei) ...
    # ... Para garantir, aqui está a continuação completa e correta ...
    
    # ADICIONEI 'dre_area' DENTRO DO values(...)
    receivables_details = receivables_query.annotate(month=TruncMonth('due_date')).values('month', 'category__name', 'dre_area').annotate(total=Sum('amount'))
    # ... logo após a linha: receivables_details = receivables_query.annotate(...) ...

    for entry in receivables_details:
        try:
            month_str = entry['month'].strftime('%b/%Y')
            idx = month_labels.index(month_str)
            
            # Pega o nome da categoria ou define padrão
            category_name = entry['category__name'] or "Sem Categoria"
            
            # Garante que o valor é decimal
            amount = Decimal(str(entry['total'])) if entry['total'] else Decimal('0')

            # --- AQUI ESTÁ A LÓGICA INTELIGENTE ---
            
            # CASO 1: Se for Venda Normal (Receita Bruta)
            if entry['dre_area'] == 'BRUTA':
                monthly_totals['RECEITAS_BRUTAS'][idx] += amount
                monthly_details['RECEITAS_BRUTAS'][idx].append({
                    'category': category_name, 
                    'amount': amount
                })

            # CASO 2: Se for Aporte/Investimento (Outras Receitas)
            elif entry['dre_area'] == 'OUTRAS_RECEITAS':
                # Agora tratamos como Resultado: Receita aumenta o saldo
                monthly_totals['NAO_OPERACIONAIS'][idx] += amount
                
                monthly_details['NAO_OPERACIONAIS'][idx].append({
                    'category': f"(+) {category_name}", 
                    'amount': amount
                })

            elif entry['dre_area'] == 'APORTE_SOCIOS':
                # Soma como positivo na nova linha
                monthly_totals['APORTES_RETIRADAS'][idx] += amount
                monthly_details['APORTES_RETIRADAS'][idx].append({
                    'category': f"(+) {category_name}", # Sinal de + visual
                    'amount': amount
                })    

        except (ValueError, IndexError, TypeError): 
            continue

    payables_details = payables_query.annotate(month=TruncMonth('due_date')).values('month', 'dre_area', 'category__name').annotate(total=Sum('amount'))
    dre_area_mapping = {'DEDUCAO': 'DEDUCAO_RECEITA_BRUTA', 'CUSTOS': 'CUSTOS_CSP_CMV', 'OPERACIONAL': 'DESPESAS_OPERACIONAIS', 'NAO_OPERACIONAL': 'NAO_OPERACIONAIS', 'RETIRADA_SOCIOS': 'APORTES_RETIRADAS', 'DISTRIBUICAO': 'DISTRIBUICAO_LUCRO_SOCIOS'}
    for entry in payables_details:
        dre_area_key = dre_area_mapping.get(entry['dre_area'], entry['dre_area'])
        if dre_area_key in monthly_totals:
            try:
                month_str = entry['month'].strftime('%b/%Y'); idx = month_labels.index(month_str)
                category_name = entry['category__name'] or "Sem Categoria"; amount = Decimal(str(entry['total'])) if entry['total'] else Decimal('0')
                
                # --- LÓGICA CORRIGIDA AQUI ---
                if dre_area_key == 'APORTES_RETIRADAS':
                    # Se for retirada (Pagamento), SUBTRAI do saldo de Aportes/Retiradas
                    monthly_totals[dre_area_key][idx] -= amount
                    monthly_details[dre_area_key][idx].append({'category': f"(-) {category_name}", 'amount': -amount})
                
                # --- NOVO BLOCO: Trata Despesa Não Operacional como redução do saldo ---
                elif dre_area_key == 'NAO_OPERACIONAIS':
                    monthly_totals[dre_area_key][idx] -= amount
                    # Adiciona o prefixo (-) no nome e inverte o valor para aparecer negativo no detalhe
                    monthly_details[dre_area_key][idx].append({'category': f"(-) {category_name}", 'amount': -amount})
                # -----------------------------------------------------------------------

                else:
                    # Demais despesas continuam somando no acumulador da despesa (ex: Custos, Despesas Op)
                    monthly_totals[dre_area_key][idx] += amount
                    monthly_details[dre_area_key][idx].append({'category': category_name, 'amount': amount})
                # -----------------------------

            except (ValueError, IndexError, TypeError): continue

    # Em views.py, na função dre_view

    for i in range(len(month_labels)):
        # Cálculos que não mudam
        monthly_totals['RECEITAS_LIQUIDAS'][i] = monthly_totals['RECEITAS_BRUTAS'][i] - monthly_totals['DEDUCAO_RECEITA_BRUTA'][i]
        monthly_totals['LUCRO_BRUTO_MC'][i] = monthly_totals['RECEITAS_LIQUIDAS'][i] - monthly_totals['CUSTOS_CSP_CMV'][i]
        # ALTERE A LINHA ACIMA PARA ISTO:
        monthly_totals['EBITDA'][i] = monthly_totals['LUCRO_BRUTO_MC'][i] - monthly_totals['DESPESAS_OPERACIONAIS'][i]
        
        monthly_totals['EBIT'][i] = monthly_totals['EBITDA'][i] - monthly_totals['DEPRECIACAO'][i]
        # LAIR = EBIT + Resultado Não Operacional (pois agora calculamos o saldo líquido)
        monthly_totals['LUCRO_ANTES_IR'][i] = monthly_totals['EBIT'][i] + monthly_totals['NAO_OPERACIONAIS'][i]
        # Lucro Líquido = LAIR - Impostos sobre o Lucro
        monthly_totals['LUCRO_LIQUIDO_FINAL'][i] = monthly_totals['LUCRO_ANTES_IR'][i] - monthly_totals['TRIBUTACAO'][i]
        # Resultado Final = Lucro Líquido - Distribuição
        monthly_totals['RESULTADO_FINAL'][i] = monthly_totals['LUCRO_LIQUIDO_FINAL'][i] - monthly_totals['DISTRIBUICAO_LUCRO_SOCIOS'][i] + monthly_totals['APORTES_RETIRADAS'][i]

    monthly_totals_anterior = {key: Decimal('0.0') for key in dre_structure.keys()}
    
    # --- CORREÇÃO: Separa Receita Bruta de Aportes no Período Anterior ---
    rec_ant_bruta = receivables_anterior_query.filter(dre_area='BRUTA').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    rec_ant_aporte = receivables_anterior_query.filter(dre_area='APORTE_SOCIOS').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    rec_ant_outras = receivables_anterior_query.filter(dre_area='OUTRAS_RECEITAS').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    monthly_totals_anterior['RECEITAS_BRUTAS'] = rec_ant_bruta
    monthly_totals_anterior['APORTES_RETIRADAS'] += rec_ant_aporte
    monthly_totals_anterior['NAO_OPERACIONAIS'] += rec_ant_outras # Receita SOMA no resultado não operacional
    # ---------------------------------------------------------------------

    payables_anterior = payables_anterior_query.values('dre_area').annotate(total=Sum('amount'))
    for entry in payables_anterior:
        dre_area_key = dre_area_mapping.get(entry['dre_area'], entry['dre_area'])
        if dre_area_key in monthly_totals_anterior and entry['total'] is not None:
            # CORREÇÃO: Se for retirada, subtrai. Se for despesa, soma.
            if dre_area_key == 'APORTES_RETIRADAS':
                monthly_totals_anterior[dre_area_key] -= entry['total']
            elif dre_area_key == 'NAO_OPERACIONAIS':
                monthly_totals_anterior[dre_area_key] -= entry['total'] # Despesa SUBTRAI do resultado
            else:
                monthly_totals_anterior[dre_area_key] += entry['total']
            
    monthly_totals_anterior['RECEITAS_LIQUIDAS'] = monthly_totals_anterior['RECEITAS_BRUTAS'] - monthly_totals_anterior['DEDUCAO_RECEITA_BRUTA']
    monthly_totals_anterior['LUCRO_BRUTO_MC'] = monthly_totals_anterior['RECEITAS_LIQUIDAS'] - monthly_totals_anterior['CUSTOS_CSP_CMV']
    # ALTERE A LINHA ACIMA PARA ISTO:
    monthly_totals_anterior['EBITDA'] = monthly_totals_anterior['LUCRO_BRUTO_MC'] - monthly_totals_anterior['DESPESAS_OPERACIONAIS']

    # ▼▼▼ NOVOS CÁLCULOS PARA A DRE COMPLETA ▼▼▼
    # ALTERE AQUI TAMBÉM
    monthly_totals_anterior['EBIT'] = monthly_totals_anterior['EBITDA'] - monthly_totals_anterior['DEPRECIACAO']
    monthly_totals_anterior['LUCRO_ANTES_IR'] = monthly_totals_anterior['EBIT'] + monthly_totals_anterior['NAO_OPERACIONAIS']
    monthly_totals_anterior['LUCRO_LIQUIDO_FINAL'] = monthly_totals_anterior['LUCRO_ANTES_IR'] - monthly_totals_anterior['TRIBUTACAO']
    monthly_totals_anterior['RESULTADO_FINAL'] = monthly_totals_anterior['LUCRO_LIQUIDO_FINAL'] - monthly_totals_anterior['DISTRIBUICAO_LUCRO_SOCIOS'] + monthly_totals_anterior['APORTES_RETIRADAS']

    dre_data_anterior_for_template = []
    for k, v in dre_structure.items():
        dre_data_anterior_for_template.append({ 'key': k, 'label': v['label'], 'is_total': v['is_total'], 'total_geral': monthly_totals_anterior[k], 'monthly_data': [{'total': monthly_totals_anterior[k], 'details': []}]})

    # COLE ESTE NOVO BLOCO COMPLETO NO LUGAR DO ANTIGO
    total_receita_bruta_periodo = sum(monthly_totals['RECEITAS_BRUTAS'])
    total_receita_liquida_periodo = sum(monthly_totals['RECEITAS_LIQUIDAS'])
    dre_data_for_template = []

    for key, values in dre_structure.items():
        total_geral_atual = sum(monthly_totals[key])
        total_geral_anterior = monthly_totals_anterior[key]

        # Análise Vertical do Período Total (coluna A.V. principal)
        base_calculo_av_total = total_receita_bruta_periodo if key in ['RECEITAS_BRUTAS', 'DEDUCAO_RECEITA_BRUTA'] else total_receita_liquida_periodo
        analise_vertical_total = (total_geral_atual / base_calculo_av_total * 100) if base_calculo_av_total > 0 else 0
        
        # Análise Horizontal do Período Total (coluna A.H. principal)
        analise_horizontal_total = None
        if total_geral_anterior is not None:
            if total_geral_anterior != 0:
                analise_horizontal_total = ((total_geral_atual - total_geral_anterior) / abs(total_geral_anterior)) * 100
            else:
                analise_horizontal_total = 100.0 if total_geral_atual != 0 else 0.0

        # --- NOVA LÓGICA: CÁLCULOS MÊS A MÊS ---
        monthly_data_combined = []
        for i in range(len(month_labels)):
            # Cálculo da Análise Vertical MENSAL
            base_bruta_mes = monthly_totals['RECEITAS_BRUTAS'][i]
            base_liquida_mes = monthly_totals['RECEITAS_LIQUIDAS'][i]
            base_av_mes = base_bruta_mes if key in ['RECEITAS_BRUTAS', 'DEDUCAO_RECEITA_BRUTA'] else base_liquida_mes
            av_mes = (monthly_totals[key][i] / base_av_mes * 100) if base_av_mes > 0 else 0

            # Cálculo da Análise Horizontal MENSAL (comparando com o mês anterior)
            ah_mes = None
            if i > 0: # Só calcula a partir do segundo mês do período
                valor_atual_mes = monthly_totals[key][i]
                valor_anterior_mes = monthly_totals[key][i-1]
                if valor_anterior_mes != 0:
                    ah_mes = ((valor_atual_mes - valor_anterior_mes) / abs(valor_anterior_mes)) * 100
                else:
                    ah_mes = 100.0 if valor_atual_mes != 0 else 0.0

            monthly_data_combined.append({
                'total': monthly_totals[key][i],
                'details': monthly_details[key][i],
                'av': av_mes,
                'ah': ah_mes
            })
        
        dre_data_for_template.append({ 
            'key': key, 
            'label': values['label'], 
            'is_total': values['is_total'], 
            'monthly_data': monthly_data_combined, 
            'total_geral': total_geral_atual, 
            'av': analise_vertical_total, 
            'ah': analise_horizontal_total, 
            'details_by_name': details_by_name.get(key, [])
        })

    # COLE ESTE BLOCO ANTES DA DEFINIÇÃO DO 'context' NA dre_view

    # --- INÍCIO: LÓGICA PARA ENVIAR OS TOTAIS ATUALIZADOS PARA OS CARDS ---
    dre_cards_data = {}
    key_map = {
        'RECEITAS_BRUTAS': 'receita_bruta',
        'DEDUCAO_RECEITA_BRUTA': 'impostos',
        'LUCRO_BRUTO_MC': 'lucro_bruto',
        'EBITDA': 'ebitda',
        'LUCRO_LIQUIDO_FINAL': 'lucro_liquido',
    }
    # Pega os totais já calculados para o período
    for item in dre_data_for_template:
        if item['key'] in key_map:
            card_key = key_map[item['key']]
            dre_cards_data[card_key] = item['total_geral']

    # Calcula o Ponto de Equilíbrio para o período (respeitando o regime)
    total_receita_liquida_cards = dre_cards_data.get('receita_bruta', Decimal('0')) - dre_cards_data.get('impostos', Decimal('0'))
    
    # Reutiliza a 'payables_query' que já está filtrada corretamente pelo regime
    total_custos_variaveis_cards = payables_query.filter(dre_area='CUSTOS', cost_type='VARIAVEL').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    total_custos_fixos_cards = payables_query.filter(cost_type='FIXO').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    total_margem_contribuicao_cards = total_receita_liquida_cards - total_custos_variaveis_cards
    
    ponto_equilibrio_cards = Decimal('0')
    if total_margem_contribuicao_cards > 0 and total_receita_liquida_cards > 0:
        mc_percentual = total_margem_contribuicao_cards / total_receita_liquida_cards
        if mc_percentual > 0:
            ponto_equilibrio_cards = total_custos_fixos_cards / mc_percentual
            
    dre_cards_data['ponto_equilibrio'] = ponto_equilibrio_cards
    # --- FIM DA LÓGICA DOS CARDS ---    

    

    context = {
        'period': period, 'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '', 'dre_data': dre_data_for_template,
        'dre_data_anterior': dre_data_anterior_for_template, 'month_labels': month_labels, 'regime': regime,
        'dre_cards_data': dre_cards_data,
    }
    return render(request, 'accounts/dre.html', context)

# Em seu arquivo views.py, substitua a função fornecedor_view existente por esta:

# Em seu arquivo views.py, substitua a função fornecedor_view por esta:

@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_fornecedores')
def fornecedor_view(request):
    # --- INÍCIO: LÓGICA DE EXPORTAÇÃO PARA EXCEL ---
    if 'export_excel' in request.GET:
        period = request.GET.get('period', '30')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        end_date = datetime.today().date()

        try:
            if period == 'custom':
                start_date = parse_date(start_date_str) if start_date_str else None
                end_date = parse_date(end_date_str) if end_date_str else datetime.today().date()
            else:
                start_date = end_date - timedelta(days=int(period))
        except (ValueError, TypeError):
            start_date = end_date - timedelta(days=30)

        accounts_to_export = PayableAccount.objects.filter(
            user=request.user, due_date__range=[start_date, end_date]
        ).order_by('due_date')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="contas_a_pagar_fornecedores.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contas a Pagar"
        headers = ['Nome', 'Descrição', 'Data de Vencimento', 'Valor (R$)', 'Status']
        ws.append(headers)
        for account in accounts_to_export:
            ws.append([
                account.name, account.description, account.due_date,
                account.amount, "Pago" if account.is_paid else "Aberto"
            ])
        wb.save(response)
        return response
    # --- FIM: LÓGICA DE EXPORTAÇÃO ---

    period = request.GET.get('period', '30')
    end_date = datetime.today().date()

    try:
        if period == 'custom':
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')
            start_date = parse_date(start_date_str) if start_date_str else end_date - timedelta(days=30)
            end_date = parse_date(end_date_str) if end_date_str else datetime.today().date()
        else:
            days = int(period)
            start_date = end_date - timedelta(days=days)
    except (ValueError, TypeError):
        period = '30'
        start_date = end_date - timedelta(days=30)

    # Cálculos para os Cards (usando todo o histórico)
    all_accounts = PayableAccount.objects.filter(user=request.user)
    today = datetime.today().date()
    total_suppliers = all_accounts.values('name').distinct().count()
    total_paid = all_accounts.filter(is_paid=True).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_overdue = all_accounts.filter(is_paid=False, due_date__lt=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_open = all_accounts.filter(is_paid=False).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # Filtrar contas para a TABELA e GRÁFICOS
    accounts_filtered = PayableAccount.objects.filter(
        user=request.user,
        due_date__range=[start_date, end_date]
    )

    accounts_query = accounts_filtered.order_by('due_date')
    total_amount_period = accounts_query.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # --- INÍCIO: LÓGICA DA PAGINAÇÃO ---
    paginator = Paginator(accounts_query, 10) # 10 itens por página
    page_number = request.GET.get('page')
    accounts_page_obj = paginator.get_page(page_number)
    # --- FIM: LÓGICA DA PAGINAÇÃO ---

    # Dados para os Gráficos (restante da lógica sem alteração)
    category_data_q = accounts_filtered.filter(is_paid=True).values('category__name').annotate(total=Sum('amount')).order_by('-total')
    category_labels = [item['category__name'] or 'Sem Categoria' for item in category_data_q]
    category_values = [float(item['total']) for item in category_data_q]

    

    # Prepara os dados para a nova tabela de Despesas por Fornecedor
    despesas_por_fornecedor = accounts_filtered.filter(is_paid=True).values('name').annotate(
        total=Sum('amount')
    ).order_by('-total')

    # Calcula o total para o rodapé da nova tabela
    total_despesas_fornecedor_periodo = despesas_por_fornecedor.aggregate(
        total_geral=Sum('total')
    )['total_geral'] or Decimal('0.00')

    PAYMENT_METHOD_CHOICES = dict(PayableAccount._meta.get_field('payment_method').choices)
    payment_method_data_q = accounts_filtered.filter(is_paid=True).values('payment_method').annotate(total=Sum('amount')).order_by('-total')
    payment_method_labels = [PAYMENT_METHOD_CHOICES.get(item['payment_method'], item['payment_method']) for item in payment_method_data_q]
    payment_method_values = [float(item['total']) for item in payment_method_data_q]

    status_data = [
        float(accounts_filtered.filter(is_paid=False).aggregate(total=Sum('amount'))['total'] or 0),
        float(accounts_filtered.filter(is_paid=True).aggregate(total=Sum('amount'))['total'] or 0)
    ]

    fixed_cost_data_q = accounts_filtered.filter(is_paid=True, cost_type='FIXO').values('category__name').annotate(total=Sum('amount')).order_by('-total')
    fixed_cost_labels = [item['category__name'] or 'Sem Categoria' for item in fixed_cost_data_q]
    fixed_cost_values = [float(item['total']) for item in fixed_cost_data_q]

    variable_cost_data_q = accounts_filtered.filter(is_paid=True, cost_type='VARIAVEL').values('category__name').annotate(total=Sum('amount')).order_by('-total')
    variable_cost_labels = [item['category__name'] or 'Sem Categoria' for item in variable_cost_data_q]
    variable_cost_values = [float(item['total']) for item in variable_cost_data_q]

    context = {
        'accounts': accounts_page_obj, # <-- MODIFICADO para usar o objeto da paginação
        'period': period,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        'total_suppliers': total_suppliers,
        'total_paid': total_paid,
        'total_overdue': total_overdue,
        'total_open': total_open,
        'category_labels': category_labels,
        'category_data': category_values,
        'despesas_por_fornecedor': despesas_por_fornecedor,
        'total_despesas_fornecedor_periodo': total_despesas_fornecedor_periodo,
        'payment_method_labels': payment_method_labels,
        'payment_method_data': payment_method_values,
        'status_data': status_data,
        'fixed_cost_labels': fixed_cost_labels,
        'fixed_cost_data': fixed_cost_values,
        'variable_cost_labels': variable_cost_labels,
        'variable_cost_data': variable_cost_values,
        'total_amount_period': total_amount_period,
    }
    return render(request, 'accounts/fornecedor.html', context)



# Em seu arquivo views.py, substitua a função clientes por esta:

@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_clientes_financeiro')
def clientes(request):

    # ▼▼▼ COLOQUE TODO O BLOCO DE EXPORTAÇÃO AQUI ▼▼▼
    if 'export_excel' in request.GET:
        period = request.GET.get('period', '30')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        try:
            end_date = datetime.today().date()
            if period == 'custom':
                start_date = parse_date(start_date_str) if start_date_str else None
                end_date = parse_date(end_date_str) if end_date_str else datetime.today().date()
            else:
                start_date = end_date - timedelta(days=int(period))
        except (ValueError, TypeError):
            start_date = datetime.today().date() - timedelta(days=30)

        accounts_to_export = ReceivableAccount.objects.filter(
            user=request.user, due_date__range=[start_date, end_date]
        ).order_by('due_date')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="contas_a_receber.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contas a Receber"
        headers = ['Nome', 'Descrição', 'Vencimento', 'Status', 'Valor (R$)']
        ws.append(headers)
        for account in accounts_to_export:
            ws.append([
                account.name, account.description, account.due_date,
                "Recebido" if account.is_received else "Aberto", account.amount
            ])
        wb.save(response)
        return response
    # --- FIM DO BLOCO DE EXPORTAÇÃO ---

    period = request.GET.get('period', '30')
    end_date = datetime.today().date()
    
    try:
        if period == 'custom':
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')
            start_date = parse_date(start_date_str) if start_date_str else None
            end_date = parse_date(end_date_str) if end_date_str else datetime.today().date()
            if start_date and end_date:
                if start_date > end_date:
                    start_date, end_date = end_date, start_date
            else:
                period = '30'
                start_date = end_date - timedelta(days=30)
        else:
            period = str(period)
            period_value = int(period)
            start_date = end_date - timedelta(days=period_value)
    except (ValueError, TypeError):
        period = '30'
        start_date = end_date - timedelta(days=30)

    

    # --- INÍCIO: Cálculos para os Cards (usando todo o histórico de Contas a Receber) ---
    all_accounts = ReceivableAccount.objects.filter(user=request.user)
    today = datetime.today().date()
    
    total_clients = all_accounts.values('name').distinct().count()
    total_received = all_accounts.filter(is_received=True).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_overdue = all_accounts.filter(is_received=False, due_date__lt=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_open = all_accounts.filter(is_received=False).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    # --- INÍCIO: Cálculo da Taxa de Inadimplência (PADRONIZADO) ---
    # O 'all_accounts' já contém todas as contas a receber do usuário
    total_faturado = all_accounts.aggregate(total=Sum('amount'))['total'] or Decimal('1')
    taxa_inadimplencia = (total_overdue / total_faturado * 100) if total_faturado > 0 else Decimal('0.00')
    # --- FIM: Cálculo da Taxa de Inadimplência (PADRONIZADO) ---

    # Filtra as contas a receber para o período selecionado (esta linha já existe, mantenha)
    accounts = ReceivableAccount.objects.filter(
        user=request.user,
        due_date__range=[start_date, end_date]
    )
    # vvv ADICIONE ESTE BLOCO ABAIXO vvv
    # --- INÍCIO: Cálculos dos totais para os rodapés das tabelas ---

    # Total para a tabela "Contas a Receber no Período" (soma de todos os valores no período)
    total_contas_receber_periodo = accounts.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    # Totais para a tabela "Previsão x Realizado"
    # O total previsto é o mesmo do cálculo anterior
    total_previsto = total_contas_receber_periodo
    # O total realizado é a soma apenas das contas recebidas no período
    total_realizado = accounts.filter(is_received=True).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    # --- FIM: Cálculos dos totais ---

    # --- INÍCIO DA ALTERAÇÃO: Dados para a nova tabela "Previsão x Realizado por Cliente" ---
    client_comparison_data = accounts.values('name').annotate(
        previsto=Sum('amount'),
        realizado=Sum('amount', filter=Q(is_received=True))
    ).order_by('-previsto')
    # --- FIM DA ALTERAÇÃO ---

    # --- INÍCIO DA CORREÇÃO: Preparação dos dados dos gráficos SEM json.dumps ---
    category_data_q = accounts.values('category__name').annotate(total=Sum('amount')).order_by('-total')
    category_labels = [item['category__name'] or 'Sem Categoria' for item in category_data_q]
    category_data = [float(item['total']) for item in category_data_q]

    # Mantemos a query original, mas agora vamos passá-la diretamente para o template
    receitas_por_cliente = accounts.values('name').annotate(total=Sum('amount')).order_by('-total')

    payment_method_data_q = accounts.values('payment_method').annotate(total=Sum('amount')).order_by('-total')
    payment_method_labels = [dict(PAYMENT_METHODS).get(item['payment_method'], item['payment_method']) for item in payment_method_data_q]
    payment_method_data = [float(item['total']) for item in payment_method_data_q]

    status_data = [
        float(accounts.filter(is_received=False).aggregate(total=Sum('amount'))['total'] or 0),
        float(accounts.filter(is_received=True).aggregate(total=Sum('amount'))['total'] or 0)
    ]

    occurrence_data = [
        float(accounts.filter(occurrence='AVULSO').aggregate(total=Sum('amount'))['total'] or 0),
        float(accounts.filter(occurrence='RECORRENTE').aggregate(total=Sum('amount'))['total'] or 0)
    ]
    accounts_query = accounts.order_by('due_date')
    # --- FIM DA CORREÇÃO ---
    paginator = Paginator(accounts_query, 10)
    page_number = request.GET.get('page')
    accounts_page_obj = paginator.get_page(page_number)
    
    context = {
        'accounts': accounts_page_obj,
        'period': period,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        
        # Dados dos gráficos (corrigidos)
        'category_labels': category_labels,
        'category_data': category_data,
        'receitas_por_cliente': receitas_por_cliente,
        'payment_method_labels': payment_method_labels,
        'payment_method_data': payment_method_data,
        'status_data': status_data,
        'occurrence_data': occurrence_data,
        # --- INÍCIO: Adicionar variáveis dos cards ---
        'total_clients': total_clients,
        'total_received': total_received,
        'total_overdue': total_overdue,
        'total_open': total_open,
        'taxa_inadimplencia': taxa_inadimplencia,
        # --- FIM: Adicionar variáveis dos cards ---
        
        # Nova variável para a tabela "Previsão x Realizado"
        'client_comparison_data': client_comparison_data,
        'total_contas_receber_periodo': total_contas_receber_periodo,
        'total_previsto': total_previsto,
        'total_realizado': total_realizado,
    }

    return render(request, 'accounts/clientes.html', context)


import logging
logger = logging.getLogger(__name__)





def login_view(request):
    # Obtém o parâmetro 'next' da URL (GET) ou do formulário (POST)
    redirect_to = request.POST.get('next', request.GET.get('next', ''))

    if request.method == 'POST':
        username = request.POST.get('username') # Use .get() para evitar erros
        password = request.POST.get('password')
        
        # Validação básica (opcional, mas boa prática)
        if not username or not password:
             return render(request, 'accounts/login.html', {
                 'error': 'Usuário e senha são obrigatórios.', 
                 'next': redirect_to # Mantenha o 'next' no contexto
             })

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            # ▼▼▼ ADICIONE ESTAS 4 LINHAS ABAIXO ▼▼▼
            # Se o usuário for um superusuário, ele não tem 'Subscription'
            # e não pode usar o app normal. Redireciona para o /admin
            if user.is_superuser:
                return redirect('/admin/')
            # --- FIM DAS LINHAS A SEREM ADICIONADAS ---

            # --- INÍCIO DA CORREÇÃO ---
            # Verificação de segurança para evitar redirecionamento para sites maliciosos
            url_is_safe = False
            if redirect_to:
                # Analisa a URL do 'next'
                url_parts = urlparse(redirect_to)
                # Permite apenas URLs relativas (ex: /contaazul/auth/)
                # OU URLs absolutas no seu próprio domínio (sistemclass.com.br)
                # Nota: a verificação de ALLOWED_HOSTS pode precisar de ajustes dependendo da sua config
                if (not url_parts.netloc or url_parts.netloc == request.get_host()):
                      # Garante que começa com / mas não com // (evita protocolo relativo)
                      if redirect_to.startswith('/') and not redirect_to.startswith('//'):
                          url_is_safe = True
                      # Permite o host exato da requisição
                      elif url_parts.netloc == request.get_host():
                          url_is_safe = True


            # --- INÍCIO DA MODIFICAÇÃO BPO ---
            # Verifica se o usuário é um BPO Admin
            is_bpo_admin = False
            try:
                if user.subscription.user_type == 'BPO':
                    is_bpo_admin = True
            except Subscription.DoesNotExist:
                pass # Deixa como Falso, é um usuário normal

            # Limpa qualquer sessão de gerenciamento antiga
            if 'real_user_id' in request.session:
                del request.session['real_user_id']
            if 'managed_user_id' in request.session:
                del request.session['managed_user_id']

            if url_is_safe:
                return redirect(redirect_to)
            elif is_bpo_admin:
                # Se for BPO, redireciona para o dashboard de seleção de clientes
                return redirect('bpo_dashboard')
            else:
                # Se for Cliente Final, redireciona para a home normal
                return redirect('smart_redirect')
            # --- FIM DA MODIFICAÇÃO BPO ---
            # --- FIM DA CORREÇÃO ---
        else:
            # Se a autenticação falhar, renderiza o login novamente com a mensagem de erro
            # e mantém o parâmetro 'next' para a próxima tentativa
            return render(request, 'accounts/login.html', {
                'error': 'Usuário ou senha inválidos', 
                'next': redirect_to
            })

    # Para a requisição GET inicial, apenas renderiza o formulário
    # passando o 'next' para o template
    return render(request, 'accounts/login.html', {'next': redirect_to})




@login_required
@owner_required
@subscription_required
@module_access_required('financial')
def cadastrar_bancos_view(request):
    bank_accounts = BankAccount.objects.filter(user=request.user).order_by('bank_name')
    
    # --- INÍCIO DA LÓGICA DE SALDO DINÂMICO ---
    # (Importe PayableAccount e ReceivableAccount no topo do seu views.py se já não estiverem)
    for account in bank_accounts:
        received = ReceivableAccount.objects.filter(
            user=request.user, bank_account=account, is_received=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        paid = PayableAccount.objects.filter(
            user=request.user, bank_account=account, is_paid=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Anexa o saldo atual calculado ao objeto da conta
        account.current_balance = account.initial_balance + received - paid
    # --- FIM DA LÓGICA DE SALDO DINÂMICO ---

    edit_account = None
    edit_id = request.GET.get('edit')
    if edit_id:
        edit_account = get_object_or_404(BankAccount, id=edit_id, user=request.user)

    if request.method == 'POST':
        if 'create_bank_account' in request.POST:
            bank_name = request.POST.get('bank_name')
            agency = request.POST.get('agency')
            account_number = request.POST.get('account_number')
            initial_balance = request.POST.get('initial_balance')
            opening_date = request.POST.get('opening_date') # 👈 CAMPO ADICIONADO
            
            try:
                initial_balance = Decimal(initial_balance.replace('.', '').replace(',', '.'))
            except (ValueError, TypeError):
                messages.error(request, 'Saldo inicial inválido. Use o formato Ex: 10.589,58.')
                return redirect('cadastrar_bancos')
            
            BankAccount.objects.create(
                user=request.user, 
                bank_name=bank_name, 
                agency=agency,
                account_number=account_number, 
                initial_balance=initial_balance,
                opening_date=opening_date # 👈 CAMPO ADICIONADO
            )
            messages.success(request, 'Conta bancária criada com sucesso.')
            return redirect('cadastrar_bancos')

        elif 'edit_bank_account' in request.POST:
            account_id = request.POST.get('account_id')
            bank_account = get_object_or_404(BankAccount, id=account_id, user=request.user)
            bank_account.bank_name = request.POST.get('bank_name')
            bank_account.agency = request.POST.get('agency')
            bank_account.account_number = request.POST.get('account_number')
            bank_account.opening_date = request.POST.get('opening_date') # 👈 CAMPO ADICIONADO
            
            initial_balance_str = request.POST.get('initial_balance')
            try:
                bank_account.initial_balance = Decimal(str(initial_balance_str).replace('.', '').replace(',', '.'))
            except (ValueError, TypeError):
                messages.error(request, 'Saldo inicial inválido. Use o formato Ex: 10.589,58.')
                return redirect('cadastrar_bancos')
            
            bank_account.save()
            messages.success(request, 'Conta bancária atualizada com sucesso.')
            return redirect('cadastrar_bancos')

        # ▼▼▼ ADICIONE ESTE BLOCO NOVO AQUI ▼▼▼
        elif 'delete_bank_account' in request.POST:
            account_id = request.POST.get('account_id')
            # Busca o banco garantindo que pertence ao usuário logado
            bank_account = get_object_or_404(BankAccount, id=account_id, user=request.user)
            
            # Opcional: Impedir exclusão se tiver lançamentos vinculados
            # if bank_account.payableaccount_set.exists() or bank_account.receivableaccount_set.exists():
            #     messages.error(request, 'Não é possível excluir este banco pois existem contas vinculadas a ele.')
            # else:
            
            bank_account.delete()
            messages.success(request, 'Conta bancária excluída com sucesso.')
            return redirect('cadastrar_bancos')
        # ▲▲▲ FIM DO BLOCO NOVO ▲▲▲    

    context = {
        'bank_accounts': bank_accounts,
        'edit_account': edit_account,
    }
    return render(request, 'accounts/cadastrar_bancos.html', context)

# Em accounts/views.py

@login_required
@owner_required
@subscription_required
@module_access_required('financial')
def importar_ofx_view(request):
    if request.method == 'POST':
        # --- [INÍCIO] COLAR O CÓDIGO AQUI ---
        if 'sync_inter' in request.POST:
            # 1. Define o período (últimos 7 dias)
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=7)
            
            # 2. Chama a API
            resultado = buscar_extrato_inter(request.user, start_date, end_date)
            
            if 'erro' in resultado:
                messages.error(request, f"Erro na integração Inter: {resultado['erro']}")
            else:
                transacoes = resultado.get('transacoes', [])
                if transacoes:
                    print(f"EXEMPLO DE TRANSAÇÃO INTER: {transacoes[0]}")
                count_importados = 0
                
                # Tenta localizar a conta bancária 'Inter' no seu sistema para vincular
                banco_inter = BankAccount.objects.filter(user=request.user, bank_name__icontains='Inter').first()

                # Contadores
                count_importados = 0
                count_duplicados = 0

                # ESTE É O LOOP DA API INTER (Note que usa 'transacao' e chaves como 'tipoOperacao')
                DATE_WINDOW_DAYS = 3 

                for transacao in transacoes:
                    data_str = transacao.get('dataEntrada') or transacao.get('dataLancamento')
                    valor = Decimal(str(transacao.get('valor', 0)))
                    descricao = transacao.get('descricao') or transacao.get('historico') or "Transação API Inter"
                    tipo_operacao = transacao.get('tipoOperacao') 
                    
                    try:
                        data_movimento = datetime.strptime(data_str, '%Y-%m-%d').date()
                    except:
                        data_movimento = end_date

                    descricao_final = f"Importado via API Inter - {descricao}"
                    
                    # [NOVO] Janela de data para busca inteligente
                    date_window_start = data_movimento - timedelta(days=DATE_WINDOW_DAYS)
                    date_window_end = data_movimento + timedelta(days=DATE_WINDOW_DAYS)

                    # --- LÓGICA PAGAR (Débito) ---
                    if tipo_operacao == 'D':
                        # Verifica se JÁ FOI IMPORTADO (evita duplicar a API)
                        existe = PayableAccount.objects.filter(
                            user=request.user,
                            amount=valor,
                            payment_date=data_movimento,
                            description=descricao_final
                        ).exists()

                        if not existe:
                            # [CIRÚRGICO] Tenta achar conta manual aberta antes de criar
                            match = PayableAccount.objects.filter(
                                user=request.user,
                                is_paid=False,          # Tem que estar ABERTA
                                amount=valor,           # Valor exato
                                due_date__range=[date_window_start, date_window_end]
                            ).first()

                            if match:
                                # ACHOU MANUAL: Só baixa!
                                match.is_paid = True
                                match.payment_date = data_movimento
                                match.bank_account = banco_inter
                                match.description += " (Conciliado Inter)" # Opcional: marca visual
                                match.save()
                                # count_importados += 1 # Opcional: pode criar um contador separado se quiser
                            else:
                                # NÃO ACHOU: Cria novo (Seu código original)
                                cat_prevista, dre_prevista, _ = prever_classificacao(request.user, descricao, 'PAYABLE')
                                PayableAccount.objects.create(
                                    user=request.user,
                                    name=descricao[:100],
                                    description=descricao_final,
                                    due_date=data_movimento,
                                    amount=valor,
                                    category=cat_prevista,
                                    dre_area=dre_prevista or 'OPERACIONAL',
                                    payment_method='DEBITO_CONTA',
                                    cost_type='VARIAVEL',
                                    occurrence='AVULSO',
                                    is_paid=True,
                                    payment_date=data_movimento,
                                    bank_account=banco_inter
                                )
                                count_importados += 1
                        else:
                            count_duplicados += 1

                    # --- LÓGICA RECEBER (Crédito) ---
                    else:
                        # Verifica se JÁ FOI IMPORTADO
                        existe = ReceivableAccount.objects.filter(
                            user=request.user,
                            amount=valor,
                            payment_date=data_movimento,
                            description=descricao_final
                        ).exists()

                        if not existe:
                            # [CIRÚRGICO] Tenta achar conta manual aberta antes de criar
                            match = ReceivableAccount.objects.filter(
                                user=request.user,
                                is_received=False,      # Tem que estar ABERTA
                                amount=valor,           # Valor exato
                                due_date__range=[date_window_start, date_window_end]
                            ).first()

                            if match:
                                # ACHOU MANUAL: Só baixa!
                                match.is_received = True
                                match.payment_date = data_movimento
                                match.bank_account = banco_inter
                                match.description += " (Conciliado Inter)"
                                match.save()
                            else:
                                # NÃO ACHOU: Cria novo (Seu código original)
                                cat_prevista, dre_prevista, _ = prever_classificacao(request.user, descricao, 'RECEIVABLE')
                                ReceivableAccount.objects.create(
                                    user=request.user,
                                    name=descricao[:100],
                                    description=descricao_final,
                                    due_date=data_movimento,
                                    amount=valor,
                                    category=cat_prevista,
                                    dre_area=dre_prevista or 'BRUTA',
                                    payment_method='DEBITO_CONTA',
                                    occurrence='AVULSO',
                                    is_received=True,
                                    payment_date=data_movimento,
                                    bank_account=banco_inter
                                )
                                count_importados += 1
                        else:
                            count_duplicados += 1
                
                # Feedback ao usuário
                if count_importados > 0:
                    messages.success(request, f"Sincronização concluída! {count_importados} novos lançamentos importados.")
                
                if count_duplicados > 0:
                    messages.info(request, f"{count_duplicados} lançamentos já existiam e foram ignorados.")
                
                if count_importados == 0 and count_duplicados == 0:
                    messages.warning(request, "Nenhuma transação encontrada no período.")
                else:
                    messages.info(request, "Conexão realizada, mas nenhuma transação nova encontrada no período.")
            
            return redirect('importar_ofx')
        # --- [FIM] COLAR O CÓDIGO AQUI ---

        # ▼▼▼ [INÍCIO] NOVO BLOCO MERCADO PAGO ▼▼▼
        elif 'sync_mp' in request.POST:
            # 1. Define o período (últimos 7 dias, igual ao Inter)
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=7)
            
            # 2. Chama a API
            resultado = buscar_extrato_mercadopago(request.user, start_date, end_date)
            
            if 'erro' in resultado:
                messages.error(request, f"Erro na integração Mercado Pago: {resultado['erro']}")
            else:
                transacoes = resultado.get('transacoes', [])
                
                # Tenta localizar ou criar uma conta bancária para o MP
                banco_mp = BankAccount.objects.filter(user=request.user, bank_name__icontains='Mercado Pago').first()
                if not banco_mp:
                    # Cria um banco virtual se não existir, para vincular os lançamentos
                    banco_mp = BankAccount.objects.create(
                        user=request.user, 
                        bank_name='Mercado Pago', 
                        agency='0001', 
                        account_number='DIGITAL', 
                        initial_balance=0
                    )

                count_importados = 0
                count_duplicados = 0

                for transacao in transacoes:
                    # Dados vindos do utils_mercadopago
                    data_movimento_str = transacao['data']
                    valor = Decimal(str(transacao['valor'])) # Valor Bruto
                    descricao_origem = transacao['descricao']
                    id_mp = transacao['id_mp']
                    
                    try:
                        data_movimento = datetime.strptime(data_movimento_str, '%Y-%m-%d').date()
                    except:
                        data_movimento = end_date

                    # Monta descrição única
                    descricao_final = f"Venda MP - {descricao_origem} (ID: {id_mp})"

                    # --- LÓGICA ANTI-DUPLICIDADE ---
                    # Verifica se já existe um lançamento com mesmo valor, data e descrição
                    existe = ReceivableAccount.objects.filter(
                        user=request.user,
                        amount=valor,
                        payment_date=data_movimento,
                        description=descricao_final
                    ).exists()

                    if not existe:
                        # --- LÓGICA INTELIGENTE (IA) ---
                        cat_prevista, dre_prevista, _ = prever_classificacao(request.user, descricao_origem, 'RECEIVABLE')
                        
                        # Categorias Padrão caso a IA não encontre
                        categoria_padrao, _ = Category.objects.get_or_create(
                            user=request.user, 
                            name='Receitas de Vendas', 
                            category_type='RECEIVABLE'
                        )

                        ReceivableAccount.objects.create(
                            user=request.user,
                            name=descricao_origem[:100],
                            description=descricao_final,
                            due_date=data_movimento,
                            amount=valor,
                            category=cat_prevista or categoria_padrao,
                            dre_area=dre_prevista or 'BRUTA',
                            payment_method='CREDITO', # Assume Crédito/Digital
                            occurrence='AVULSO',
                            is_received=True, # Já entra como Recebido (Aprovado)
                            payment_date=data_movimento,
                            bank_account=banco_mp,
                            external_id=id_mp # Salva o ID do MP para referência futura
                        )
                        count_importados += 1
                    else:
                        count_duplicados += 1
                
                # Feedback
                if count_importados > 0:
                    messages.success(request, f"Mercado Pago: {count_importados} vendas importadas com sucesso!")
                elif count_duplicados > 0:
                    messages.info(request, "Mercado Pago: Nenhuma venda nova. Transações já importadas.")
                else:
                    messages.warning(request, "Mercado Pago: Nenhuma transação encontrada nos últimos 7 dias.")
            
            return redirect('importar_ofx')
        # ▲▲▲ [FIM] NOVO BLOCO MERCADO PAGO ▲▲▲

        # ▼▼▼ [INÍCIO] NOVO BLOCO ASAAS ▼▼▼
        elif 'sync_asaas' in request.POST:
            # 1. Define o período (Últimos 30 dias, pois Asaas tem taxas antigas que são importantes)
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)
            
            # 2. Chama a API
            resultado = buscar_extrato_asaas(request.user, start_date, end_date)
            
            if 'erro' in resultado:
                messages.error(request, f"Erro na integração Asaas: {resultado['erro']}")
            else:
                transacoes = resultado.get('transacoes', [])
                
                # Cria/Busca banco virtual Asaas
                banco_asaas, _ = BankAccount.objects.get_or_create(
                    user=request.user, 
                    bank_name='Asaas IP S.A.',
                    defaults={'agency': '0001', 'account_number': 'DIGITAL', 'initial_balance': 0}
                )

                count_importados = 0
                count_duplicados = 0

                for transacao in transacoes:
                    data_movimento_str = transacao['data']
                    valor = Decimal(str(transacao['valor']))
                    descricao_origem = transacao['descricao']
                    id_asaas = transacao['id_asaas']
                    tipo_op = transacao['tipo'] # 'C' ou 'D'
                    
                    try:
                        data_movimento = datetime.strptime(data_movimento_str, '%Y-%m-%d').date()
                    except:
                        data_movimento = end_date

                    descricao_final = f"Asaas - {descricao_origem} (ID: {id_asaas})"

                    # --- LÓGICA DE ENTRADA (RECEIVABLE) ---
                    if tipo_op == 'C':
                        existe = ReceivableAccount.objects.filter(
                            user=request.user,
                            external_id=id_asaas # Verifica pelo ID único do Asaas
                        ).exists()

                        if not existe:
                            cat_prevista, dre_prevista, _ = prever_classificacao(request.user, descricao_origem, 'RECEIVABLE')
                            
                            categoria_padrao, _ = Category.objects.get_or_create(
                                user=request.user, name='Receitas Asaas', category_type='RECEIVABLE'
                            )

                            ReceivableAccount.objects.create(
                                user=request.user,
                                name=descricao_origem[:100],
                                description=descricao_final,
                                due_date=data_movimento,
                                amount=valor,
                                category=cat_prevista or categoria_padrao,
                                dre_area=dre_prevista or 'BRUTA',
                                payment_method='BOLETO', # Padrão Asaas
                                occurrence='AVULSO',
                                is_received=True,
                                payment_date=data_movimento,
                                bank_account=banco_asaas,
                                external_id=id_asaas
                            )
                            count_importados += 1
                        else:
                            count_duplicados += 1

                    # --- LÓGICA DE SAÍDA (PAYABLE) ---
                    # Ex: Taxas de boleto, Transferências, Saques
                    elif tipo_op == 'D':
                        existe = PayableAccount.objects.filter(
                            user=request.user,
                            external_id=id_asaas
                        ).exists()

                        if not existe:
                            cat_prevista, dre_prevista, _ = prever_classificacao(request.user, descricao_origem, 'PAYABLE')
                            
                            categoria_padrao, _ = Category.objects.get_or_create(
                                user=request.user, name='Taxas Bancárias', category_type='PAYABLE'
                            )

                            PayableAccount.objects.create(
                                user=request.user,
                                name=descricao_origem[:100],
                                description=descricao_final,
                                due_date=data_movimento,
                                amount=valor,
                                category=cat_prevista or categoria_padrao,
                                dre_area=dre_prevista or 'DESPESAS_OPERACIONAIS',
                                payment_method='DEBITO_CONTA',
                                cost_type='VARIAVEL',
                                occurrence='AVULSO',
                                is_paid=True,
                                payment_date=data_movimento,
                                bank_account=banco_asaas,
                                external_id=id_asaas
                            )
                            count_importados += 1
                        else:
                            count_duplicados += 1
                
                if count_importados > 0:
                    messages.success(request, f"Asaas: {count_importados} movimentações importadas com sucesso!")
                elif count_duplicados > 0:
                    messages.info(request, "Asaas: Nenhuma movimentação nova.")
                else:
                    messages.warning(request, "Asaas: Nenhuma transação encontrada no período.")
            
            return redirect('importar_ofx')
        # ▲▲▲ [FIM] NOVO BLOCO ASAAS ▲▲▲

        # ▼▼▼ [INÍCIO] NOVO BLOCO OMIE ▼▼▼
        elif 'sync_omie' in request.POST:
            resultado = sincronizar_omie_completo(request.user)
            
            if 'erro' in resultado:
                messages.error(request, f"Erro na integração Omie: {resultado['erro']}")
            else:
                total_novos = resultado['pagar_novos'] + resultado['receber_novos']
                total_atualizados = resultado['pagar_atualizados'] + resultado['receber_atualizados']
                
                if total_novos > 0 or total_atualizados > 0:
                    msg = f"Omie Sincronizado: {total_novos} novos lançamentos e {total_atualizados} atualizados."
                    messages.success(request, msg)
                else:
                    messages.info(request, "Omie conectado, mas não houve alterações nos dados.")
                
                if resultado.get('erros'):
                    messages.warning(request, f"Alguns itens tiveram erro: {resultado['erros'][:3]}...") # Mostra os 3 primeiros erros
            
            return redirect('importar_ofx')
        # ▲▲▲ [FIM] NOVO BLOCO OMIE ▲▲▲
        # ▼▼▼ [INÍCIO] NOVO BLOCO NIBO ▼▼▼
        elif 'sync_nibo' in request.POST:
            resultado = sincronizar_nibo_completo(request.user)
            
            if 'erro' in resultado:
                messages.error(request, f"Erro na integração Nibo: {resultado['erro']}")
            else:
                total_novos = resultado['pagar_novos'] + resultado['receber_novos']
                total_atualizados = resultado['pagar_atualizados'] + resultado['receber_atualizados']
                
                if total_novos > 0 or total_atualizados > 0:
                    msg = f"Nibo Sincronizado: {total_novos} novos lançamentos e {total_atualizados} atualizados."
                    messages.success(request, msg)
                else:
                    messages.info(request, "Nibo conectado, mas não houve alterações nos dados.")
                
                if resultado.get('erros'):
                    # Mostra os 3 primeiros erros para não poluir
                    messages.warning(request, f"Alguns itens tiveram erro: {resultado['erros'][:3]}...")
            
            return redirect('importar_ofx')
        # ▲▲▲ [FIM] NOVO BLOCO NIBO ▲▲▲

        # ▼▼▼ [INÍCIO] NOVO BLOCO TINY ERP ▼▼▼
        elif 'sync_tiny' in request.POST:
            resultado = sincronizar_tiny_completo(request.user)
            
            if 'erro' in resultado:
                messages.error(request, f"Erro na integração Tiny: {resultado['erro']}")
            else:
                total_novos = resultado['pagar_novos'] + resultado['receber_novos']
                total_atualizados = resultado['pagar_atualizados'] + resultado['receber_atualizados']
                
                if total_novos > 0 or total_atualizados > 0:
                    msg = f"Tiny Sincronizado: {total_novos} novos lançamentos e {total_atualizados} atualizados."
                    messages.success(request, msg)
                else:
                    messages.info(request, "Tiny conectado, mas não houve alterações nos dados.")
                
                if resultado.get('erros'):
                    # Mostra os 3 primeiros erros para não poluir a tela
                    messages.warning(request, f"Alguns itens tiveram erro: {resultado['erros'][:3]}...")
            
            return redirect('importar_ofx')
        # ▲▲▲ [FIM] NOVO BLOCO TINY ERP ▲▲▲


        form = OFXImportForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            ofx_import = form.save(commit=False)
            ofx_import.user = request.user
            ofx_import.bank_name = ofx_import.bank_account.bank_name
            ofx_import.save()
            try:
                ofx_file = ofx_import.file
                ofx_file.seek(0)
                # ... (bloco de decodificação do OFX continua igual) ...
                raw_content = ofx_file.read()
                content = None
                encodings = ['utf-8', 'cp1252', 'iso-8859-1', 'latin-1']
                for encoding in encodings:
                    try:
                        content = raw_content.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                if content is None:
                    content = raw_content.decode('utf-8', errors='replace')

                # ▼▼▼▼▼▼▼▼▼▼ INÍCIO DA CORREÇÃO ▼▼▼▼▼▼▼▼▼▼
                # Corrige erro "Empty transaction name" preenchendo tags vazias
                content = re.sub(r'<NAME>\s*</NAME>', '<NAME>Nao Informado</NAME>', content, flags=re.IGNORECASE)
                # ▲▲▲▲▲▲▲▲▲▲ FIM DA CORREÇÃO ▲▲▲▲▲▲▲▲▲▲    

                if not content.strip().startswith('<OFX'):
                     content = re.sub(r'^OFXHEADER:.*?(?=<)', '', content, flags=re.MULTILINE | re.DOTALL)
                     content = re.sub(r'(<[A-Z/][^>]*>)\s*([^<]+)\s*(?=<|$)', r'\1\2</\1>', content)
                     content = content.encode('ascii', errors='ignore').decode('ascii')
                     if not content.strip().startswith('<OFX'):
                        content = f"<OFX><BANKMSGSRSV1><STMTTRNRS><STMTRS><BANKTRANLIST>{content}</BANKTRANLIST></STMTRS></STMTTRNRS></BANKMSGSRSV1></OFX>"

                ofx_file_content = StringIO(content)
                ofx = OfxParser.parse(ofx_file_content)
                if not hasattr(ofx, 'account'):
                    raise ValueError("O arquivo OFX não contém informações de conta válidas.")
                
                bank_account = ofx_import.bank_account
                
                # --- INÍCIO DA LÓGICA DE RECONCILIAÇÃO ---
                
                payable_category, _ = Category.objects.get_or_create(name="Despesas Administrativas")
                receivable_category, _ = Category.objects.get_or_create(name="Receitas sobre Vendas")
                
                # 1. NOVOS CONTADORES
                new_transactions_count = 0
                skipped_transactions_count = 0
                reconciled_count = 0 # <<< NOVO

                existing_fitids = set(
                    PayableAccount.objects.filter(user=request.user, fitid__isnull=False).values_list('fitid', flat=True)
                )
                existing_fitids.update(
                    ReceivableAccount.objects.filter(user=request.user, fitid__isnull=False).values_list('fitid', flat=True)
                )
                
                # Define a "janela" de dias para procurar (ex: 3 dias antes e 3 dias depois)
                DATE_WINDOW_DAYS = 3

                for transaction in ofx.account.statement.transactions:
                    
                    fitid = transaction.id
                    
                    # Passo 1: Lógica anti-duplicação (que já tínhamos)
                    if fitid in existing_fitids:
                        skipped_transactions_count += 1
                        continue 

                    amount = Decimal(str(transaction.amount))
                    date = transaction.date.date() # Data real da transação
                    name = (transaction.memo or f"Transação OFX {transaction.id}").encode('ascii', errors='ignore').decode('ascii')
                    description = f"Transação OFX {transaction.id}"
                    
                    # Define a janela de datas para a busca
                    date_window_start = date - timedelta(days=DATE_WINDOW_DAYS)
                    date_window_end = date + timedelta(days=DATE_WINDOW_DAYS)

                    match_found = False

                    # Passo 2: Lógica de "Smart-Matching"
                    if amount < 0: # É um DÉBITO (Contas a Pagar)
                        # Procura uma conta a pagar correspondente
                        match = PayableAccount.objects.filter(
                            Q(bank_account=bank_account) | Q(bank_account__isnull=True), # Do banco certo OU sem banco
                            user=request.user,
                            is_paid=False,                     # Que esteja em aberto
                            fitid__isnull=True,                # Que não tenha sido conciliada ainda
                            amount=abs(amount),                # Com o valor EXATO
                            due_date__range=[date_window_start, date_window_end] # E na janela de data
                        ).order_by('due_date').first() # Pega a mais próxima

                        if match:
                            match.is_paid = True
                            match.payment_date = date # Data real do pagamento (do OFX)
                            match.fitid = fitid       # Vincula o ID do OFX
                            match.bank_account = bank_account # Confirma o banco
                            match.ofx_import = ofx_import
                            match.save()
                            
                            reconciled_count += 1
                            match_found = True
                            existing_fitids.add(fitid) # Adiciona para não ser processado de novo

                    else: # É um CRÉDITO (Contas a Receber)
                        # Procura uma conta a receber correspondente
                        match = ReceivableAccount.objects.filter(
                            Q(bank_account=bank_account) | Q(bank_account__isnull=True),
                            user=request.user,
                            is_received=False,
                            fitid__isnull=True,
                            amount=amount, # Valor exato
                            due_date__range=[date_window_start, date_window_end]
                        ).order_by('due_date').first()

                        if match:
                            match.is_received = True
                            match.payment_date = date
                            match.fitid = fitid
                            match.bank_account = bank_account
                            match.ofx_import = ofx_import
                            match.save()
                            
                            reconciled_count += 1
                            match_found = True
                            existing_fitids.add(fitid)
                    
                    # Passo 3: Se não encontrou match, cria um novo lançamento
                    if not match_found:
                        new_transactions_count += 1
                        existing_fitids.add(fitid)

                        if amount < 0:
                            # --- LÓGICA INTELIGENTE PARA PAGAR ---
                            # 1. Tenta prever com base no dicionário
                            cat_prevista, dre_prevista, _ = prever_classificacao(request.user, name, 'PAYABLE')
                            
                            # 2. Define os valores finais (usa o previsto ou o padrão)
                            categoria_final = cat_prevista if cat_prevista else payable_category
                            dre_final = dre_prevista if dre_prevista else 'OPERACIONAL'
                            
                            PayableAccount.objects.create(
                                user=request.user, 
                                name=name, 
                                description=description, 
                                due_date=date,
                                amount=abs(amount), 
                                category=categoria_final, # <--- USA A PREVISÃO
                                dre_area=dre_final,       # <--- USA A PREVISÃO
                                payment_method='PIX', 
                                occurrence='AVULSO', 
                                
                                is_paid=True,       # Antes era False
                                payment_date=date,  # Adicionado: Data da baixa = Data do extrato
                                
                                cost_type='VARIAVEL', 
                                bank_account=bank_account,
                                ofx_import=ofx_import, 
                                fitid=fitid
                            )
                        else:
                            # --- LÓGICA INTELIGENTE PARA RECEBER ---
                            cat_prevista, dre_prevista, _ = prever_classificacao(request.user, name, 'RECEIVABLE')
                            
                            categoria_final = cat_prevista if cat_prevista else receivable_category
                            dre_final = dre_prevista if dre_prevista else 'BRUTA'

                            ReceivableAccount.objects.create(
                                user=request.user, 
                                name=name, 
                                description=description, 
                                due_date=date,
                                amount=amount, 
                                category=categoria_final, # <--- USA A PREVISÃO
                                dre_area=dre_final,       # <--- USA A PREVISÃO
                                payment_method='PIX', 
                                occurrence='AVULSO', 
                                is_received=True, 
                                payment_date=date,  # Adicionado: Data da baixa = Data do extrato
                                bank_account=bank_account,
                                ofx_import=ofx_import, 
                                fitid=fitid
                            )
                
                # --- MENSAGENS DE SUCESSO MELHORADAS ---
                if new_transactions_count > 0:
                    messages.success(request, f'{new_transactions_count} nova(s) transação(ões) importada(s) com sucesso.')
                
                if reconciled_count > 0: # <<< NOVO
                    messages.info(request, f'{reconciled_count} transação(ões) existente(s) foram conciliadas (marcadas como pagas/recebidas).')
                
                if skipped_transactions_count > 0:
                    messages.warning(request, f'{skipped_transactions_count} transação(ões) duplicada(s) foram ignoradas.')

                if new_transactions_count == 0 and reconciled_count == 0 and skipped_transactions_count == 0:
                    messages.info(request, 'O arquivo foi processado, mas nenhuma transação foi encontrada.')
                elif new_transactions_count == 0 and reconciled_count == 0 and skipped_transactions_count > 0:
                     messages.info(request, 'Nenhuma transação nova foi encontrada. Todas as transações do arquivo já haviam sido importadas.')

                return redirect('importar_ofx') # Redireciona para a própria página de importação

            except Exception as e:
                messages.error(request, f'Erro ao processar o arquivo OFX: {str(e)}')
                ofx_import.delete()
                return redirect('importar_ofx')
        else:
             messages.error(request, f'Erro no formulário: {form.errors.as_text()}')
    else:
        form = OFXImportForm(user=request.user)

    # --- NOVA LÓGICA DE SALDO ---
    saldo_inter = None
    # Verifica se o usuário tem credenciais configuradas antes de tentar buscar
    if hasattr(request.user, 'inter_creds'):
        try:
            # Busca o saldo
            resp_saldo = buscar_saldo_inter(request.user)
            
            # Se a API retornou o campo 'disponivel', salvamos na variável
            if 'disponivel' in resp_saldo:
                saldo_inter = resp_saldo['disponivel']
        except Exception as e:
            print(f"Não foi possível carregar o saldo: {e}")
            # Não vamos travar a página se o saldo falhar, apenas segue sem ele

    # ▼▼▼ NOVO: SALDO MERCADO PAGO ▼▼▼
    saldo_mp = None
    if hasattr(request.user, 'mercadopago_creds'):
        try:
            resp_saldo_mp = buscar_saldo_mercadopago(request.user)
            if 'disponivel' in resp_saldo_mp:
                saldo_mp = resp_saldo_mp['disponivel']
        except Exception as e:
            print(f"Erro saldo MP: {e}")
    # ▲▲▲ FIM NOVO BLOCO ▲▲▲   

    # ▼▼▼ NOVO: SALDO ASAAS ▼▼▼
    saldo_asaas = None
    if hasattr(request.user, 'asaas_creds'):
        try:
            resp_saldo_asaas = buscar_saldo_asaas(request.user)
            if 'disponivel' in resp_saldo_asaas:
                saldo_asaas = resp_saldo_asaas['disponivel']
        except Exception as e:
            print(f"Erro saldo Asaas: {e}")
    # ▲▲▲ FIM NOVO BLOCO ▲▲▲     

    context = {
        'ofx_form': form,
        'saldo_inter': saldo_inter,
        'saldo_mp': saldo_mp,
        'saldo_asaas': saldo_asaas,
    }

    return render(request, 'accounts/importar_ofx.html', context)



@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_contas_pagar')
def contas_pagar(request):
    edit_id_str = request.GET.get('edit') # 1. Pega o ID como string (ex: "1.000")
    instance = None
    if edit_id_str:
        try:
            # 2. Limpa a string, removendo pontos e vírgulas
            cleaned_id = int(re.sub(r'[.,]', '', edit_id_str))
            
            # 3. Usa o ID limpo (ex: 1000) para buscar
            instance = get_object_or_404(PayableAccount, id=cleaned_id, user=request.user)
        except (ValueError, TypeError):
            # 4. Se o ID for inválido (ex: "abc"), não quebra e apenas exibe uma mensagem
            messages.error(request, "O ID da conta para edição é inválido.")
            instance = None

    # Lógica de exportação refatorada
    if request.GET.get('export_pdf') or request.GET.get('export_excel'):
        filter_status = request.GET.get('status', 'all')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        accounts = PayableAccount.objects.filter(user=request.user).order_by('due_date')
        if filter_status == 'open': accounts = accounts.filter(is_paid=False)
        elif filter_status == 'paid': accounts = accounts.filter(is_paid=True)
        if start_date: accounts = accounts.filter(due_date__gte=parse_date(start_date))
        if end_date: accounts = accounts.filter(due_date__lte=parse_date(end_date))

        if request.GET.get('export_pdf'):
            return gerar_pdf_generic(accounts, 'pagar')
        elif request.GET.get('export_excel'):
            return gerar_excel_generic(accounts, 'pagar')


    if request.method == 'POST':
        # --- LÓGICA DE REDIRECIONAMENTO INTELIGENTE ---
        query_params = request.GET.copy()
        
        # COMENTE OU REMOVA A LINHA ABAIXO PARA MANTER A PÁGINA ATUAL
        # query_params.pop('page', None) 
        
        query_params.pop('edit', None) # Remove 'edit' para sair do modo de edição (isso deve ficar)

        # Isso garante que status, data, banco, busca e PÁGINA sejam mantidos
        redirect_query_string = query_params.urlencode()
        redirect_url = f"{request.path}?{redirect_query_string}"

        # Cria a URL de redirecionamento com os filtros
        redirect_url = f"{request.path}?{redirect_query_string}"
        # --- FIM DA LÓGICA DE REDIRECIONAMENTO ---

        # ▼▼▼ SUBSTITUA O BLOCO 'import_excel' INTEIRO POR ESTE NOVO BLOCO MELHORADO ▼▼▼
        if request.POST.get('action') == 'import_excel':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Nenhum arquivo de planilha foi enviado.")
                return redirect(redirect_url)

            try:
                df = pd.read_excel(excel_file)
                
                column_map = {
                    'nome': 'name', 'cliente': 'name', 'fornecedor': 'name',
                    'descrição': 'description', 'histórico': 'description',
                    'vencimento': 'due_date', 'data do pagamento': 'due_date', 'data do vencimento': 'due_date',
                    'valor': 'amount', 'pagamento': 'amount',
                    'categoria': 'category',
                    'conta bancária': 'bank_account', 'banco': 'bank_account',
                    'forma de pagamento': 'payment_method',
                    'custo': 'cost_type',
                }

                # --- INÍCIO DA CORREÇÃO DA MENSAGEM DE ERRO ---
                
                # 1. Guarda os nomes originais das colunas da planilha (em minúsculo).
                original_columns = {str(col).strip().lower() for col in df.columns}
                
                # 2. Dicionário para criar mensagens de erro amigáveis.
                friendly_names_map = {
                    'name': "'Nome', 'Cliente' ou 'Fornecedor'",
                    'due_date': "'Vencimento', 'Data do pagamento' ou 'Data do vencimento'",
                    'amount': "'Valor' ou 'Pagamento'"
                }
                
                # 3. Verifica ANTES de renomear se as colunas obrigatórias existem.
                required_internal_names = {'name', 'due_date', 'amount'}
                
                # Inverte o mapa de colunas para facilitar a busca.
                reverse_map = {v: k for k, v in column_map.items()}
                
                # Pega todas as chaves do mapa original (ex: 'nome', 'cliente', 'vencimento'...).
                valid_original_names = set(column_map.keys())
                
                # Renomeia as colunas da planilha para os nomes internos.
                df.rename(columns={col: column_map.get(col.strip().lower()) for col in df.columns}, inplace=True)

                # 4. Verifica DEPOIS de renomear se os nomes internos foram criados.
                for internal_name in required_internal_names:
                    if internal_name not in df.columns:
                        # Se um nome interno não foi criado, significa que nenhum dos seus nomes amigáveis existia na planilha.
                        # Agora, a mensagem de erro usa o mapa amigável!
                        raise ValueError(f"A coluna obrigatória {friendly_names_map[internal_name]} não foi encontrada. Verifique os nomes das colunas da sua planilha.")
                        
                # --- FIM DA CORREÇÃO DA MENSAGEM DE ERRO ---

                successful_imports = 0
                failed_rows = []

                with transaction.atomic():
                    for index, row in df.iterrows():
                        excel_row_number = index + 2
                        try:
                            # 1. Tratamento Básico de Dados (Nome, Descrição, Valor, Data)
                            # -----------------------------------------------------------
                            name = str(row['name']).strip()
                            description = str(row.get('description', 'Importado via Excel'))
                            
                            due_date = pd.to_datetime(row['due_date'], errors='coerce', dayfirst=True).date()
                            if pd.isnull(due_date):
                                raise ValueError("Formato de data inválido.")

                            amount_str = str(row['amount']).replace(',', '.')
                            amount = float(amount_str)

                            payment_method = str(row.get('payment_method', 'BOLETO')).strip().upper()
                            cost_type_value = str(row.get('cost_type', 'FIXO')).strip().upper()
                            if cost_type_value not in ['FIXO', 'VARIAVEL']:
                                cost_type_value = 'FIXO'

                            # 2. Lógica Inteligente de Classificação (Categoria, DRE, Banco)
                            # -----------------------------------------------------------
                            
                            # A. Tenta prever com base no histórico (Dicionário)
                            cat_smart, dre_smart, bank_smart = prever_classificacao(request.user, name, 'PAYABLE')

                            # B. Define a Categoria Final
                            # Prioridade: Excel > Inteligência > Padrão (Criar nova)
                            category = None
                            if pd.notna(row.get('category')):
                                cat_name_excel = str(row.get('category')).strip()
                                category, _ = Category.objects.get_or_create(name=cat_name_excel, category_type='PAYABLE', user=request.user)
                            elif cat_smart:
                                category = cat_smart
                            else:
                                # Se não tem no Excel e nem no Smart, cria/usa uma padrão
                                category, _ = Category.objects.get_or_create(name='Despesas Gerais', category_type='PAYABLE', user=request.user)

                            # C. Define a Área DRE Final
                            # Prioridade: Inteligência > Padrão ('OPERACIONAL')
                            # (Nota: Dificilmente vem DRE no Excel, então confiamos no Smart)
                            dre_final = dre_smart if dre_smart else 'OPERACIONAL'

                            # D. Define o Banco Final
                            # Prioridade: Excel (Nome do Banco) > Inteligência > None
                            bank_account = None
                            if pd.notna(row.get('bank_account')):
                                bank_name_excel = str(row.get('bank_account')).strip()
                                bank_account = BankAccount.objects.filter(user=request.user, bank_name__icontains=bank_name_excel).first()
                            
                            if not bank_account and bank_smart:
                                bank_account = bank_smart

                            # 3. Criação do Objeto
                            # -----------------------------------------------------------
                            PayableAccount.objects.create(
                                user=request.user, 
                                name=name, 
                                description=description,
                                due_date=due_date, 
                                amount=amount, 
                                category=category, 
                                bank_account=bank_account, # Banco Inteligente ou do Excel
                                dre_area=dre_final,        # DRE Inteligente ou Padrão
                                payment_method=payment_method,
                                cost_type=cost_type_value, 
                                occurrence='AVULSO', 
                                is_paid=False
                            )
                            successful_imports += 1

                        except Exception as e:
                            failed_rows.append(f"Linha {excel_row_number}: {e}")

                if successful_imports > 0:
                    messages.success(request, f"{successful_imports} conta(s) foi(ram) importada(s) com sucesso!")
                
                if failed_rows:
                    error_details = "; ".join(failed_rows)
                    messages.error(request, f"Falha ao importar {len(failed_rows)} linha(s). Detalhes: {error_details}")
                
                if successful_imports == 0 and not failed_rows:
                    messages.warning(request, "A planilha foi processada, mas nenhuma linha válida para importação foi encontrada.")

            except Exception as e:
                messages.error(request, f"Ocorreu um erro geral ao processar o arquivo: {e}")

            return redirect(redirect_url)
        # --- FIM DO BLOCO DE SUBSTITUIÇÃO ---


        if 'action_pay' in request.POST:
            account_id = request.POST.get('action_pay')
            account = get_object_or_404(PayableAccount, id=account_id, user=request.user)
            account.is_paid = True
            account.payment_date = datetime.today().date()
            account.save()
            messages.success(request, 'Conta marcada como paga.')
            return redirect(redirect_url) # ALTERADO

        elif 'action_undo' in request.POST:
            account_id = request.POST.get('action_undo')
            account = get_object_or_404(PayableAccount, id=account_id, user=request.user)
            
            original_fitid = account.fitid
            original_payment_date = account.payment_date # Captura a data antes de limpar
            
            # 1. Reseta a conta manual (a conta original)
            account.is_paid = False
            account.payment_date = None
            account.fitid = None
            account.ofx_import = None
            account.save()
            messages.success(request, f'A baixa da conta manual "{account.name}" foi desfeita.')

            # 2. Se a conta TINHA um fitid, ela foi conciliada por OFX.
            #    Agora, criamos a "outra ponta" (a transação do OFX)
            #    como um novo lançamento em aberto.
            if original_fitid:
                transaction_date = original_payment_date
                if not transaction_date:
                    transaction_date = account.due_date # Fallback
                
                # Pega a categoria padrão
                payable_category, _ = Category.objects.get_or_create(name="Despesas Administrativas")

                PayableAccount.objects.create(
                    user=request.user,
                    name=f"Transação OFX ({original_fitid})", # Novo nome
                    description=f"Lançamento separado da conta '{account.name}'",
                    due_date=transaction_date, # Data em que ocorreu
                    amount=account.amount, # O mesmo valor
                    category=payable_category, # Categoria padrão
                    dre_area=account.dre_area, # Herda a área DRE
                    payment_method=account.payment_method, # Herda forma de pgto
                    occurrence='AVULSO', 
                    is_paid=False, # <<< Fica em aberto
                    cost_type=account.cost_type, 
                    bank_account=account.bank_account, # Herda o banco
                    ofx_import=None, 
                    fitid=original_fitid # <<< CRUCIAL: Mantém o fitid
                )
                messages.info(request, f'A transação do OFX foi criada como um novo lançamento em aberto.')

            return redirect(redirect_url)

        elif 'action_delete' in request.POST:
            account_id = request.POST.get('action_delete')
            PayableAccount.objects.filter(id=account_id, user=request.user).delete()
            messages.success(request, 'Conta excluída com sucesso.')
            return redirect(redirect_url) # ALTERADO

        elif 'action_attach' in request.POST and 'file' in request.FILES:
            account_id = request.POST.get('attach_account_id')
            account = get_object_or_404(PayableAccount, id=account_id, user=request.user)
            account.file = request.FILES['file']
            account.save()
            messages.success(request, 'Arquivo anexado com sucesso.')
            return redirect(redirect_url) # ALTERADO
        
        # --- NOVO BLOCO: Excluir apenas o arquivo (anexo) ---
        elif 'action_delete_file' in request.POST:
            account_id = request.POST.get('action_delete_file')
            account = get_object_or_404(PayableAccount, id=account_id, user=request.user)
            
            # Define o campo file como None (nulo)
            # Isso remove a referência ao arquivo sem tentar acessá-lo no S3
            account.file = None 
            account.save()
            
            messages.success(request, 'Anexo removido com sucesso.')
            return redirect(redirect_url)
        # ----------------------------------------------------

        elif 'delete_selected' in request.POST:
            account_ids = request.POST.getlist('account_ids')
            if account_ids:
                count, _ = PayableAccount.objects.filter(id__in=account_ids, user=request.user).delete()
                messages.success(request, f'{count} conta(s) selecionada(s) foram excluídas.')
            return redirect(redirect_url) # ALTERADO

        # Lógica para o formulário principal (Salvar ou Salvar e Pagar)
        # Lógica para o formulário principal (Salvar ou Salvar e Pagar)
        else:
            post_instance_id = request.POST.get('account_id')

            # --- INÍCIO DA CORREÇÃO ---
            instance = None
            original_occurrence = None # Padrão para novas contas
            if post_instance_id:
                instance = get_object_or_404(PayableAccount, id=post_instance_id, user=request.user)
                original_occurrence = instance.occurrence # Captura o estado ANTES do form
            # --- FIM DA CORREÇÃO ---
            
            form = PayableAccountForm(request.POST, request.FILES, instance=instance, user=request.user)
            
            if form.is_valid():
                is_new_account = instance is None

                # --- INÍCIO DA CORREÇÃO ---
                # Compara com o estado original, não com o estado em-memória do form
                was_avulso = (not is_new_account) and (original_occurrence == 'AVULSO')
                # --- FIM DA CORREÇÃO ---
                new_category_name = form.cleaned_data.get('new_category')

                if new_category_name:
                    # 2. ADICIONE 'user=request.user' AQUI e corrija o get_or_create
                    category, _ = Category.objects.get_or_create(
                        name=new_category_name, 
                        category_type='PAYABLE',
                        user=request.user  # <-- Linha adicionada
                    )
                    form.instance.category = category
                
                form.instance.user = request.user
                account = form.save()

                # ▼▼▼ ADICIONE ESTE BLOCO ▼▼▼
                # Ensina o dicionário
                aprender_classificacao(
                    user=request.user,
                    nome=account.name,
                    categoria=account.category,
                    dre_area=account.dre_area,
                    tipo='PAYABLE',
                    bank_account=account.bank_account
                )
                # ▲▲▲ FIM DO BLOCO ▲▲▲
                
                is_recorrente_agora = form.cleaned_data.get('occurrence') == 'RECORRENTE'
                recurrence_count = form.cleaned_data.get('recurrence_count')
                
                # Roda a lógica se:
                # 1. É uma conta nova E é recorrente
                # 2. OU era avulsa E foi mudada para recorrente
                if (is_new_account and is_recorrente_agora) or (was_avulso and is_recorrente_agora):
                    if recurrence_count: # Garante que temos um número para iterar
                        for i in range(1, recurrence_count):
                            new_due_date = account.due_date + relativedelta(months=i)
                            
                            PayableAccount.objects.create(
                                user=request.user, name=account.name, description=account.description,
                                due_date=new_due_date, amount=account.amount,
                                category=account.category, dre_area=account.dre_area, payment_method=account.payment_method,
                                occurrence='RECORRENTE', 
                                recurrence_count=recurrence_count,
                                cost_type=account.cost_type, bank_account=account.bank_account
                            )
                
                if 'save_and_pay' in request.POST:
                    account.is_paid = True
                    if not account.payment_date:
                        account.payment_date = datetime.today().date()
                    account.save()
                    messages.success(request, 'Conta salva e marcada como paga.')
                else:
                    messages.success(request, 'Conta salva com sucesso.')
                
                return redirect(redirect_url)
            else:
                messages.error(request, f'Erro no formulário: {form.errors.as_json()}')

    # Em views.py, na função contas_pagar

    # ... (lógica do POST) ...

    # Adicione esse bloco
    # Filtros e renderização final (GET request)
    filter_status = request.GET.get('status', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_query = request.GET.get('search_query', '') # <-- LINHA ADICIONADA
    bank_filter = request.GET.get('bank', '')           # <-- NOVA LINHA

    # Se nenhuma data for fornecida E não houver busca, define o padrão para o mês atual.
    if not start_date and not end_date and not search_query: # <-- CONDIÇÃO MODIFICADA
        today = datetime.today().date()
        # Define a data inicial como o dia 1 do mês corrente.
        start_date = today.replace(day=1)
        # Para o fim do mês, vamos para o próximo mês e voltamos um dia.
        end_of_month = (start_date + relativedelta(months=1)) - timedelta(days=1)
        
        # Converte as datas para o formato de string 'YYYY-MM-DD' para preencher o formulário de filtro.
        end_date = end_of_month.strftime('%Y-%m-%d')
        start_date = start_date.strftime('%Y-%m-%d')
    # Adicione esse bloco
    accounts_query = PayableAccount.objects.filter(user=request.user).order_by('due_date')

    # ▼▼▼ ADICIONE ESTE NOVO BLOCO DE BUSCA ▼▼▼
    if search_query:
        accounts_query = accounts_query.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    # --- FIM DO BLOCO DE BUSCA ---

    # ▼▼▼ ADICIONE ESTE NOVO BLOCO DE FILTRO DE BANCO ▼▼▼
    if bank_filter:
        accounts_query = accounts_query.filter(bank_account__id=bank_filter)
    # --- FIM DO BLOCO DE FILTRO DE BANCO ---

    if filter_status == 'open': accounts_query = accounts_query.filter(is_paid=False)
    elif filter_status == 'paid': accounts_query = accounts_query.filter(is_paid=True)
    if start_date: accounts_query = accounts_query.filter(due_date__gte=parse_date(start_date))
    if end_date: accounts_query = accounts_query.filter(due_date__lte=parse_date(end_date))

    # ▼▼▼ ADICIONE ESTE BLOCO ABAIXO ▼▼▼
    # 1. Calcula o valor total ANTES da paginação
    total_amount = accounts_query.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # 2. Aplica a paginação
    paginator = Paginator(accounts_query, 10) # 10 itens por página
    page_number = request.GET.get('page')
    accounts_page_obj = paginator.get_page(page_number)
    # ▲▲▲ FIM DO BLOCO 
    user_banks = BankAccount.objects.filter(user=request.user).order_by('bank_name') # <-- NOVA LINHA

    form = PayableAccountForm(instance=instance, user=request.user)

    return render(request, 'accounts/contas_pagar.html', {
        'form': form,
        'accounts': accounts_page_obj, # <-- ALTERADO para o objeto da página
        'filter_status': filter_status,
        'start_date': start_date,
        'end_date': end_date,
        'total_amount': total_amount, # <-- ADICIONADO o valor total
        'search_query': search_query,
        'user_banks': user_banks,        
        'bank_filter': bank_filter,
    })


@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_contas_receber')
def contas_receber(request):
    # --- 1. LÓGICA DE FILTRAGEM (GET) ---
    filter_status = request.GET.get('status', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    accounts = ReceivableAccount.objects.filter(user=request.user)

    if filter_status == 'open':
        accounts = accounts.filter(is_received=False)
    elif filter_status == 'received':
        accounts = accounts.filter(is_received=True)
    
    if start_date_str:
        accounts = accounts.filter(due_date__gte=parse_date(start_date_str))
    if end_date_str:
        accounts = accounts.filter(due_date__lte=parse_date(end_date_str))

    accounts_query = accounts.order_by('due_date') # Renomeado para _query
    accounts_for_export = accounts_query

    # Em accounts/views.py, dentro da função contas_receber

    if 'export_pdf' in request.GET:
        return gerar_pdf_generic(accounts_for_export, 'receber')

    if 'export_excel' in request.GET:
        return gerar_excel_generic(accounts_for_export, 'receber')

    # --- 3. LÓGICA DE AÇÕES (POST) ---
    if request.method == 'POST':
        # --- INÍCIO DA CORREÇÃO ---
        # Captura a URL da página atual (incluindo filtros e paginação) para redirecionar de volta
        # Adicione esse bloco (A nova forma "inteligente" que lê os filtros da URL)
        # --- LÓGICA DE REDIRECIONAMENTO INTELIGENTE ---
        query_params = request.GET.copy()
        
        query_params.pop('edit', None) 

        redirect_query_string = query_params.urlencode()
        redirect_url = f"{request.path}?{redirect_query_string}"
        # --- FIM DA LÓGICA DE REDIRECIONAMENTO ---
        # --- FIM DA CORREÇÃO ---
        # ▼▼▼ SUBSTITUA O BLOCO 'import_excel' INTEIRO POR ESTE NOVO BLOCO CORRIGIDO ▼▼▼
        if request.POST.get('action') == 'import_excel':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Nenhum arquivo de planilha foi enviado.")
                return redirect(redirect_url)

            try:
                df = pd.read_excel(excel_file)
                
                column_map = {
                    'nome': 'name', 'cliente': 'name', 'fornecedor': 'name',
                    'descrição': 'description', 'histórico': 'description',
                    'vencimento': 'due_date', 'data do pagamento': 'due_date', 'data do vencimento': 'due_date',
                    'valor': 'amount', 'pagamento': 'amount',
                    'categoria': 'category',
                    'conta bancária': 'bank_account', 'banco': 'bank_account',
                    'forma de pagamento': 'payment_method',
                }

                # --- INÍCIO DA CORREÇÃO DA MENSAGEM DE ERRO ---
                friendly_names_map = {
                    'name': "'Nome', 'Cliente' ou 'Fornecedor'",
                    'due_date': "'Vencimento', 'Data do pagamento' ou 'Data do vencimento'",
                    'amount': "'Valor' ou 'Pagamento'"
                }
                
                # Renomeia as colunas da planilha para os nomes internos.
                df.rename(columns={col: column_map.get(str(col).strip().lower()) for col in df.columns}, inplace=True)

                # Verifica se os nomes internos obrigatórios foram criados após o renomeio.
                required_internal_names = {'name', 'due_date', 'amount'}
                for internal_name in required_internal_names:
                    if internal_name not in df.columns:
                        # Se não foi criado, lança um erro com a mensagem amigável.
                        raise ValueError(f"A coluna obrigatória {friendly_names_map[internal_name]} não foi encontrada. Verifique os nomes das colunas da sua planilha.")
                # --- FIM DA CORREÇÃO DA MENSAGEM DE ERRO ---

                successful_imports = 0
                failed_rows = []

                with transaction.atomic():
                    for index, row in df.iterrows():
                        excel_row_number = index + 2
                        try:
                            # 1. Tratamento Básico
                            name = str(row['name']).strip()
                            description = str(row.get('description', 'Importado via Excel'))
                            
                            due_date = pd.to_datetime(row['due_date'], errors='coerce', dayfirst=True).date()
                            if pd.isnull(due_date):
                                raise ValueError("Formato de data inválido.")

                            amount_str = str(row['amount']).replace(',', '.')
                            amount = float(amount_str)
                            
                            payment_method = str(row.get('payment_method', 'BOLETO')).strip().upper()

                            # 2. Lógica Inteligente (Categoria, DRE, Banco)
                            # -----------------------------------------------------------
                            
                            # A. Previsão
                            cat_smart, dre_smart, bank_smart = prever_classificacao(request.user, name, 'RECEIVABLE')

                            # B. Categoria
                            category = None
                            if pd.notna(row.get('category')):
                                cat_name_excel = str(row.get('category')).strip()
                                category, _ = Category.objects.get_or_create(name=cat_name_excel, category_type='RECEIVABLE', user=request.user)
                            elif cat_smart:
                                category = cat_smart
                            else:
                                category, _ = Category.objects.get_or_create(name='Receitas Gerais', category_type='RECEIVABLE', user=request.user)

                            # C. DRE
                            dre_final = dre_smart if dre_smart else 'BRUTA'

                            # D. Banco
                            bank_account = None
                            if pd.notna(row.get('bank_account')):
                                bank_name_excel = str(row.get('bank_account')).strip()
                                bank_account = BankAccount.objects.filter(user=request.user, bank_name__icontains=bank_name_excel).first()
                            
                            if not bank_account and bank_smart:
                                bank_account = bank_smart

                            # 3. Criação
                            ReceivableAccount.objects.create(
                                user=request.user, 
                                name=name, 
                                description=description,
                                due_date=due_date, 
                                amount=amount, 
                                category=category, 
                                bank_account=bank_account, # Banco Inteligente
                                dre_area=dre_final,        # DRE Inteligente
                                payment_method=payment_method,
                                occurrence='AVULSO',
                                is_received=False
                            )
                            successful_imports += 1

                        except Exception as e:
                            failed_rows.append(f"Linha {excel_row_number}: {e}")

                if successful_imports > 0:
                    messages.success(request, f"{successful_imports} conta(s) foi(ram) importada(s) com sucesso!")
                
                if failed_rows:
                    error_details = "; ".join(failed_rows)
                    messages.error(request, f"Falha ao importar {len(failed_rows)} linha(s). Detalhes: {error_details}")
                
                if successful_imports == 0 and not failed_rows:
                    messages.warning(request, "A planilha foi processada, mas nenhuma linha válida para importação foi encontrada.")

            except Exception as e:
                messages.error(request, f"Ocorreu um erro geral ao processar o arquivo: {e}")

            return redirect(redirect_url)
        # --- FIM DO BLOCO DE SUBSTITUIÇÃO ---
        # ▼▼▼ ADICIONE ESTE BLOCO ▼▼▼
        elif 'action_attach' in request.POST and 'file' in request.FILES:
            account_id = request.POST.get('attach_account_id')
            account = get_object_or_404(ReceivableAccount, id=account_id, user=request.user)
            account.file = request.FILES['file']
            account.save()
            messages.success(request, 'Arquivo anexado com sucesso.')
            return redirect(redirect_url) # Redireciona para a URL com filtros

        elif 'action_delete_file' in request.POST:
            account_id = request.POST.get('action_delete_file')
            account = get_object_or_404(ReceivableAccount, id=account_id, user=request.user)
            
            # Define o campo file como None (nulo)
            account.file = None 
            account.save()
            
            messages.success(request, 'Anexo removido com sucesso.')
            return redirect(redirect_url) # Redireciona para a URL com filtros
        # ▲▲▲ FIM DO BLOCO ADICIONADO ▲▲▲


        if 'action_receive' in request.POST:
            account_id = request.POST.get('action_receive')
            account = get_object_or_404(ReceivableAccount, id=account_id, user=request.user)
            account.is_received = True
            account.payment_date = datetime.today().date()
            account.save()
            messages.success(request, 'Conta marcada como recebida.')
            return redirect(redirect_url) # ALTERADO

        elif 'action_undo' in request.POST:
            account_id = request.POST.get('action_undo')
            account = get_object_or_404(ReceivableAccount, id=account_id, user=request.user)
            
            original_fitid = account.fitid
            original_payment_date = account.payment_date # Captura a data
            
            # 1. Reseta a conta manual
            account.is_received = False
            account.payment_date = None
            account.fitid = None
            account.ofx_import = None
            account.save()
            messages.success(request, f'A baixa da conta manual "{account.name}" foi desfeita.')

            # 2. Se a conta TINHA um fitid, cria o lançamento do OFX
            if original_fitid:
                transaction_date = original_payment_date
                if not transaction_date:
                    transaction_date = account.due_date # Fallback
                
                # Pega a categoria padrão
                receivable_category, _ = Category.objects.get_or_create(name="Receitas sobre Vendas")

                ReceivableAccount.objects.create(
                    user=request.user,
                    name=f"Transação OFX ({original_fitid})", 
                    description=f"Lançamento separado da conta '{account.name}'",
                    due_date=transaction_date,
                    amount=account.amount,
                    category=receivable_category,
                    dre_area=account.dre_area,
                    payment_method=account.payment_method,
                    occurrence='AVULSO', 
                    is_received=False, # <<< Fica em aberto
                    bank_account=account.bank_account,
                    ofx_import=None, 
                    fitid=original_fitid # <<< Mantém o fitid
                )
                messages.info(request, f'A transação do OFX foi criada como um novo lançamento em aberto.')

            return redirect(redirect_url)

        elif 'action_delete' in request.POST:
            account_id = request.POST.get('action_delete')
            get_object_or_404(ReceivableAccount, id=account_id, user=request.user).delete()
            messages.success(request, 'Conta a receber excluída com sucesso.')
            return redirect(redirect_url) # ALTERADO

        elif 'delete_selected' in request.POST:
            account_ids = request.POST.getlist('account_ids')
            if account_ids:
                count, _ = ReceivableAccount.objects.filter(id__in=account_ids, user=request.user).delete()
                messages.success(request, f'{count} conta(s) selecionada(s) foram excluídas.')
            else:
                messages.warning(request, 'Nenhuma conta foi selecionada para exclusão.')
            return redirect(redirect_url) # ALTERADO

        # Se nenhuma ação foi acionada, processa o formulário principal
        # Se nenhuma ação foi acionada, processa o formulário principal
        else:
            # --- INÍCIO DA CORREÇÃO ---
            instance = None
            original_occurrence = None # Padrão
            account_id = request.POST.get('account_id')
            if account_id:
                instance = get_object_or_404(ReceivableAccount, id=account_id, user=request.user)
                original_occurrence = instance.occurrence # Captura o estado original
            # --- FIM DA CORREÇÃO ---

            form = ReceivableAccountForm(request.POST, request.FILES, instance=instance, user=request.user)
            
            if form.is_valid():
                is_new_account = instance is None
                
                # --- INÍCIO DA CORREÇÃO ---
                # Compara com o estado original, não com o estado em-memória do form
                was_avulso = (not is_new_account) and (original_occurrence == 'AVULSO')
                # --- FIM DA CORREÇÃO ---
                new_category_name = form.cleaned_data.get('new_category')
                if new_category_name:
                    # 2. ADICIONE 'user=request.user' AQUI e corrija o get_or_create
                    category, _ = Category.objects.get_or_create(
                        name=new_category_name, 
                        category_type='RECEIVABLE',
                        user=request.user  # <-- Linha adicionada
                    )
                    form.instance.category = category
                
                form.instance.user = request.user
                account = form.save()

                # ▼▼▼ ADICIONE ESTE BLOCO ▼▼▼
                # Ensina o dicionário
                aprender_classificacao(
                    user=request.user,
                    nome=account.name,
                    categoria=account.category,
                    dre_area=account.dre_area,
                    tipo='RECEIVABLE',
                    bank_account=account.bank_account
                )
                # ▲▲▲ FIM DO BLOCO ▲▲▲

                is_recorrente_agora = form.cleaned_data.get('occurrence') == 'RECORRENTE'
                recurrence_count = form.cleaned_data.get('recurrence_count')
                
                # Roda a lógica se:
                # 1. É uma conta nova E é recorrente
                # 2. OU era avulsa E foi mudada para recorrente
                if (is_new_account and is_recorrente_agora) or (was_avulso and is_recorrente_agora):
                    if recurrence_count: # Garante que temos um número para iterar
                        for i in range(1, recurrence_count):
                            new_due_date = account.due_date + relativedelta(months=i)
                            ReceivableAccount.objects.create(
                                user=request.user, name=account.name, description=account.description,
                                due_date=new_due_date, amount=account.amount, category=account.category,
                                dre_area=account.dre_area, payment_method=account.payment_method,
                                occurrence='RECORRENTE', recurrence_count=recurrence_count,
                                bank_account=account.bank_account
                            )
                if 'save_and_receive' in request.POST:
                    account.is_received = True
                    if not account.payment_date:
                        account.payment_date = datetime.today().date()
                    account.save()
                    messages.success(request, 'Conta salva e recebida com sucesso.')
                else:
                    messages.success(request, 'Conta salva com sucesso.')
                return redirect(redirect_url)
            else:
                messages.error(request, 'Erro no formulário. Verifique os campos e tente novamente.')

    # Adicione esse bloco
    # --- LÓGICA GET COMPLETA E CORRIGIDA ---

    # Adicione esse bloco
    # Pega os filtros da URL.
    filter_status = request.GET.get('status', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_query = request.GET.get('search_query', '') # <-- LINHA ADICIONADA
    bank_filter = request.GET.get('bank', '')           # <-- NOVA LINHA

    # Se nenhuma data for fornecida, define o padrão para o mês atual.
    # Adicione esse bloco
    # Se nenhuma data for fornecida E não houver busca, define o padrão para o mês atual.
    if not start_date and not end_date and not search_query: # <-- CONDIÇÃO MODIFICADA
        today = datetime.today().date()
        start_date_obj = today.replace(day=1)
        end_of_month = (start_date_obj + relativedelta(months=1)) - timedelta(days=1)
        # Converte para string para preencher o formulário e usar no filtro.
        start_date = start_date_obj.strftime('%Y-%m-%d')
        end_date = end_of_month.strftime('%Y-%m-%d')

    # Lógica de edição (se houver o parâmetro 'edit' na URL).
    instance_to_edit = None
    edit_id_str = request.GET.get('edit') # 1. Pega o ID "sujo" (ex: "1.000")
    if edit_id_str:
        try:
            # 2. Limpa o ID, removendo pontos e vírgulas
            cleaned_id = int(re.sub(r'[.,]', '', edit_id_str))
            
            # 3. Usa o ID limpo para buscar
            instance_to_edit = get_object_or_404(ReceivableAccount, id=cleaned_id, user=request.user)
        except (ValueError, TypeError):
            # 4. Se o ID for inválido (ex: "abc"), não quebra e apenas exibe uma mensagem
            messages.error(request, "O ID da conta para edição é inválido.")
            instance_to_edit = None # Garante que o formulário apareça como "novo"

    form = ReceivableAccountForm(instance=instance_to_edit, user=request.user)

    # Adicione esse bloco
    # Inicia a query base.
    accounts_query = ReceivableAccount.objects.filter(user=request.user).order_by('due_date')

    # ▼▼▼ ADICIONE ESTE NOVO BLOCO DE BUSCA ▼▼▼
    if search_query:
        accounts_query = accounts_query.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    # --- FIM DO BLOCO DE BUSCA ---

    # ▼▼▼ ADICIONE ESTE NOVO BLOCO DE FILTRO DE BANCO ▼▼▼
    if bank_filter:
        accounts_query = accounts_query.filter(bank_account__id=bank_filter)
    # --- FIM DO BLOCO DE FILTRO DE BANCO ---

    # Aplica os filtros.
    if filter_status == 'open':
        accounts_query = accounts_query.filter(is_received=False)
    elif filter_status == 'received':
        accounts_query = accounts_query.filter(is_received=True)

    if start_date:
        accounts_query = accounts_query.filter(due_date__gte=parse_date(start_date))
    if end_date:
        accounts_query = accounts_query.filter(due_date__lte=parse_date(end_date))

    # Calcula o total ANTES da paginação.
    total_amount = accounts_query.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # Aplica a paginação.
    paginator = Paginator(accounts_query, 10)
    page_number = request.GET.get('page')
    accounts_page_obj = paginator.get_page(page_number)

    user_banks = BankAccount.objects.filter(user=request.user).order_by('bank_name') # <-- NOVA LINHA

    # Monta o contexto final para o template.
    context = {
        'form': form,
        'accounts': accounts_page_obj, 
        'filter_status': filter_status,
        'start_date': start_date, # Agora as datas sempre terão um valor.
        'end_date': end_date,
        'total_amount': total_amount,
        'search_query': search_query,
        'user_banks': user_banks,     # <-- NOVA LINHA
        'bank_filter': bank_filter,   # <-- NOVA LINHA
    }
    return render(request, 'accounts/contas_receber.html', context)

# Em accounts/views.py, dentro da função dashboards

@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_painel_financeiro')
def dashboards(request):

    # --- CORREÇÃO: Define o idioma para Português do Brasil para formatação de datas ---
    try:
        # Tenta configurar o locale, mas não quebra a aplicação se falhar
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
    except locale.Error:
        # Se falhar, apenas imprime um aviso no log e continua
        print("Aviso: locale 'pt_BR.UTF-8' não encontrado. Usando o padrão do sistema.")
        pass


    # --- 1. LÓGICA DE FILTRAGEM UNIFICADA (existente, sem alterações) ---
    regime = request.GET.get('regime', 'caixa') 
    view_type = request.GET.get('view_type', 'receber')
    period = request.GET.get('period', '30')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if period == 'custom' and start_date_str and end_date_str:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
    else:
        end_date = datetime.today().date()
        days = int(period) if period in ['90', '180'] else 30
        start_date = end_date - timedelta(days=days)
        period = str(days)
    
    today = datetime.today().date()
    period_label = (end_date - start_date).days

    # ▼▼▼ ADICIONE ESTE BLOCO DE CÓDIGO AQUI ▼▼▼
    # --- LÓGICA PARA O RÓTULO DE COMPARAÇÃO DINÂMICO ---
    comparison_period_label = "período anterior" # Padrão para 'custom'
    if period == '30':
        comparison_period_label = "mês anterior"
    elif period == '90':
        comparison_period_label = "trimestre anterior"
    elif period == '180':
        comparison_period_label = "semestre anterior"
    # --- FIM DA LÓGICA DO RÓTULO ---

    # --- 2. CÁLCULOS DO DASHBOARD FINANCEIRO (existente, sem alterações) ---
    total_receivable_filt = ReceivableAccount.objects.filter(user=request.user, is_received=True, due_date__gte=start_date, due_date__lte=end_date).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    total_payable_filt = PayableAccount.objects.filter(user=request.user, is_paid=True, due_date__gte=start_date, due_date__lte=end_date).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    balance_filt = total_receivable_filt - total_payable_filt
    year_start = today.replace(month=1, day=1)
    total_receivable_year = ReceivableAccount.objects.filter(user=request.user, is_received=True, due_date__gte=year_start, due_date__lte=today).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    total_payable_year = PayableAccount.objects.filter(user=request.user, is_paid=True, due_date__gte=year_start, due_date__lte=today).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    balance_year = total_receivable_year - total_payable_year
    expected_receivable_filt = ReceivableAccount.objects.filter(user=request.user, due_date__gte=start_date, due_date__lte=end_date).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    expected_payable_filt = PayableAccount.objects.filter(user=request.user, due_date__gte=start_date, due_date__lte=end_date).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    # ▼▼▼ COLE ESTE NOVO BLOCO CORRIGIDO NO LUGAR ▼▼▼
    # --- INÍCIO DA LÓGICA CORRIGIDA PARA O CARD INSIGHTS ---

    # 1. Calcula a duração do período filtrado
    duration = end_date - start_date

    # 2. Calcula as datas do período anterior equivalente
    previous_period_end = start_date - timedelta(days=1)
    previous_period_start = previous_period_end - duration

    # 3. Calcula o total de ENTRADAS para o período ATUAL e ANTERIOR
    current_period_receivable = ReceivableAccount.objects.filter(
        user=request.user, is_received=True, due_date__range=[start_date, end_date]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    previous_period_receivable = ReceivableAccount.objects.filter(
        user=request.user, is_received=True, due_date__range=[previous_period_start, previous_period_end]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # 4. Calcula a variação das ENTRADAS
    receivable_variation = 0
    if previous_period_receivable > 0:
        receivable_variation = ((current_period_receivable - previous_period_receivable) / previous_period_receivable) * 100

    # 5. Calcula o total de SAÍDAS para o período ATUAL e ANTERIOR
    current_period_payable = PayableAccount.objects.filter(
        user=request.user, is_paid=True, due_date__range=[start_date, end_date]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    previous_period_payable = PayableAccount.objects.filter(
        user=request.user, is_paid=True, due_date__range=[previous_period_start, previous_period_end]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # 6. Calcula a variação das SAÍDAS
    payable_variation = 0
    if previous_period_payable > 0:
        payable_variation = ((current_period_payable - previous_period_payable) / previous_period_payable) * 100
    # --- FIM DA LÓGICA CORRIGIDA ---
    month_abbr_pt = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    payable_by_month_q = PayableAccount.objects.filter(user=request.user, is_paid=True, due_date__year=today.year).exclude(dre_area='NAO_CONSTAR').annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    receivable_by_month_q = ReceivableAccount.objects.filter(user=request.user, is_received=True, due_date__year=today.year).exclude(dre_area='NAO_CONSTAR').annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    payable_map = {item['month'].strftime('%b'): float(item['total']) for item in payable_by_month_q}
    receivable_map = {item['month'].strftime('%b'): float(item['total']) for item in receivable_by_month_q}
    cashflow_labels_financeiro = [month_abbr_pt[i-1] for i in range(1, 13)]
    cashflow_payable_financeiro = [payable_map.get(datetime(today.year, i, 1).strftime('%b'), 0) for i in range(1, 13)]
    cashflow_receivable_financeiro = [receivable_map.get(datetime(today.year, i, 1).strftime('%b'), 0) for i in range(1, 13)]

    # --- 3. CÁLCULOS DO DASHBOARD FLUXO DE CAIXA (COM ALTERAÇÕES) ---
    receivable_open_filt = ReceivableAccount.objects.filter(user=request.user, is_received=False, due_date__gte=start_date, due_date__lte=end_date).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    payable_open_filt = PayableAccount.objects.filter(user=request.user, is_paid=False, due_date__gte=start_date, due_date__lte=end_date).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    monthly_data_fc = {}
    receivables_by_month_fc = ReceivableAccount.objects.filter(user=request.user, is_received=True, due_date__year=today.year).annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    payables_by_month_fc = PayableAccount.objects.filter(user=request.user, is_paid=True, due_date__year=today.year).annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    for r in receivables_by_month_fc:
        monthly_data_fc.setdefault(r['month'].strftime('%b/%Y'), {'receivable': Decimal(0), 'payable': Decimal(0)})['receivable'] = r['total']
    for p in payables_by_month_fc:
        monthly_data_fc.setdefault(p['month'].strftime('%b/%Y'), {'receivable': Decimal(0), 'payable': Decimal(0)})['payable'] = p['total']
    for data in monthly_data_fc.values(): data['balance'] = data['receivable'] - data['payable']
    def get_best_worst(metric):
        if not monthly_data_fc: return '-', '-'
        valid_months = {m: d[metric] for m, d in monthly_data_fc.items() if metric in d and d[metric] != 0}
        if not valid_months: return '-', '-'
        if metric == 'payable': worst_month, best_month = max(valid_months, key=valid_months.get), min(valid_months, key=valid_months.get)
        else: best_month, worst_month = max(valid_months, key=valid_months.get), min(valid_months, key=valid_months.get)
        return best_month, worst_month
    receivable_best_month, receivable_worst_month = get_best_worst('receivable')
    payable_best_month, payable_worst_month = get_best_worst('payable')
    balance_best_month, balance_worst_month = get_best_worst('balance')
    # COLE ESTE NOVO BLOCO COMPLETO E CORRIGIDO NA SUA VIEW 'dashboards'

    # --- INÍCIO: LÓGICA REFEITA E CORRIGIDA PARA PREVISÃO VS REALIZADO ---

    # 1. PREVISÃO VS REALIZADO DE RECEITAS (com valores positivos)
    pvr_receitas_query = ReceivableAccount.objects.filter(
        user=request.user, due_date__range=[start_date, end_date]
    ).values('category__name').annotate(
        expected=Sum('amount'),
        actual=Sum('amount', filter=Q(is_received=True))
    ).order_by('-expected')

    pvr_receitas = [
        {'category': item['category__name'] or 'Sem Categoria', 
        'expected': item['expected'] or Decimal('0'), 
        'actual': item['actual'] or Decimal('0')}
        for item in pvr_receitas_query
    ]
    pvr_receitas_total = {
        'expected': sum(item['expected'] for item in pvr_receitas),
        'actual': sum(item['actual'] for item in pvr_receitas)
    }

    # 2. PREVISÃO VS REALIZADO DE DESPESAS (com valores positivos)
    pvr_despesas_query = PayableAccount.objects.filter(
        user=request.user, due_date__range=[start_date, end_date]
    ).exclude(dre_area='NAO_CONSTAR').values('category__name').annotate(
        expected=Sum('amount'),
        actual=Sum('amount', filter=Q(is_paid=True))
    ).order_by('-expected')

    pvr_despesas = [
        {'category': item['category__name'] or 'Sem Categoria', 
        'expected': item['expected'] or Decimal('0'), 
        'actual': item['actual'] or Decimal('0')}
        for item in pvr_despesas_query
    ]
    pvr_despesas_total = {
        'expected': sum(item['expected'] for item in pvr_despesas),
        'actual': sum(item['actual'] for item in pvr_despesas)
    }

    # 3. DADOS PARA OS NOVOS GRÁFICOS SEPARADOS
    pvr_receitas_chart_data = {
        'labels': [item['category'] for item in pvr_receitas],
        'expected': [float(item['expected']) for item in pvr_receitas],
        'actual': [float(item['actual']) for item in pvr_receitas]
    }
    pvr_despesas_chart_data = {
        'labels': [item['category'] for item in pvr_despesas],
        'expected': [float(item['expected']) for item in pvr_despesas],
        'actual': [float(item['actual']) for item in pvr_despesas]
    }

    # 4. RECONSTRUÇÃO DAS VARIÁVEIS ANTIGAS PARA COMPATIBILIDADE (A CORREÇÃO)
    #    Esta parte recria a variável 'category_comparison' que outras partes da view usam.
    category_comparison_data = {}
    for item in pvr_receitas:
        cat = item['category']
        category_comparison_data.setdefault(cat, {'expected': Decimal('0'), 'actual': Decimal('0')})
        category_comparison_data[cat]['expected'] += item['expected']
        category_comparison_data[cat]['actual'] += item['actual']

    for item in pvr_despesas:
        cat = item['category']
        category_comparison_data.setdefault(cat, {'expected': Decimal('0'), 'actual': Decimal('0')})
        # Subtrai os valores de despesa para que fiquem negativos, como a análise DFC espera
        category_comparison_data[cat]['expected'] -= item['expected']
        category_comparison_data[cat]['actual'] -= item['actual']

    category_comparison = [{'category': k, **v} for k, v in category_comparison_data.items()]
    pvr_total_previsto = sum(item['expected'] for item in category_comparison)
    pvr_total_realizado = sum(item['actual'] for item in category_comparison)

    # --- FIM DA NOVA LÓGICA ---
    receivable_by_category = ReceivableAccount.objects.filter(user=request.user, is_received=True, due_date__gte=start_date, due_date__lte=end_date).values('category__name').annotate(total=Sum('amount')).order_by('-total')
    payable_by_category = PayableAccount.objects.filter(user=request.user, is_paid=True, due_date__gte=start_date, due_date__lte=end_date).values('category__name').annotate(total=Sum('amount')).order_by('-total')

    # --- INÍCIO DA NOVA LÓGICA: DADOS PARA GRÁFICO DE EVOLUÇÃO (FLUXO DE CAIXA) ---
    # 1. Geração de dados DIÁRIOS (respeitando o filtro de período)
    fc_daily_labels, fc_daily_receivable, fc_daily_payable = [], [], []
    payable_by_date_q = PayableAccount.objects.filter(user=request.user, is_paid=True, due_date__gte=start_date, due_date__lte=end_date).exclude(dre_area='NAO_CONSTAR').values('due_date').annotate(total=Sum('amount')).order_by('due_date')
    receivable_by_date_q = ReceivableAccount.objects.filter(user=request.user, is_received=True, due_date__gte=start_date, due_date__lte=end_date).exclude(dre_area='NAO_CONSTAR').values('due_date').annotate(total=Sum('amount')).order_by('due_date')
    payable_map_daily = {item['due_date']: float(item['total']) for item in payable_by_date_q}
    receivable_map_daily = {item['due_date']: float(item['total']) for item in receivable_by_date_q}
    current_date_loop = start_date
    while current_date_loop <= end_date:
        fc_daily_labels.append(current_date_loop.strftime('%d/%m'))
        fc_daily_payable.append(payable_map_daily.get(current_date_loop, 0))
        fc_daily_receivable.append(receivable_map_daily.get(current_date_loop, 0))
        current_date_loop += timedelta(days=1)

    # 2. Geração de dados MENSAIS (COM FILTROS DE ANO E VISÃO)
    
    # --- Captura os filtros da URL (HTML) ---
    fc_year_str = request.GET.get('fc_year')
    fc_view = request.GET.get('fc_view', 'realizado') # Padrão: realizado

    # Define o ano (se não vier na URL, usa o ano atual)
    try:
        fc_year = int(fc_year_str) if fc_year_str else today.year
    except ValueError:
        fc_year = today.year

    # Gera rótulos para todos os 12 meses do ano SELECIONADO
    fc_monthly_labels = [datetime(fc_year, i, 1).strftime('%b/%Y') for i in range(1, 13)]

    # --- Configura os filtros base (Pagar e Receber) ---
    receivables_qs = ReceivableAccount.objects.filter(
        user=request.user, 
        due_date__year=fc_year
    ).exclude(dre_area='NAO_CONSTAR')

    payables_qs = PayableAccount.objects.filter(
        user=request.user, 
        due_date__year=fc_year
    ).exclude(dre_area='NAO_CONSTAR')

    # --- Aplica a lógica do botão Realizado vs Previsto ---
    if fc_view == 'realizado':
        # Se for realizado, filtra apenas o que foi pago/recebido
        receivables_qs = receivables_qs.filter(is_received=True)
        payables_qs = payables_qs.filter(is_paid=True)
    else:
        # Se for 'previsto', NÃO filtra por status (mostra tudo que vence no ano)
        pass 

    # --- Executa as buscas no banco de dados ---
    receivable_by_month_q_fc = receivables_qs.annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    payable_by_month_q_fc = payables_qs.annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))

    # Mapeia os resultados para facilitar a busca
    receivable_map_monthly = {item['month'].strftime('%b/%Y'): float(item['total']) for item in receivable_by_month_q_fc}
    payable_map_monthly = {item['month'].strftime('%b/%Y'): float(item['total']) for item in payable_by_month_q_fc}
    
    # Preenche as listas de dados (coloca 0 se o mês não tiver valor)
    fc_monthly_receivable = [receivable_map_monthly.get(label, 0) for label in fc_monthly_labels]
    fc_monthly_payable = [payable_map_monthly.get(label, 0) for label in fc_monthly_labels]
    
    # 1. Calcula o Saldo (Entrada - Saída) mês a mês
    saldos_mensais_tabela = [r - p for r, p in zip(fc_monthly_receivable, fc_monthly_payable)]

    # 2. Calcula os totais do ano (coluna Total)
    total_entradas_ano = sum(fc_monthly_receivable)
    total_saidas_ano = sum(fc_monthly_payable)
    total_geracao_caixa_ano = total_entradas_ano - total_saidas_ano

    # Monta o dicionário final para o template
    fluxo_caixa_tabela_data = {
        'labels': fc_monthly_labels,
        'entradas': fc_monthly_receivable,
        'saidas': fc_monthly_payable,
        'geracao_caixa': saldos_mensais_tabela,
        'totais': {
            'entradas': total_entradas_ano,
            'saidas': total_saidas_ano,
            'geracao_caixa': total_geracao_caixa_ano,
        }
    }
    
    # --- 4. CÁLCULOS DO DASHBOARD KPIS (existente, COM ALTERAÇÕES) ---
    insights_data_dre, monthly_data_dre = {}, {}
    loop_start_date = start_date.replace(day=1)
    while loop_start_date <= end_date:
        month_end = (loop_start_date + relativedelta(months=1)) - timedelta(days=1)
        if month_end > end_date: month_end = end_date
        month_key = loop_start_date.strftime('%b/%Y')
        receita_bruta = ReceivableAccount.objects.filter(user=request.user, is_received=True, dre_area='BRUTA', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        impostos = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DEDUCAO', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        custos = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='CUSTOS', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        despesas_operacionais = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='OPERACIONAL', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        nao_operacionais = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='NAO_OPERACIONAL', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        depreciacao = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DEPRECIACAO', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        tributacao = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='TRIBUTACAO', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        distribuicao_lucro = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DISTRIBUICAO', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        receita_liquida = receita_bruta - impostos
        lucro_bruto = receita_liquida - custos
        ebitda = lucro_bruto - despesas_operacionais
        ebit = ebitda - depreciacao
        lair = ebit - nao_operacionais
        lucro_liquido = lair - tributacao # Este é o Lucro Líquido real
        resultado_final = lucro_liquido - distribuicao_lucro
        monthly_data_dre[month_key] = {'receita_bruta': receita_bruta, 'impostos': impostos, 'lucro_bruto': lucro_bruto, 'ebitda': ebitda, 'lucro_liquido': lucro_liquido, 'resultado_final': resultado_final}
        custos_variaveis = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='CUSTOS', cost_type='VARIAVEL', due_date__gte=loop_start_date, due_date__lte=month_end).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        custos_fixos_total = PayableAccount.objects.filter(user=request.user, is_paid=True, cost_type='FIXO', due_date__gte=loop_start_date, due_date__lte=month_end).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        margem_contribuicao = receita_liquida - custos_variaveis
        
        # LÓGICA DO PONTO DE EQUILÍBRIO CORRIGIDA
        ponto_equilibrio_mensal = Decimal('0')
        if margem_contribuicao > 0 and receita_liquida > 0:
            margem_contribuicao_percentual = margem_contribuicao / receita_liquida
            if margem_contribuicao_percentual > 0:
                 ponto_equilibrio_mensal = custos_fixos_total / margem_contribuicao_percentual

        insights_data_dre[month_key] = {
            'margem_bruta_percentual': (lucro_bruto / receita_liquida * 100) if receita_liquida else 0, 
            'ebitda_percentual': (ebitda / receita_liquida * 100) if receita_liquida else 0, 
            'margem_liquida_percentual': (lucro_liquido / receita_liquida * 100) if receita_liquida else 0, 
            'custos': custos, 'custos_percentual': (custos / receita_liquida * 100) if receita_liquida else 0, 
            'margem_contribuicao': margem_contribuicao, 
            'margem_contribuicao_percentual': (margem_contribuicao / receita_liquida * 100) if receita_liquida else 0, 
            'ponto_equilibrio': ponto_equilibrio_mensal # Usa a variável corrigida
        }
        loop_start_date += relativedelta(months=1)

    dre_results = {}
    
    if monthly_data_dre:
        dre_results = {k: sum(d[k] for d in monthly_data_dre.values()) for k in monthly_data_dre.get(next(iter(monthly_data_dre)), {})}
        
        # Zera o dicionário de margens para garantir que não haja dados antigos
        dre_results['margens'] = {}

        if dre_results.get('receita_bruta', 0) > 0:
            receita_bruta_total = dre_results['receita_bruta']
            # Dicionário de margens agora completo
            dre_results['margens'] = {
                'receita_bruta': 100.00,  # A margem da receita bruta sobre ela mesma é 100%
                'impostos': (dre_results.get('impostos', 0) / receita_bruta_total) * 100,
                'lucro_bruto': (dre_results.get('lucro_bruto', 0) / receita_bruta_total) * 100,
                'ebitda': (dre_results.get('ebitda', 0) / receita_bruta_total) * 100,
                'lucro_liquido': (dre_results.get('lucro_liquido', 0) / receita_bruta_total) * 100
            }
# ...
        # NOVO CÁLCULO: PONTO DE EQUILÍBRIO PARA O PERÍODO TODO (PARA O CARD)
        total_receita_liquida_periodo = dre_results.get('receita_bruta', Decimal('0')) - dre_results.get('impostos', Decimal('0'))
        total_custos_variaveis_periodo = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='CUSTOS', cost_type='VARIAVEL', due_date__gte=start_date, due_date__lte=end_date).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        total_custos_fixos_periodo = PayableAccount.objects.filter(user=request.user, is_paid=True, cost_type='FIXO', due_date__gte=start_date, due_date__lte=end_date).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        
        total_margem_contribuicao = total_receita_liquida_periodo - total_custos_variaveis_periodo
        
        ponto_equilibrio_total = Decimal('0')
        if total_margem_contribuicao > 0 and total_receita_liquida_periodo > 0:
            total_margem_contribuicao_percentual = total_margem_contribuicao / total_receita_liquida_periodo
            if total_margem_contribuicao_percentual > 0:
                ponto_equilibrio_total = total_custos_fixos_periodo / total_margem_contribuicao_percentual
        
        dre_results['ponto_equilibrio'] = ponto_equilibrio_total # Adiciona ao dicionário
        
        dre_results['melhores_meses'], dre_results['piores_meses'] = {}, {}
        for metric in ['receita_bruta', 'impostos', 'lucro_bruto', 'ebitda', 'lucro_liquido']:
            sorted_months = sorted([(m, d[metric]) for m, d in monthly_data_dre.items()], key=lambda x: x[1])
            dre_results['piores_meses'][metric] = sorted_months[0] if sorted_months else (None, 0); dre_results['melhores_meses'][metric] = sorted_months[-1] if sorted_months else (None, 0)

    # --- INÍCIO: LÓGICA PARA GRÁFICO DE INDICADORES CHAVE COM 12 MESES FIXOS ---
    current_year = today.year
    
    # 1. Gera rótulos e estrutura de dados para os 12 meses do ano corrente
    insights_labels_kpis = [datetime(current_year, i, 1).strftime('%b/%Y') for i in range(1, 13)]
    insights_data_anual = OrderedDict((month, {}) for month in insights_labels_kpis)

    # 2. Calcula os indicadores para cada mês do ano, ignorando filtros de data
    for month_key in insights_labels_kpis:
        month_date = datetime.strptime(month_key, '%b/%Y')
        start_of_month = month_date.replace(day=1)
        end_of_month = (start_of_month + relativedelta(months=1)) - timedelta(days=1)

        # Busca os dados financeiros apenas para o mês do loop
        receita_bruta = ReceivableAccount.objects.filter(user=request.user, is_received=True, dre_area='BRUTA', due_date__gte=start_of_month, due_date__lte=end_of_month).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        impostos = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DEDUCAO', due_date__gte=start_of_month, due_date__lte=end_of_month).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        custos = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='CUSTOS', due_date__gte=start_of_month, due_date__lte=end_of_month).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        despesas_operacionais = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='OPERACIONAL', due_date__gte=start_of_month, due_date__lte=end_of_month).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        nao_operacionais = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='NAO_OPERACIONAL', due_date__gte=start_of_month, due_date__lte=end_of_month).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        custos_variaveis = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='CUSTOS', cost_type='VARIAVEL', due_date__gte=start_of_month, due_date__lte=end_of_month).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        custos_fixos_total = PayableAccount.objects.filter(user=request.user, is_paid=True, cost_type='FIXO', due_date__gte=start_of_month, due_date__lte=end_of_month).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

        # Cálculos dos indicadores
        receita_liquida = receita_bruta - impostos
        lucro_bruto = receita_liquida - custos
        ebitda = lucro_bruto - despesas_operacionais
        lucro_liquido = ebitda - nao_operacionais
        margem_contribuicao = receita_liquida - custos_variaveis
        ponto_equilibrio_mensal = Decimal('0')
        if margem_contribuicao > 0 and receita_liquida > 0:
            margem_contribuicao_percentual = margem_contribuicao / receita_liquida
            if margem_contribuicao_percentual > 0:
                 ponto_equilibrio_mensal = custos_fixos_total / margem_contribuicao_percentual

        # Armazena os indicadores do mês
        insights_data_anual[month_key] = {
            'margem_bruta_percentual': (lucro_bruto / receita_liquida * 100) if receita_liquida else 0, 
            'ebitda_percentual': (ebitda / receita_liquida * 100) if receita_liquida else 0, 
            'margem_liquida_percentual': (lucro_liquido / receita_liquida * 100) if receita_liquida else 0, 
            'custos': custos, 'custos_percentual': (custos / receita_liquida * 100) if receita_liquida else 0, 
            'margem_contribuicao': margem_contribuicao, 
            'margem_contribuicao_percentual': (margem_contribuicao / receita_liquida * 100) if receita_liquida else 0, 
            'ponto_equilibrio': ponto_equilibrio_mensal
        }

    # 3. Gera os dados finais para o Chart.js a partir da estrutura anual
    first_month_key = insights_labels_kpis[0] if insights_labels_kpis else None
    insights_charts_data_kpis = {}
    if first_month_key:
        # Pega as chaves de indicadores do primeiro mês como referência
        indicadores = insights_data_anual[first_month_key].keys()
        insights_charts_data_kpis = {
            key: [float(insights_data_anual.get(m, {}).get(key, 0)) for m in insights_labels_kpis]
            for key in indicadores
        }
    # --- INÍCIO: LÓGICA ATUALIZADA PARA GRÁFICO WATERFALL DE LUCRO LÍQUIDO ---
    current_year = today.year
    waterfall_labels, waterfall_data, waterfall_bg_colors = [], [], []
    cumulative_profit = Decimal('0')

    # 1. Gera rótulos para os 12 meses (sem alterações)
    monthly_profit_map = OrderedDict()
    for i in range(1, 13):
        month_key = f"{MESES_ABREVIADOS[i]}/{current_year}"
        monthly_profit_map[month_key] = Decimal('0')

    # 2. Busca TODOS os componentes anuais da DRE (base para competência)
    receitas_anuais = ReceivableAccount.objects.filter(user=request.user, dre_area='BRUTA', due_date__year=current_year).annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    impostos_anuais = PayableAccount.objects.filter(user=request.user, dre_area='DEDUCAO', due_date__year=current_year).annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    custos_anuais = PayableAccount.objects.filter(user=request.user, dre_area='CUSTOS', due_date__year=current_year).annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    despesas_op_anuais = PayableAccount.objects.filter(user=request.user, dre_area='OPERACIONAL', due_date__year=current_year).annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))
    nao_op_anuais = PayableAccount.objects.filter(user=request.user, dre_area='NAO_OPERACIONAL', due_date__year=current_year).annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount'))

    # ---> NOVO BLOCO CONDICIONAL <---
    # Este é o "coração" da mudança. Ele aplica o filtro de 'caixa' apenas se for o regime selecionado.
    if regime == 'caixa':
        receitas_anuais = receitas_anuais.filter(is_received=True)
        impostos_anuais = impostos_anuais.filter(is_paid=True)
        custos_anuais = custos_anuais.filter(is_paid=True)
        despesas_op_anuais = despesas_op_anuais.filter(is_paid=True)
        nao_op_anuais = nao_op_anuais.filter(is_paid=True)

    # 3. Mapeia os totais mensais (sem alterações)
    receitas_map = {f"{MESES_ABREVIADOS[r['month'].month]}/{r['month'].year}": r['total'] for r in receitas_anuais}
    impostos_map = {f"{MESES_ABREVIADOS[i['month'].month]}/{i['month'].year}": i['total'] for i in impostos_anuais}
    custos_map = {f"{MESES_ABREVIADOS[c['month'].month]}/{c['month'].year}": c['total'] for c in custos_anuais}
    despesas_op_map = {f"{MESES_ABREVIADOS[d['month'].month]}/{d['month'].year}": d['total'] for d in despesas_op_anuais}
    nao_op_map = {f"{MESES_ABREVIADOS[n['month'].month]}/{n['month'].year}": n['total'] for n in nao_op_anuais}

    # 4. Calcula o LUCRO LÍQUIDO para cada mês do ano (sem alterações)
    for month_key in monthly_profit_map.keys():
        receita_bruta = receitas_map.get(month_key, Decimal('0'))
        impostos = impostos_map.get(month_key, Decimal('0'))
        custos = custos_map.get(month_key, Decimal('0'))
        despesas_operacionais = despesas_op_map.get(month_key, Decimal('0'))
        nao_operacionais = nao_op_map.get(month_key, Decimal('0'))
        
        receita_liquida = receita_bruta - impostos
        lucro_bruto = receita_liquida - custos
        ebitda = lucro_bruto - despesas_operacionais
        lucro_liquido = ebitda - nao_operacionais
        
        monthly_profit_map[month_key] = lucro_liquido

    # 5. Gera os dados para o gráfico (sem alterações)
    for month_key, profit in monthly_profit_map.items():
        waterfall_labels.append(month_key.split('/')[0])
        waterfall_data.append([float(cumulative_profit), float(cumulative_profit + profit)])
        waterfall_bg_colors.append('green' if profit >= 0 else 'red')
        cumulative_profit += profit
    
    waterfall_labels.append('Total')
    waterfall_data.append([0, float(cumulative_profit)])
    waterfall_bg_colors.append("#2F88CD")
    # --- FIM: LÓGICA ATUALIZADA PARA GRÁFICO WATERFALL ---

    # --- 5. CÁLCULOS DA ANÁLISE DE CONTAS (LÓGICA ATUALIZADA) ---
    insights_data_analise = {}
    analysis_data = {}
    today = timezone.now().date()  # Garante que 'today' está definido aqui

    if view_type == 'pagar':
        # Bloco para Contas a Pagar
        contas_em_aberto_geral = PayableAccount.objects.filter(user=request.user, is_paid=False)
        total_atrasado = contas_em_aberto_geral.filter(due_date__lt=today).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_a_vencer = contas_em_aberto_geral.filter(due_date__gte=today).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        # ... (cálculos de insights que não mudam)
        contas_pagas = PayableAccount.objects.filter(user=request.user, is_paid=True, payment_date__isnull=False)
        pagos_em_atraso_q = contas_pagas.filter(payment_date__gt=F('due_date'))
        percentual_pago_atrasado = (pagos_em_atraso_q.count() / contas_pagas.count() * 100) if contas_pagas.count() > 0 else "N/A"
        dias_para_pagar_list = [(c.payment_date - c.due_date).days for c in contas_pagas]
        media_dias_pagamento = sum(dias_para_pagar_list) / len(dias_para_pagar_list) if dias_para_pagar_list else "N/A"
        taxa_atraso = (total_atrasado / (PayableAccount.objects.filter(user=request.user).aggregate(total=Sum('amount'))['total'] or Decimal('1'))) * 100
        insights_data_analise = {'total_a_vencer': total_a_vencer, 'total_atrasado': total_atrasado, 'percentual_pago_atrasado': percentual_pago_atrasado, 'media_dias_pagamento': media_dias_pagamento, 'taxa_inadimplencia': taxa_atraso}
        supplier_open_amounts = contas_em_aberto_geral.values('name').annotate(total_open=Sum('amount')).order_by('-total_open')
        
        # --- INÍCIO DA NOVA LÓGICA DE ENVELHECIMENTO DA DÍVIDA ---
        contas_atrasadas_geral = contas_em_aberto_geral.filter(due_date__lt=today)
        overdue_1_to_30_days = contas_atrasadas_geral.filter(due_date__gte=today - timedelta(days=30)).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        overdue_31_to_60_days = contas_atrasadas_geral.filter(due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        overdue_61_to_90_days = contas_atrasadas_geral.filter(due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        overdue_over_90_days = contas_atrasadas_geral.filter(due_date__lt=today - timedelta(days=90)).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        analysis_data = {
            'client_open_amount_labels': [item['name'] or 'Fornecedor' for item in supplier_open_amounts],
            'client_open_amount_data': [float(item['total_open']) for item in supplier_open_amounts],
            'open_vs_overdue_data': [float(total_a_vencer), float(total_atrasado)],
            'overdue_by_period_labels': ['1 a 30 dias', '31 a 60 dias', '61 a 90 dias', 'Acima de 90 dias'],
            'overdue_by_period_data': [float(overdue_1_to_30_days), float(overdue_31_to_60_days), float(overdue_61_to_90_days), float(overdue_over_90_days)]
        }

    else: # receber
        # Bloco para Contas a Receber
        contas_em_aberto_geral = ReceivableAccount.objects.filter(user=request.user, is_received=False)
        total_atrasado = contas_em_aberto_geral.filter(due_date__lt=today).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_a_vencer = contas_em_aberto_geral.filter(due_date__gte=today).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        # ... (cálculos de insights que não mudam)
        contas_recebidas = ReceivableAccount.objects.filter(user=request.user, is_received=True, payment_date__isnull=False)
        recebidos_em_atraso_q = contas_recebidas.filter(payment_date__gt=F('due_date'))
        percentual_pago_atrasado = (recebidos_em_atraso_q.count() / contas_recebidas.count() * 100) if contas_recebidas.count() > 0 else "N/A"
        dias_para_receber_list = [(c.payment_date - c.due_date).days for c in contas_recebidas]
        media_dias_pagamento = sum(dias_para_receber_list) / len(dias_para_receber_list) if dias_para_receber_list else "N/A"
        taxa_inadimplencia = (total_atrasado / (ReceivableAccount.objects.filter(user=request.user).aggregate(total=Sum('amount'))['total'] or Decimal('1'))) * 100
        insights_data_analise = {'total_a_vencer': total_a_vencer, 'total_atrasado': total_atrasado, 'percentual_pago_atrasado': percentual_pago_atrasado, 'media_dias_pagamento': media_dias_pagamento, 'taxa_inadimplencia': taxa_inadimplencia}
        client_open_amounts = contas_em_aberto_geral.values('name').annotate(total_open=Sum('amount')).order_by('-total_open')

        # --- INÍCIO DA NOVA LÓGICA DE ENVELHECIMENTO DA DÍVIDA ---
        contas_atrasadas_geral = contas_em_aberto_geral.filter(due_date__lt=today)
        overdue_1_to_30_days = contas_atrasadas_geral.filter(due_date__gte=today - timedelta(days=30)).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        overdue_31_to_60_days = contas_atrasadas_geral.filter(due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        overdue_61_to_90_days = contas_atrasadas_geral.filter(due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        overdue_over_90_days = contas_atrasadas_geral.filter(due_date__lt=today - timedelta(days=90)).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        # Em accounts/views.py -> dashboards -> else: # receber:

        analysis_data = {
            'client_open_amount_labels': [item['name'] or 'Cliente' for item in client_open_amounts],
            'client_open_amount_data': [float(item['total_open']) for item in client_open_amounts],
            'open_vs_overdue_data': [float(total_a_vencer), float(total_atrasado)],
            'overdue_by_period_labels': ['1 a 30 dias', '31 a 60 dias', '61 a 90 dias', 'Acima de 90 dias'],
            'overdue_by_period_data': [float(overdue_1_to_30_days), float(overdue_31_to_60_days), float(overdue_61_to_90_days), float(overdue_over_90_days)]
        }
        # --- FIM DA NOVA LÓGICA ---

    # vvv ADICIONE ESTE BLOCO ABAIXO vvv
    # --- LÓGICA PARA O TERMÔMETRO DE INSIGHTS ---
    # Define uma escala: -25% de variação (ou pior) será 0% no termômetro.
    # +25% de variação (ou melhor) será 100% no termômetro.
    min_variation = -25.0
    max_variation = 25.0

    # Converte o valor da variação para a escala de 0 a 100
    try:
        # Garante que a variação é um float para o cálculo
        current_variation = float(receivable_variation)

        # Fórmula de normalização
        percentage = ((current_variation - min_variation) / (max_variation - min_variation)) * 100

        # Garante que o resultado fique sempre entre 0 e 100
        thermometer_percentage = max(0, min(100, percentage))
    except (ValueError, TypeError):
        thermometer_percentage = 50 # Valor padrão (meio) em caso de erro

    # --- FIM DA LÓGICA DO TERMÔMETRO ---

    

    # --- NOVOS CÁLCULOS PARA GRÁFICOS DA SEÇÃO DRE ---

    # 1. Gráfico Top 5 Despesas por Categoria (no período filtrado)
    top_despesas_query = PayableAccount.objects.filter(
        user=request.user,
        is_paid=True,
        due_date__range=[start_date, end_date]
    ).exclude(
        # Exclui itens que não são despesas operacionais típicas
        Q(dre_area='NAO_CONSTAR') | Q(category__name__iexact='Retirada sócio')
    ).values('category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')[:5] # Pega os 5 maiores

    top_despesas_labels = [item['category__name'] or 'Sem Categoria' for item in top_despesas_query]
    top_despesas_data = [float(item['total']) for item in top_despesas_query]


    # 2. Gráfico de Retirada de Sócios (no período filtrado)
    # IMPORTANTE: Confirme se o nome da categoria é exatamente "Retirada sócio". Se for diferente, ajuste abaixo.
    retirada_socios_query = PayableAccount.objects.filter(
        user=request.user,
        is_paid=True,
        due_date__range=[start_date, end_date],
        category__name__iexact='Retirada sócio'
    ).values('name').annotate( # Agrupa pelo campo 'name', que deve ter o nome do sócio
        total=Sum('amount')
    ).order_by('-total')

    retirada_socios_labels = [item['name'] or 'Sócio não identificado' for item in retirada_socios_query]
    retirada_socios_data = [float(item['total']) for item in retirada_socios_query]

    # --- FIM DOS NOVOS CÁLCULOS ---

    # Em views.py, dentro da view dashboards, antes da definição do 'context'

    # --- INÍCIO: CÁLCULOS PARA OS NOVOS CARDS DE KPI ---

    # 1. Buscamos os valores base da DRE para o período filtrado
    lucro_liquido_final_kpi = dre_results.get('lucro_liquido', Decimal('0'))
    
    # Precisamos recalcular a distribuição de lucro para o período, pois não está no `dre_results`
    distribuicao_lucro_kpi = PayableAccount.objects.filter(
        user=request.user, is_paid=True, dre_area='DISTRIBUICAO', due_date__range=[start_date, end_date]
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    # O lucro ANTES da distribuição é o lucro líquido final SOMADO de volta a distribuição
    lucro_antes_distribuicao_kpi = lucro_liquido_final_kpi + distribuicao_lucro_kpi

    # 2. Calculamos os indicadores possíveis
    payout_ratio = (distribuicao_lucro_kpi / lucro_antes_distribuicao_kpi * 100) if lucro_antes_distribuicao_kpi > 0 else Decimal('0')
    lucros_retidos_ratio = 100 - payout_ratio

    # 3. Montamos o dicionário que será enviado ao template
    new_kpi_cards = {
        'roi': {'value': 'Não aplicável', 'description': 'Requer o valor total do Investimento.', 'is_na': True},
        'roa': {'value': 'Não aplicável', 'description': 'Requer o valor dos Ativos Totais.', 'is_na': True},
        'roe': {'value': 'Não aplicável', 'description': 'Requer o valor do Patrimônio Líquido.', 'is_na': True},
        'cac': {'value': 'Não aplicável', 'description': 'Requer dados detalhados de Marketing e Vendas.', 'is_na': True},
        'payout_ratio': {
            'value': f"{payout_ratio:.2f}%",
            'description': 'Percentual do lucro distribuído aos sócios.',
            'is_na': False
        },
        'lucros_retidos': {
            'value': f"{lucros_retidos_ratio:.2f}%",
            'description': 'Percentual do lucro reinvestido na empresa.',
            'is_na': False
        },
        'divida_ebitda': {'value': 'Não aplicável', 'description': 'Requer o valor da Dívida Líquida.', 'is_na': True},
        'liquidez': {'value': 'Não aplicável', 'description': 'Requer dados do Balanço Patrimonial.', 'is_na': True},
        'taxa_crescimento_receita': {
            'value': f"{receivable_variation:.2f}%",
            'description': f"Crescimento da receita em relação ao {comparison_period_label}.",
            'is_na': False
        },
        'geracao_caixa': {
            'value': f"R$ {balance_filt:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            'description': 'Caixa gerado no período, disponível para operações ou distribuição.',
            'is_na': False
        },
    }
    # --- FIM: CÁLCULOS PARA OS NOVOS CARDS DE KPI ---

    # --- INÍCIO: CÁLCULOS PARA OS INSIGHTS DOS KPIS ---
    
    # 1. Calcula as datas do período anterior para comparação
    duration = end_date - start_date
    previous_period_end = start_date - timedelta(days=1)
    previous_period_start = previous_period_end - duration

    # 2. Busca os dados financeiros do período anterior
    receita_bruta_anterior = ReceivableAccount.objects.filter(user=request.user, is_received=True, dre_area='BRUTA', due_date__range=[previous_period_start, previous_period_end]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    impostos_anterior = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DEDUCAO', due_date__range=[previous_period_start, previous_period_end]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    custos_anterior = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='CUSTOS', due_date__range=[previous_period_start, previous_period_end]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    despesas_op_anterior = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='OPERACIONAL', due_date__range=[previous_period_start, previous_period_end]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    # 3. Calcula os KPIs do período anterior
    receita_liquida_anterior = receita_bruta_anterior - impostos_anterior
    lucro_bruto_anterior = receita_liquida_anterior - custos_anterior
    ebitda_anterior = lucro_bruto_anterior - despesas_op_anterior

    margem_bruta_anterior = (lucro_bruto_anterior / receita_liquida_anterior * 100) if receita_liquida_anterior else 0
    ebitda_percentual_anterior = (ebitda_anterior / receita_liquida_anterior * 100) if receita_liquida_anterior else 0

    # 4. Cria um dicionário com os dados para a análise
    kpi_analysis_data = {
        'current_margem_bruta': float(dre_results.get('margens', {}).get('lucro_bruto', 0)),
        'previous_margem_bruta': float(margem_bruta_anterior),
        'current_ebitda_percentual': float(dre_results.get('margens', {}).get('ebitda', 0)),
        'previous_ebitda_percentual': float(ebitda_percentual_anterior),
        'period_label': period_label,
        'comparison_period_label': comparison_period_label,
    }
    # --- FIM: CÁLCULOS PARA OS INSIGHTS DOS KPIS ---
    # --- INÍCIO: CÁLCULOS PARA O CARD DE VALUATION ---
    today = timezone.now().date()
    start_date_ltm = today - relativedelta(years=1) # LTM = Last Twelve Months (Últimos 12 Meses)

    # 1. Calcula os componentes do EBITDA dos últimos 12 meses
    # 1. Calcula os componentes do EBITDA dos últimos 12 meses (AGORA POR COMPETÊNCIA)
    receita_bruta_ltm = ReceivableAccount.objects.filter(user=request.user, dre_area='BRUTA', due_date__range=[start_date_ltm, today]).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    impostos_ltm = PayableAccount.objects.filter(user=request.user, dre_area='DEDUCAO', due_date__range=[start_date_ltm, today]).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    custos_ltm = PayableAccount.objects.filter(user=request.user, dre_area='CUSTOS', due_date__range=[start_date_ltm, today]).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    despesas_op_ltm = PayableAccount.objects.filter(user=request.user, dre_area='OPERACIONAL', due_date__range=[start_date_ltm, today]).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # 2. Calcula o EBITDA LTM
    receita_liquida_ltm = receita_bruta_ltm - impostos_ltm
    lucro_bruto_ltm = receita_liquida_ltm - custos_ltm
    ebitda_ltm = lucro_bruto_ltm - despesas_op_ltm

    # 3. Define um múltiplo padrão e calcula a estimativa de valuation
    multiplo_ebitda = 6.0 # Um múltiplo conservador e comum para PMEs
    valuation_ebitda = ebitda_ltm * Decimal(multiplo_ebitda)
    # --- FIM: CÁLCULOS PARA O CARD DE VALUATION --


    # --- INÍCIO: CÁLCULO BASE PARA O FLUXO DE CAIXA DESCONTADO (DCF) ---
    # Reutiliza o período de 12 meses (LTM) já definido
    entradas_realizadas_ltm = ReceivableAccount.objects.filter(
        user=request.user, is_received=True, due_date__range=[start_date_ltm, today]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    saidas_realizadas_ltm = PayableAccount.objects.filter(
        user=request.user, is_paid=True, due_date__range=[start_date_ltm, today]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # Geração de Caixa (Fluxo de Caixa Livre) dos últimos 12 meses
    geracao_caixa_ltm = entradas_realizadas_ltm - saidas_realizadas_ltm
    # --- FIM DO CÁLCULO BASE ---

    # Geração de Caixa (Fluxo de Caixa Livre) dos últimos 12 meses
    geracao_caixa_ltm = entradas_realizadas_ltm - saidas_realizadas_ltm

    # ▼▼▼ ADICIONE ESTA NOVA LINHA AQUI ▼▼▼
    geracao_caixa_ltm_js = f"{geracao_caixa_ltm:.2f}".replace(',', '.')

    # Em views.py, na função dashboards
    # --- INÍCIO: DADOS PARA O SELETOR DE SETOR DO VALUATION ---
    setores_valuation_raw = [
        {'nome': 'Selecione um setor para uma sugestão...', 'multiplo': multiplo_ebitda}, 
        {'nome': 'Tecnologia (Software, SaaS)', 'multiplo': 8.5},
        {'nome': 'Varejo (Comércio em geral)', 'multiplo': 5.5},
        {'nome': 'Serviços (Consultoria, Agências, etc.)', 'multiplo': 5.0},
        {'nome': 'Saúde (Clínicas, Laboratórios)', 'multiplo': 7.0},
        {'nome': 'Indústria (Manufatura)', 'multiplo': 6.5},
        {'nome': 'Construção Civil', 'multiplo': 5.0},
        {'nome': 'Alimentação (Restaurantes, Bares)', 'multiplo': 4.5},
        {'nome': 'Educação', 'multiplo': 7.5},
    ]
    
    # Adiciona uma chave '_js' formatada com ponto para cada setor
    setores_valuation = [
        {**setor, 'multiplo_js': f"{setor['multiplo']:.1f}".replace(',', '.')}
        for setor in setores_valuation_raw
    ]
   

    # --- INÍCIO: NOVOS DADOS CORRIGIDOS PARA A ANÁLISE DO FLUXO DE CAIXA (DFC) ---
    
    # 1. Pega os dados principais do período atual que já foram calculados
    entradas_atuais = total_receivable_filt
    saidas_atuais = total_payable_filt
    saldo_atual = balance_filt

    # 2. Calcula o saldo do período anterior
    saldo_anterior = previous_period_receivable - previous_period_payable

    # 3. Identifica as top 3 categorias de Receita e Despesa do período
    # Filtra para excluir categorias de despesa da lista de receitas
    top_receitas = sorted(
        [item for item in category_comparison if item.get('expected', 0) > 0 and item.get('actual', 0) > 0], 
        key=lambda x: x['actual'], 
        reverse=True
    )[:3]
    
    # Filtra para pegar apenas categorias de despesa
    top_despesas = sorted(
        [item for item in category_comparison if item.get('expected', 0) < 0 or item.get('actual', 0) < 0], 
        key=lambda x: x['actual']
    )[:3]
    
    # 4. Analisa a tendência mensal do ano (LÓGICA CORRIGIDA)
    # Calcula os saldos mensais a partir das listas que já existem
    saldos_mensais = [r - p for r, p in zip(fc_monthly_receivable, fc_monthly_payable)]
    meses_positivos = sum(1 for saldo in saldos_mensais if saldo > 0)
    meses_negativos = sum(1 for saldo in saldos_mensais if saldo < 0)

    

    # 5. Agrupa todos os dados em um dicionário para enviar ao JavaScript
    dfc_analysis_data = {
        'entradas_atuais': float(entradas_atuais),
        'saidas_atuais': float(saidas_atuais),
        'saldo_atual': float(saldo_atual),
        'saldo_anterior': float(saldo_anterior),
        'top_receitas': [{'category': item['category'], 'actual': float(item['actual'])} for item in top_receitas], # <-- CORRIGIDO AQUI
        'top_despesas': [{'category': item['category'], 'actual': float(abs(item['actual']))} for item in top_despesas], # <-- E CORRIGIDO AQUI
        'meses_positivos': meses_positivos,
        'meses_negativos': meses_negativos,
        'total_meses': len(saldos_mensais),
    }

    # --- 6. CONTEXTO FINAL UNIFICADO ---
    context = {
        'period': period, 'start_date': start_date.strftime('%Y-%m-%d'), 'end_date': end_date.strftime('%Y-%m-%d'), 'view_type': view_type,
        # Financeiro
        'period_label': period_label, 'total_receivable': total_receivable_filt, 'expected_receivable': expected_receivable_filt,
        'total_payable': total_payable_filt, 'expected_payable': expected_payable_filt, 'balance': balance_filt,
        'total_receivable_year': total_receivable_year, 'total_payable_year': total_payable_year, 'balance_year': balance_year,
        'receivable_variation': receivable_variation, 'payable_variation': 0, 
        'payable_variation': payable_variation,
        'comparison_period_label': comparison_period_label,
        'cashflow_labels': json.dumps(cashflow_labels_financeiro), 'cashflow_receivable': json.dumps(cashflow_receivable_financeiro), 'cashflow_payable': json.dumps(cashflow_payable_financeiro),
        # Fluxo de Caixa
        'receivable_difference': total_receivable_filt - expected_receivable_filt, 'payable_difference': total_payable_filt - expected_payable_filt,
        'expected_balance': expected_receivable_filt - expected_payable_filt, 'balance_difference': balance_filt - (expected_receivable_filt - expected_payable_filt),
        'receivable_open': receivable_open_filt, 'payable_open': payable_open_filt, 'balance_open': receivable_open_filt - payable_open_filt,
        'receivable_trend': total_receivable_filt + receivable_open_filt, 'payable_trend': total_payable_filt + payable_open_filt, 'balance_trend': (total_receivable_filt + receivable_open_filt) - (total_payable_filt + payable_open_filt),
        'receivable_best_month': receivable_best_month, 'receivable_worst_month': receivable_worst_month,
        'payable_best_month': payable_best_month, 'payable_worst_month': payable_worst_month,
        'balance_best_month': balance_best_month, 'balance_worst_month': balance_worst_month,
        'category_comparison': category_comparison,
        'receivable_category_labels': json.dumps([item['category__name'] or 'Sem Categoria' for item in receivable_by_category]),
        'receivable_category_data': json.dumps([float(item['total']) for item in receivable_by_category]),
        'payable_category_labels': json.dumps([item['category__name'] or 'Sem Categoria' for item in payable_by_category]),
        'payable_category_data': json.dumps([float(item['total']) for item in payable_by_category]),
        'thermometer_percentage': thermometer_percentage,
        'new_kpi_cards': new_kpi_cards,
        # --- ADIÇÃO DAS NOVAS VARIÁVEIS DE CONTEXTO ---
        'fc_daily_labels': json.dumps(fc_daily_labels),
        'fc_daily_receivable': json.dumps(fc_daily_receivable),
        'fc_daily_payable': json.dumps(fc_daily_payable),
        'fc_monthly_labels': json.dumps(fc_monthly_labels),
        'fc_monthly_receivable': json.dumps(fc_monthly_receivable),
        'fc_monthly_payable': json.dumps(fc_monthly_payable),
        'current_period_receivable': current_period_receivable,
        'previous_period_receivable': previous_period_receivable,
        'current_period_payable': current_period_payable,
        'previous_period_payable': previous_period_payable,
        'top_despesas_labels_json': json.dumps(top_despesas_labels),
        'top_despesas_data_json': json.dumps(top_despesas_data),
        'retirada_socios_labels_json': json.dumps(retirada_socios_labels),
        'retirada_socios_data_json': json.dumps(retirada_socios_data),
        'regime': regime,

        # KPIs
        'dre': dre_results, 'insights_labels': json.dumps(insights_labels_kpis), 'insights_charts_data': json.dumps(insights_charts_data_kpis),
        'waterfall_labels': json.dumps(waterfall_labels), 'waterfall_data': json.dumps(waterfall_data), 'waterfall_bg_colors': json.dumps(waterfall_bg_colors),
        # Análise de Contas
        'kpi_analysis_data': json.dumps(kpi_analysis_data),
        'ebitda_ltm': ebitda_ltm,
        'ebitda_ltm_js': f"{ebitda_ltm:.2f}".replace(',', '.'),          # <-- NOVA LINHA
        'multiplo_ebitda': multiplo_ebitda,
        'multiplo_ebitda_js': f"{multiplo_ebitda:.1f}".replace(',', '.'),# <-- NOVA LINHA
        'valuation_ebitda': valuation_ebitda,
        'setores_valuation': setores_valuation,
        'pvr_total_previsto': pvr_total_previsto,
        'pvr_total_realizado': pvr_total_realizado,
        'dfc_analysis_data_json': json.dumps(dfc_analysis_data),
        'fluxo_caixa_tabela_data': fluxo_caixa_tabela_data,
        'geracao_caixa_ltm': geracao_caixa_ltm,
        'geracao_caixa_ltm_js': geracao_caixa_ltm_js,
        'pvr_receitas': pvr_receitas,
        'pvr_receitas_total': pvr_receitas_total,
        'pvr_despesas': pvr_despesas,
        'pvr_despesas_total': pvr_despesas_total,
        'pvr_receitas_chart_json': json.dumps(pvr_receitas_chart_data),
        'pvr_despesas_chart_json': json.dumps(pvr_despesas_chart_data),
        'insights_data': insights_data_analise,
        **analysis_data
        
        
    }
    
    return render(request, 'accounts/dashboards.html', context)




@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_painel_vendas')
def faturamento_dashboard_view(request):

    # --- CORREÇÃO: Define o idioma para Português do Brasil para formatação de datas ---
    try:
        # Tenta configurar o locale, mas não quebra a aplicação se falhar
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
    except locale.Error:
        # Se falhar, apenas imprime um aviso no log e continua
        print("Aviso: locale 'pt_BR.UTF-8' não encontrado. Usando o padrão do sistema.")
        pass


    df = pd.DataFrame()
    # --- BASE DE DADOS ---
    # `all_sales` é usado para gráficos que mostram o histórico completo (Ex: Vendas por Mês/Ano)
    all_sales = Venda.objects.filter(user=request.user)
    
    # --- LÓGICA PARA O GRÁFICO DE FATURAMENTO E INSIGHTS (TOPO) ---
    faturamento_periodo = request.GET.get('faturamento_periodo', 'semanal')
    today = timezone.now().date()
    
    if faturamento_periodo == 'mensal':
        period_days = 30
        period_text = "últimos 30 dias"
        prev_period_text = "30 dias anteriores"
        current_start_date = today - timedelta(days=period_days - 1)
        prev_end_date = current_start_date - timedelta(days=1)
        prev_start_date = prev_end_date - timedelta(days=period_days - 1)
        
        chart_sales_query = all_sales.filter(data_venda__date__gte=current_start_date).values('data_venda__date').annotate(total=Sum('valor_total_liquido')).order_by('data_venda__date')
        sales_by_day = {item['data_venda__date']: float(item['total']) for item in chart_sales_query}
        faturamento_chart_labels = [(current_start_date + timedelta(days=i)).strftime('%d/%m') for i in range(period_days)]
        faturamento_chart_data = [sales_by_day.get(current_start_date + timedelta(days=i), 0) for i in range(period_days)]

    elif faturamento_periodo == 'anual':
        period_text = "ano corrente"
        ano_anterior = today.year - 1
        prev_period_text = f"ano de {ano_anterior}"
        
        # Define o período atual como o ano corrente
        ano_corrente = today.year
        current_start_date = today.replace(year=ano_corrente, month=1, day=1)
        
        # Define o período anterior como o ano anterior completo
        prev_start_date = current_start_date - relativedelta(years=1)
        prev_end_date = current_start_date - timedelta(days=1)
        
        # 1. Cria os rótulos de Jan a Dez do ano corrente
        faturamento_chart_labels = [datetime(ano_corrente, i, 1).strftime('%b/%Y') for i in range(1, 13)]
        
        # 2. Busca as vendas do ano corrente
        chart_sales_query = all_sales.filter(data_venda__year=ano_corrente).annotate(
            month=TruncMonth('data_venda')
        ).values('month').annotate(total=Sum('valor_total_liquido')).order_by('month')
        
        # 3. Mapeia os resultados
        sales_by_month = {v['month'].strftime('%b/%Y'): float(v['total']) for v in chart_sales_query}
        
        # 4. Gera os dados finais, preenchendo com 0 os meses sem vendas
        faturamento_chart_data = [sales_by_month.get(label, 0) for label in faturamento_chart_labels]


    else: # Default para 'semanal'
        faturamento_periodo = 'semanal'
        period_days = 7
        period_text = "últimos 7 dias"
        prev_period_text = "7 dias anteriores"
        current_start_date = today - timedelta(days=period_days - 1)
        prev_end_date = current_start_date - timedelta(days=1)
        prev_start_date = prev_end_date - timedelta(days=period_days - 1)

        dias_semana_pt = { 0: 'Segunda', 1: 'Terça', 2: 'Quarta', 3: 'Quinta', 4: 'Sexta', 5: 'Sábado', 6: 'Domingo' }
        chart_sales_query = all_sales.filter(data_venda__date__gte=current_start_date).values('data_venda__date').annotate(total=Sum('valor_total_liquido')).order_by('data_venda__date')
        sales_by_day = {item['data_venda__date']: float(item['total']) for item in chart_sales_query}
        faturamento_chart_labels = [dias_semana_pt[(current_start_date + timedelta(days=i)).weekday()] for i in range(period_days)]
        faturamento_chart_data = [sales_by_day.get(current_start_date + timedelta(days=i), 0) for i in range(period_days)]

    current_period_sales = all_sales.filter(data_venda__date__gte=current_start_date, data_venda__date__lte=today).aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')
    previous_period_sales = all_sales.filter(data_venda__date__range=[prev_start_date, prev_end_date]).aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')
    sales_variation = ((current_period_sales - previous_period_sales) / previous_period_sales * 100) if previous_period_sales > 0 else 0
    faturamento_chart_json = json.dumps({'labels': faturamento_chart_labels, 'data': faturamento_chart_data})
    
    # --- LÓGICA DO FILTRO DE PERÍODO PRINCIPAL (CORRIGIDA) ---
    period = request.GET.get('period', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Inicializa as datas
    start_date, end_date = None, None
    today = timezone.now().date()

    vendas_concluidas = Venda.objects.filter(user=request.user)

    if period == 'custom':
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
    elif period.isdigit():
        days = int(period)
        end_date = today
        start_date = end_date - timedelta(days=days)
    # Se for 'all', start_date e end_date permanecem None e o filtro não é aplicado

    # Aplica o filtro de data se as datas forem válidas
    if start_date and end_date:
        vendas_concluidas = vendas_concluidas.filter(data_venda__date__range=[start_date, end_date])

    # Coloque este bloco dentro de faturamento_dashboard_view, após a linha que define 'vendas_concluidas'

    # --- CÁLCULOS PARA OS NOVOS CARDS DE INFORMAÇÃO ---
    melhor_mes_valor = pior_mes_valor = None
    melhor_mes_qtd = pior_mes_qtd = None
    melhor_mes_ticket = pior_mes_ticket = None
    melhor_mes_prazo = pior_mes_prazo = None
    prazo_medio_recebimento = 0

    # Agrupa as vendas do período por mês para análise de melhor/pior
    vendas_por_mes = vendas_concluidas.annotate(mes=TruncMonth('data_venda')) \
        .values('mes') \
        .annotate(
            valor_total=Sum('valor_total_liquido'),
            qtd_total=Count('id')
        ).order_by('mes')

    if vendas_por_mes:
        # Análise de Valor
        vendas_ordenadas_valor = sorted(vendas_por_mes, key=lambda v: v['valor_total'])
        pior_mes_valor = {'mes': vendas_ordenadas_valor[0]['mes'].strftime('%b/%Y'), 'valor': vendas_ordenadas_valor[0]['valor_total']}
        melhor_mes_valor = {'mes': vendas_ordenadas_valor[-1]['mes'].strftime('%b/%Y'), 'valor': vendas_ordenadas_valor[-1]['valor_total']}

        # Análise de Quantidade
        vendas_ordenadas_qtd = sorted(vendas_por_mes, key=lambda v: v['qtd_total'])
        pior_mes_qtd = {'mes': vendas_ordenadas_qtd[0]['mes'].strftime('%b/%Y'), 'qtd': vendas_ordenadas_qtd[0]['qtd_total']}
        melhor_mes_qtd = {'mes': vendas_ordenadas_qtd[-1]['mes'].strftime('%b/%Y'), 'qtd': vendas_ordenadas_qtd[-1]['qtd_total']}

        # Análise de Ticket Médio
        tickets_por_mes = [
            {'mes': v['mes'].strftime('%b/%Y'), 'ticket': v['valor_total'] / v['qtd_total'] if v['qtd_total'] > 0 else 0}
            for v in vendas_por_mes
        ]
        tickets_ordenados = sorted(tickets_por_mes, key=lambda t: t['ticket'])
        pior_mes_ticket = tickets_ordenados[0]
        melhor_mes_ticket = tickets_ordenados[-1]

    # --- INÍCIO DA CORREÇÃO PARA VENDA INEXISTENTE ---
    # 1. Tenta buscar a data da primeira venda de forma segura
    try:
        earliest_sale_date = all_sales.earliest('data_venda').data_venda.date()
    except Venda.DoesNotExist:
        earliest_sale_date = timezone.now().date() # Se não houver vendas, usa a data de hoje como padrão

    # 2. Define o range de datas
    date_range = [start_date, end_date] if period != 'all' else [earliest_sale_date, timezone.now().date()]

    # 3. Aplica o range de datas na consulta
    contas_recebidas_periodo = ReceivableAccount.objects.filter(
        user=request.user, 
        is_received=True, 
        payment_date__isnull=False,
        payment_date__range=date_range
    )
    # --- FIM DA CORREÇÃO ---

    if contas_recebidas_periodo.exists():
        duracao_media = contas_recebidas_periodo.aggregate(
            media=Avg(ExpressionWrapper(F('payment_date') - F('due_date'), output_field=fields.DurationField()))
        )['media']
        prazo_medio_recebimento = duracao_media.days if duracao_media else 0

        # Análise de Melhor/Pior mês para PMR
        pmr_por_mes = contas_recebidas_periodo.annotate(mes=TruncMonth('payment_date')) \
            .values('mes') \
            .annotate(media_prazo=Avg(ExpressionWrapper(F('payment_date') - F('due_date'), output_field=fields.DurationField()))) \
            .order_by('mes')
        
        if pmr_por_mes:
            pmr_ordenado = sorted(pmr_por_mes, key=lambda p: p['media_prazo'])
            # Melhor prazo é o menor (mais rápido para receber)
            melhor_mes_prazo = {'mes': pmr_ordenado[0]['mes'].strftime('%b/%Y'), 'dias': pmr_ordenado[0]['media_prazo'].days}
            # Pior prazo é o maior (mais demorado para receber)
            pior_mes_prazo = {'mes': pmr_ordenado[-1]['mes'].strftime('%b/%Y'), 'dias': pmr_ordenado[-1]['media_prazo'].days}


    # --- INÍCIO DO NOVO CÓDIGO ---
    # 1. Busca dados para o gráfico de Vendas por Cliente
    vendas_por_cliente = vendas_concluidas.values('cliente__nome') \
        .annotate(valor_total=Sum('valor_total_liquido')) \
        .order_by('-valor_total')
    vendas_por_cliente_json = json.dumps({
        'labels': [item['cliente__nome'] for item in vendas_por_cliente],
        'data': [float(item['valor_total']) for item in vendas_por_cliente]
    })

    # 2. Busca dados para o gráfico de Vendas por Produto/Serviço
    vendas_por_produto_servico = ItemVenda.objects.filter(venda__in=vendas_concluidas) \
        .values('produto__nome') \
        .annotate(valor=Sum(F('quantidade') * F('preco_unitario'))) \
        .order_by('-valor')
    vendas_por_produto_servico_json = json.dumps({
        'labels': [item['produto__nome'] for item in vendas_por_produto_servico],
        'data': [float(item['valor']) for item in vendas_por_produto_servico]
    })
    # --- FIM DO NOVO CÓDIGO ---


    # --- NOVO: Lógica para o gráfico Top 5 Cidades ---
    # CÓDIGO CORRIGIDO
    top_5_cidades = vendas_concluidas.exclude(cidade__isnull=True).exclude(cidade__exact='') \
                                    .values('cidade') \
                                    .annotate(total_vendido=Sum('valor_total_liquido')) \
                                    .order_by('-total_vendido')[:5]

    top_cidades_labels = [item['cidade'] for item in top_5_cidades]
    top_cidades_data = [float(item['total_vendido']) for item in top_5_cidades]
    top_cidades_json = json.dumps({
        'labels': top_cidades_labels,
        'data': top_cidades_data
    })
    # --- FIM DO NOVO CÓDIGO ---    

    # --- CÁLCULOS PARA OS CARDS SUPERIORES ---
    # --- NOVOS CÁLCULOS PARA OS CARDS DE ANÁLISE DE CLIENTES ---
    
    # Card 1: Total de clientes cadastrados na base de dados.
    total_clientes_base = Cliente.objects.filter(user=request.user).count()

    # Card 2: Contagem de clientes únicos que realizaram compras no período filtrado.
    clientes_no_periodo = vendas_concluidas.values('cliente').distinct().count()

    # Card 3: Contagem de clientes que fizeram mais de uma compra no período.
    clientes_recorrentes = vendas_concluidas.values('cliente').annotate(num_compras=Count('id')).filter(num_compras__gt=1).count()

    # Card 4: Percentual de clientes da base que compraram no período.
    cobertura_carteira = (clientes_no_periodo / total_clientes_base * 100) if total_clientes_base > 0 else 0

    # Mantemos estes cálculos pois são usados em outras partes do dashboard
    total_vendido = vendas_concluidas.aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')
    qtd_vendas = vendas_concluidas.count()
    ticket_medio = total_vendido / qtd_vendas if qtd_vendas > 0 else Decimal('0')

    # --- CORREÇÃO DEFINITIVA: Gráfico Vendas por Mês/Ano (Mostra todos os 12 meses do ano) ---
    ano_corrente = timezone.now().year
    labels_meses = [datetime(ano_corrente, i, 1).strftime('%b/%Y') for i in range(1, 13)]
    vendas_por_mes_query = all_sales.filter(data_venda__year=ano_corrente).annotate(
        mes=TruncMonth('data_venda')
    ).values('mes').annotate(total=Sum('valor_total_liquido')).order_by('mes')
    vendas_por_mes_dict = {v['mes'].strftime('%b/%Y'): float(v['total']) for v in vendas_por_mes_query}
    data_meses = [vendas_por_mes_dict.get(label, 0) for label in labels_meses]
    vendas_por_mes_json = json.dumps({'labels': labels_meses, 'data': data_meses})
    
    # --- NOVO: Lógica para o mapa do Brasil (Usa vendas_concluidas para respeitar o filtro) ---
    vendas_por_estado = vendas_concluidas.values('estado').annotate(total=Sum('valor_total_liquido')).order_by('-total')
    vendas_por_estado_data = [['Estado', 'Valor Vendido']]
    for venda in vendas_por_estado:
        if venda['estado'] and len(venda['estado']) == 2:
            estado_code = f"BR-{venda['estado'].upper()}"
            vendas_por_estado_data.append([estado_code, float(venda['total'])])
    vendas_por_estado_json = json.dumps(vendas_por_estado_data)

    # --- LÓGICA EXISTENTE PARA OS DEMAIS COMPONENTES DO DASHBOARD ---
    itens_vendidos_agrupados = (ItemVenda.objects.filter(venda__in=vendas_concluidas).values('produto__nome').annotate(valor=Sum(F('quantidade') * F('preco_unitario'))).order_by('-valor'))
    abc_data = []
    abc_chart_data = {'labels': [], 'data_valor': [], 'data_acumulado': []}
    if itens_vendidos_agrupados:
        df = pd.DataFrame(list(itens_vendidos_agrupados))
        df.rename(columns={'produto__nome': 'produto'}, inplace=True)
        total_geral_vendas_abc = df['valor'].sum()
        if total_geral_vendas_abc > 0:
            df['acumulado'] = df['valor'].cumsum()
            df['acumulado_percentual'] = 100 * df['acumulado'] / total_geral_vendas_abc
            df['classe'] = df['acumulado_percentual'].apply(lambda p: 'A' if p <= 80 else ('B' if p <= 95 else 'C'))
            abc_data = df.to_dict('records')
            abc_chart_data['labels'] = df['produto'].tolist()
            abc_chart_data['data_valor'] = [float(v) for v in df['valor'].tolist()]
            abc_chart_data['data_acumulado'] = [round(float(p), 2) for p in df['acumulado_percentual'].tolist()]
    abc_chart_json = json.dumps(abc_chart_data)
    # --- INÍCIO DO NOVO CÓDIGO: Resumo para os cards ABC ---
    abc_summary = {
        'A': {'count': 0, 'value': Decimal('0.0'), 'percentage': Decimal('0.0')},
        'B': {'count': 0, 'value': Decimal('0.0'), 'percentage': Decimal('0.0')},
        'C': {'count': 0, 'value': Decimal('0.0'), 'percentage': Decimal('0.0')}
    }

    if not df.empty and total_geral_vendas_abc > 0:
        summary_df = df.groupby('classe')['valor'].agg(['count', 'sum'])
        for classe, data in summary_df.iterrows():
            abc_summary[classe] = {
                'count': data['count'],
                'value': data['sum'],
                'percentage': (data['sum'] / total_geral_vendas_abc) * 100
            }
    # --- FIM DO NOVO CÓDIGO ---
    total_itens_vendidos = int(ItemVenda.objects.filter(venda__in=vendas_concluidas).aggregate(total=Sum('quantidade'))['total'] or 0)
    
    # --- LÓGICA CORRIGIDA PARA O GRÁFICO "ANÁLISE DE VENDAS" COM MESES FIXOS ---
    # Usaremos `all_sales` aqui para ignorar o filtro de período principal.
    
    # --- LÓGICA CORRIGIDA PARA O GRÁFICO "ANÁLISE DE VENDAS" COM MESES FIXOS (JANEIRO A DEZEMBRO) ---
    
    meses_fixos = OrderedDict()
    ano_corrente = timezone.now().year
    
    # 1. Cria a estrutura com os 12 meses do ano corrente, todos com valor zero.
    for i in range(1, 13):
        mes_chave = datetime(ano_corrente, i, 1).strftime('%b/%Y')
        meses_fixos[mes_chave] = {'valor_vendas': 0.0, 'qtd_vendas': 0, 'itens_vendidos': 0}
        
    # 2. Busca os dados reais do ano corrente no banco de dados.
    vendas_do_ano_corrente = all_sales.filter(data_venda__year=ano_corrente)

    vendas_agrupadas = vendas_do_ano_corrente.annotate(mes=TruncMonth('data_venda')).values('mes').annotate(
        valor_total=Sum('valor_total_liquido'),
        qtd_total=Count('id')
    ).order_by('mes')

    itens_agrupados = ItemVenda.objects.filter(venda__in=vendas_do_ano_corrente).annotate(mes=TruncMonth('venda__data_venda')).values('mes').annotate(
        itens_total=Sum('quantidade')
    ).order_by('mes')

    # 3. Preenche a estrutura com os dados reais encontrados.
    for v in vendas_agrupadas:
        mes_chave = v['mes'].strftime('%b/%Y')
        if mes_chave in meses_fixos:
            meses_fixos[mes_chave]['valor_vendas'] = float(v['valor_total'])
            meses_fixos[mes_chave]['qtd_vendas'] = v['qtd_total']
            
    for i in itens_agrupados:
        mes_chave = i['mes'].strftime('%b/%Y')
        if mes_chave in meses_fixos:
            meses_fixos[mes_chave]['itens_vendidos'] = float(i['itens_total'])

    # 4. Monta o dicionário final para o template.
    indicadores_data = {
        'labels': list(meses_fixos.keys()),
        'datasets': {
            'valor_vendas': [d['valor_vendas'] for d in meses_fixos.values()],
            'qtd_vendas': [d['qtd_vendas'] for d in meses_fixos.values()],
            'ticket_medio': [(d['valor_vendas'] / d['qtd_vendas'] if d.get('qtd_vendas', 0) > 0 else 0) for d in meses_fixos.values()],
            'itens_vendidos': [d['itens_vendidos'] for d in meses_fixos.values()],
            'valor_medio_item': [(d['valor_vendas'] / d['itens_vendidos'] if d.get('itens_vendidos', 0) > 0 else 0) for d in meses_fixos.values()],
        }
    }

    # 3. Preenche a estrutura com os dados reais encontrados.
    for v in vendas_agrupadas:
        mes_chave = v['mes'].strftime('%b/%Y')
        if mes_chave in meses_fixos:
            meses_fixos[mes_chave]['valor_vendas'] = float(v['valor_total'])
            meses_fixos[mes_chave]['qtd_vendas'] = v['qtd_total']
            
    for i in itens_agrupados:
        mes_chave = i['mes'].strftime('%b/%Y')
        if mes_chave in meses_fixos:
            meses_fixos[mes_chave]['itens_vendidos'] = float(i['itens_total'])

    # 4. Monta o dicionário final para o template.
    indicadores_data = {
        'labels': list(meses_fixos.keys()),
        'datasets': {
            'valor_vendas': [d['valor_vendas'] for d in meses_fixos.values()],
            'qtd_vendas': [d['qtd_vendas'] for d in meses_fixos.values()],
            'ticket_medio': [(d['valor_vendas'] / d['qtd_vendas'] if d.get('qtd_vendas', 0) > 0 else 0) for d in meses_fixos.values()],
            'itens_vendidos': [d['itens_vendidos'] for d in meses_fixos.values()],
            'valor_medio_item': [(d['valor_vendas'] / d['itens_vendidos'] if d.get('itens_vendidos', 0) > 0 else 0) for d in meses_fixos.values()],
        }
    }
    
    top_produtos_desc = list(itens_vendidos_agrupados[:10])
    top_produtos_servicos_chart_json = json.dumps({'labels': [item['produto__nome'] for item in top_produtos_desc], 'data': [float(item['valor']) for item in top_produtos_desc]})
    
    top_clientes_desc = list(vendas_concluidas.values('cliente__nome').annotate(valor_total=Sum('valor_total_liquido')).order_by('-valor_total')[:10])
    top_clientes_chart_json = json.dumps({'labels': [item['cliente__nome'] for item in top_clientes_desc], 'data': [float(item['valor_total']) for item in top_clientes_desc]})
    
    vendas_produtos_desc = list(ItemVenda.objects.filter(venda__in=vendas_concluidas, produto__tipo='PRODUTO').values('produto__nome').annotate(valor=Sum(F('quantidade') * F('preco_unitario'))).order_by('-valor')[:10])
    vendas_por_produto_data_json = json.dumps({'labels': [item['produto__nome'] for item in vendas_produtos_desc], 'data': [float(item['valor']) for item in vendas_produtos_desc]})
    
    vendas_servicos_desc = list(ItemVenda.objects.filter(venda__in=vendas_concluidas, produto__tipo='SERVICO').values('produto__nome').annotate(valor=Sum(F('quantidade') * F('preco_unitario'))).order_by('-valor')[:10])
    vendas_por_servico_data_json = json.dumps({'labels': [item['produto__nome'] for item in vendas_servicos_desc], 'data': [float(item['valor']) for item in vendas_servicos_desc]})
    
    top_clientes_valor = list(vendas_concluidas.values('cliente__nome').annotate(total=Sum('valor_total_liquido')).order_by('-total')[:8])
    top_clientes_qtd = list(vendas_concluidas.values('cliente__nome').annotate(total=Count('id')).order_by('-total')[:8])
    top_clientes_ticket = list(vendas_concluidas.values('cliente__nome').annotate(total=Avg('valor_total_liquido')).order_by('-total')[:8])
    top_clientes_itens = list(ItemVenda.objects.filter(venda__in=vendas_concluidas).values('venda__cliente__nome').annotate(total=Sum('quantidade')).order_by('-total')[:8])
    top_clientes_valor_medio_item = list(ItemVenda.objects.filter(venda__in=vendas_concluidas).values('venda__cliente__nome').annotate(valor_total=Sum(F('quantidade') * F('preco_unitario')), itens_totais=Sum('quantidade')).annotate(total=F('valor_total') / F('itens_totais')).order_by('-total')[:8])
    
    analise_cliente_data = {
        'valor_vendas': {'labels': [c['cliente__nome'] for c in top_clientes_valor], 'data': [float(c['total']) for c in top_clientes_valor]},
        'qtd_vendas': {'labels': [c['cliente__nome'] for c in top_clientes_qtd], 'data': [int(c['total']) for c in top_clientes_qtd]},
        'ticket_medio': {'labels': [c['cliente__nome'] for c in top_clientes_ticket], 'data': [float(c['total']) for c in top_clientes_ticket]},
        'itens_vendidos': {'labels': [c['venda__cliente__nome'] for c in top_clientes_itens], 'data': [float(c['total']) for c in top_clientes_itens]},
        'valor_medio_produto': {'labels': [c['venda__cliente__nome'] for c in top_clientes_valor_medio_item], 'data': [float(c['total']) for c in top_clientes_valor_medio_item]}
    }
    analise_cliente_json = json.dumps(analise_cliente_data)
    
    vendas_com_vendedor = vendas_concluidas.filter(vendedor__isnull=False)
    analise_vendedores_query = vendas_com_vendedor.values('vendedor__nome').annotate(valor_vendas=Sum('valor_total_liquido'), qtd_vendas=Count('id'), ticket_medio=Avg('valor_total_liquido'), total_clientes=Count('cliente', distinct=True)).order_by('-valor_vendas')
    # --- INÍCIO: LÓGICA PARA FILTRO DE METAS BATIDAS ---
    metas_dict = {}
    # Se houver um filtro de data aplicado, busca as metas do período
    if start_date and end_date:
        # Soma as metas mensais que se encaixam no período para cada vendedor
        metas_periodo = MetaFaturamento.objects.filter(
            user=request.user,
            vendedor__isnull=False,
            mes_ano__gte=start_date,
            mes_ano__lte=end_date
        ).values('vendedor__nome').annotate(total_meta=Sum('valor_meta'))
        metas_dict = {item['vendedor__nome']: item['total_meta'] for item in metas_periodo}

    vendedores_meta_batida = []
    vendedores_meta_nao_batida = []

    # Itera sobre os vendedores que tiveram vendas para comparar com suas metas
    for vendedor_data in analise_vendedores_query:
        nome = vendedor_data['vendedor__nome']
        valor_vendas = vendedor_data.get('valor_vendas', Decimal('0'))
        meta = metas_dict.get(nome) # Pega a meta do vendedor

        # Se a meta existe e foi batida, adiciona à lista de sucesso
        if meta is not None and valor_vendas >= meta:
            vendedores_meta_batida.append(vendedor_data)
        else:
            vendedor_data['meta'] = meta if meta is not None else Decimal('0')
            # Se a meta não foi batida ou não existe, adiciona à outra lista
            vendedores_meta_nao_batida.append(vendedor_data)
    # --- FIM: LÓGICA PARA FILTRO DE METAS BATIDAS ---
    top_3_vendedores = list(analise_vendedores_query[:3])
    valor_vendido_top_3 = sum(v.get('valor_vendas', 0) or 0 for v in top_3_vendedores)
    qtd_vendas_top_3 = sum(v.get('qtd_vendas', 0) for v in top_3_vendedores)
    if len(top_3_vendedores) >= 1: top_3_vendedores[0]['rank'] = 1
    if len(top_3_vendedores) >= 2: top_3_vendedores[1]['rank'] = 2
    if len(top_3_vendedores) >= 3: top_3_vendedores[2]['rank'] = 3
    
    vendedores_data_table = list(analise_vendedores_query[:10])
    total_geral_vendas_vendedores = vendas_com_vendedor.aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('1')
    for v in vendedores_data_table: v['percentual_valor'] = (v.get('valor_vendas', 0) / total_geral_vendas_vendedores) * 100
        
    # Em views.py, na função faturamento_dashboard_view, substitua o bloco de análise de vendedor por este:

    # --- INÍCIO DO NOVO CÓDIGO CORRIGIDO PARA O GRÁFICO DINÂMICO ---

    # 1. Query principal de vendedores, agora buscando também a comissão de cada um
    analise_vendedores_query = vendas_com_vendedor.values(
        'vendedor__nome', 
        'vendedor__comissao_percentual'  # <-- Buscando o % de comissão do modelo Vendedor
    ).annotate(
        valor_vendas=Sum('valor_total_liquido'),
        qtd_vendas=Count('id'),
        ticket_medio=Avg('valor_total_liquido'),
        total_clientes=Count('cliente', distinct=True)
    ).order_by('-valor_vendas')

    vendedores_data_table = list(analise_vendedores_query) # Remove o slice para usar a query completa abaixo

    # 2. Prepara os dados para o gráfico, calculando o valor da comissão
    valor_vendas_data = {
        'labels': [],
        'data': [],
        'comissoes_valor': [],
        'comissoes_percentual': []
    }
    # Itera sobre a lista de vendedores para popular os dados do gráfico
    for v in vendedores_data_table:
        valor_vendas_data['labels'].append(v['vendedor__nome'])
        valor_vendas_data['data'].append(float(v.get('valor_vendas', 0)))
        
        # Pega a comissão do vendedor (ou 0 se não estiver definida)
        comissao_percent = v.get('vendedor__comissao_percentual', 0) or 0
        # Calcula o valor da comissão
        valor_comissao = float(v.get('valor_vendas', 0) or 0) * (float(comissao_percent) / 100.0)
        
        valor_vendas_data['comissoes_valor'].append(valor_comissao)
        valor_vendas_data['comissoes_percentual'].append(float(comissao_percent))
        # ▼▼▼ ADICIONE ESTA LINHA QUE ESTAVA FALTANDO ▼▼▼
        # Calcula o % do valor para a TABELA
        v['percentual_valor'] = (v.get('valor_vendas', 0) / total_geral_vendas_vendedores) * 100 if total_geral_vendas_vendedores else 0
    vendedores_nao_batida_sorted = sorted(vendedores_meta_nao_batida, key=lambda x: x.get('valor_vendas', 0), reverse=True)
    # 3. Monta o dicionário final para o template
    analise_vendedor_chart_data = {
        'valor_vendas': valor_vendas_data,
        'qtd_vendas': {
            'labels': [v['vendedor__nome'] for v in sorted(vendedores_data_table, key=lambda x: x.get('qtd_vendas', 0), reverse=True)],
            'data': [v.get('qtd_vendas', 0) for v in sorted(vendedores_data_table, key=lambda x: x.get('qtd_vendas', 0), reverse=True)]
        },
        'ticket_medio': {
            'labels': [v['vendedor__nome'] for v in sorted(vendedores_data_table, key=lambda x: x.get('ticket_medio', 0), reverse=True)],
            'data': [float(v.get('ticket_medio', 0)) for v in sorted(vendedores_data_table, key=lambda x: x.get('ticket_medio', 0), reverse=True)]
        },
        'metas_batidas': {
            'labels': [v['vendedor__nome'] for v in sorted(vendedores_meta_batida, key=lambda x: x.get('valor_vendas', 0), reverse=True)],
            'data': [float(v.get('valor_vendas', 0)) for v in sorted(vendedores_meta_batida, key=lambda x: x.get('valor_vendas', 0), reverse=True)]
        },
        'metas_nao_batidas': {
            'labels': [v['vendedor__nome'] for v in vendedores_nao_batida_sorted],
            'data': [float(v.get('valor_vendas', 0)) for v in vendedores_nao_batida_sorted],
            'metas': [float(v.get('meta', 0)) for v in vendedores_nao_batida_sorted]
        },
    }
    analise_vendedor_chart_json = json.dumps(analise_vendedor_chart_data)
    # --- FIM DO NOVO CÓDIGO CORRIGIDO ---



    doze_meses_atras = timezone.now().date() - relativedelta(months=11)
    vendas_periodo = (vendas_com_vendedor.filter(data_venda__date__gte=doze_meses_atras).annotate(mes=TruncMonth('data_venda')).values('mes').annotate(total=Sum('valor_total_liquido')).order_by('mes'))
    vendas_periodo_json = json.dumps({'labels': [v['mes'].strftime('%b/%Y') for v in vendas_periodo], 'data': [float(v['total']) for v in vendas_periodo]})
    
    resumo_vendedores = vendas_com_vendedor.aggregate(valor_total=Sum('valor_total_liquido'), qtd_total=Count('id'))
    resumo_valor_vendido = resumo_vendedores['valor_total'] or Decimal('0')
    resumo_qtd_vendas = resumo_vendedores['qtd_total'] or 0
    resumo_ticket_medio = (resumo_valor_vendido / resumo_qtd_vendas) if resumo_qtd_vendas > 0 else Decimal('0')

    # --- INÍCIO DA NOVA LÓGICA ---
    # Card 4: Cobertura da carteira pelos vendedores
    total_clientes_base = Cliente.objects.filter(user=request.user).count()
    clientes_atendidos_por_vendedores = vendas_com_vendedor.values('cliente').distinct().count()
    resumo_cobertura_carteira = (clientes_atendidos_por_vendedores / total_clientes_base * 100) if total_clientes_base > 0 else 0
    # --- FIM DA NOVA LÓGICA ---

    percentual_valor_top_3 = (valor_vendido_top_3 / resumo_valor_vendido * 100) if resumo_valor_vendido > 0 else 0

    # --- INÍCIO DO NOVO CÓDIGO PARA O RODAPÉ DA TABELA ---
    vendedor_table_totals = {
        'total_valor': Decimal('0'),
        'total_qtd': 0,
        'total_ticket_medio': Decimal('0'),
        'total_clientes': 0
    }
    # Usa a mesma query da tabela para garantir que os totais correspondam aos dados filtrados
    if analise_vendedores_query.exists():
        totais_agregados = analise_vendedores_query.aggregate(
            total_valor=Sum('valor_vendas'),
            total_qtd=Sum('qtd_vendas')
        )
        vendedor_table_totals['total_valor'] = totais_agregados.get('total_valor', Decimal('0'))
        vendedor_table_totals['total_qtd'] = totais_agregados.get('total_qtd', 0)
        
        # O total de clientes é a contagem de clientes únicos atendidos por todos os vendedores
        vendedor_table_totals['total_clientes'] = vendas_com_vendedor.values('cliente').distinct().count()
        
        # O Ticket Médio geral é o valor total dividido pela quantidade total
        if vendedor_table_totals['total_qtd'] > 0:
            vendedor_table_totals['total_ticket_medio'] = vendedor_table_totals['total_valor'] / vendedor_table_totals['total_qtd']
    # --- FIM DO NOVO CÓDIGO ---

    

    # 1. Define os períodos de tempo de forma robusta, considerando timezones
    now = timezone.now()
    # Para "Hoje", criamos um range do início ao fim do dia atual
    start_of_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_today = start_of_today + timedelta(days=1)
    
    # Para "Neste Mês", criamos um range do primeiro dia do mês até agora
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    all_sales = Venda.objects.filter(user=request.user)

    # Filtra as vendas para os períodos de uma só vez para otimizar
    sales_today = all_sales.filter(data_venda__gte=start_of_today, data_venda__lt=end_of_today)
    sales_this_month = all_sales.filter(data_venda__gte=start_of_month)

    # Card 1: Valor Vendido
    valor_vendido_hoje = sales_today.aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')
    valor_vendido_mes = sales_this_month.aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')

    # Card 2: Quantidade de Vendas
    qtd_vendas_hoje = sales_today.count()
    qtd_vendas_mes = sales_this_month.count()

    # Card 3 e 4: Lógica final e simplificada para Clientes Novos e Recorrentes
    
    # --- Lógica para "Hoje" ---
    # Pega os IDs de clientes únicos que compraram hoje
    clientes_que_compraram_hoje_ids = sales_today.values_list('cliente_id', flat=True).distinct()
    # Anota o número total de vendas para cada um desses clientes
    clientes_hoje_com_contagem = Cliente.objects.filter(id__in=clientes_que_compraram_hoje_ids).annotate(
        num_vendas=Count('venda')
    )
    # É NOVO se o total de vendas da sua vida é exatamente 1
    novos_clientes_hoje = clientes_hoje_com_contagem.filter(num_vendas=1).count()
    # É RECORRENTE se o total de vendas da sua vida é maior que 1
    clientes_recorrentes_hoje = clientes_hoje_com_contagem.filter(num_vendas__gt=1).count()
    
    # --- Lógica para "Neste Mês" ---
    # Pega os IDs de clientes únicos que compraram neste mês
    clientes_que_compraram_no_mes_ids = sales_this_month.values_list('cliente_id', flat=True).distinct()
    # Anota o número total de vendas para cada um desses clientes
    clientes_mes_com_contagem = Cliente.objects.filter(id__in=clientes_que_compraram_no_mes_ids).annotate(
        num_vendas=Count('venda')
    )
    # É NOVO se o total de vendas da sua vida é exatamente 1
    novos_clientes_mes = clientes_mes_com_contagem.filter(num_vendas=1).count()
    # É RECORRENTE se o total de vendas da sua vida é maior que 1
    clientes_recorrentes_mes = clientes_mes_com_contagem.filter(num_vendas__gt=1).count()

    # Em views.py, dentro de faturamento_dashboard_view, ANTES da definição do context

    # NOVO CÓDIGO CORRIGIDO
    # --- INÍCIO: NOVOS CÁLCULOS PARA A SEÇÃO "ANÁLISE POR CLIENTE" ---

    # 1. Dados para a tabela e gráfico de evolução mensal (Janeiro a Dezembro do ano corrente)
    dados_evolucao_cliente = []
    current_year = timezone.now().year

    # Anota a data da primeira compra em cada cliente (consulta feita uma vez fora do loop)
    clientes_com_primeira_compra = Cliente.objects.filter(user=request.user).annotate(
        primeira_compra=Min('venda__data_venda__date')
    )

    # Loop de 1 (Janeiro) a 12 (Dezembro)
    for month_num in range(1, 13):
        # Define o início e o fim de cada mês do loop
        inicio_mes = datetime(current_year, month_num, 1).date()
        fim_mes = (inicio_mes + relativedelta(months=1)) - timedelta(days=1)

        # Filtra as vendas apenas para o mês atual do loop
        vendas_no_mes = all_sales.filter(data_venda__date__range=[inicio_mes, fim_mes])

        # a. Qtd. de Clientes que compraram no mês
        clientes_compra_mes = vendas_no_mes.values('cliente').distinct().count()

        # b. Clientes Recorrentes (que compraram mais de uma vez no mês)
        recorrentes_na_tabela = vendas_no_mes.values('cliente').annotate(
            num_compras=Count('id')
        ).filter(num_compras__gt=1).count()

        # c. Clientes Novos (cuja primeira compra foi neste mês)
        clientes_novos_mes = clientes_com_primeira_compra.filter(
            primeira_compra__range=[inicio_mes, fim_mes]
        ).count()

        dados_evolucao_cliente.append({
            'mes_ano': inicio_mes.strftime('%B/%Y').capitalize(),
            'qtd_clientes_compra': clientes_compra_mes,
            'clientes_recorrentes_mes': recorrentes_na_tabela,
            'clientes_novos_mes': clientes_novos_mes,
        })

    # A lista já está em ordem cronológica, então a inversão não é mais necessária.

    # 2. Prepara os dados para o gráfico de linha em formato JSON
    evolucao_clientes_chart_json = json.dumps({
        'labels': [d['mes_ano'] for d in dados_evolucao_cliente],
        'data': [d['qtd_clientes_compra'] for d in dados_evolucao_cliente]
    })
    # --- FIM DOS NOVOS CÁLCULOS ---

    # --- INÍCIO DA LÓGICA ATUALIZADA (METAS E VELOCÍMETRO) ---

    # Pega o filtro da URL, o padrão é 'ano'
    meta_periodo = request.GET.get('meta_periodo', 'ano')
    today = timezone.now().date()
    current_year = today.year
    current_month = today.month

    # Variáveis genéricas para armazenar os resultados
    valor_acumulado = Decimal('0')
    valor_meta_atual = Decimal('0')

    if meta_periodo == 'mes':
        # LÓGICA PARA O FILTRO "MÊS"
        meta_obj = MetaFaturamento.objects.filter(
            user=request.user, vendedor__isnull=True, mes_ano__year=current_year, mes_ano__month=current_month
        ).first()
        valor_meta_atual = meta_obj.valor_meta if meta_obj else Decimal('0')

        valor_acumulado = Venda.objects.filter(
            user=request.user, data_venda__year=current_year, data_venda__month=current_month
        ).aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')

    else: # LÓGICA PARA O FILTRO "ANO" (comportamento padrão)
        valor_acumulado = Venda.objects.filter(
            user=request.user, data_venda__year=current_year, data_venda__month__lte=current_month
        ).aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')

        valor_meta_atual = MetaFaturamento.objects.filter(
            user=request.user, vendedor__isnull=True, mes_ano__year=current_year, mes_ano__month__lte=current_month
        ).aggregate(total=Sum('valor_meta'))['total'] or Decimal('0')

    # Cálculos que servem para ambos os filtros
    diferenca_meta_valor = valor_acumulado - valor_meta_atual
    # VVVV--- ADICIONE ESTE BLOCO DE CÓDIGO ABAIXO ---VVVV

    diferenca_meta_percentual = 0  # Inicia com 0
    if valor_meta_atual > 0:
        # Calcula o percentual apenas se a meta for maior que zero
        diferenca_meta_percentual = (diferenca_meta_valor / valor_meta_atual) * 100

    # ^^^^--- FIM DO BLOCO DE CÓDIGO A SER ADICIONADO ---^^^^
    atingimento_percentual = (valor_acumulado / valor_meta_atual * 100) if valor_meta_atual > 0 else 0
    # --- INÍCIO DA NOVA LÓGICA PARA O RODAPÉ ---
    # Cria o texto para o rodapé do card de metas
    if meta_periodo == 'mes':
        # Formata a data para "Mês/Ano", ex: "Setembro/2025"
        data_mes_atual = today.replace(day=1)
        # Usa o locale 'pt_BR.UTF-8' já configurado no início da view
        texto_rodape_meta = data_mes_atual.strftime('%B/%Y').capitalize()
    else: # meta_periodo == 'ano'
        # Formata a data para "Ano/Ano", ex: "Ano/2025"
        texto_rodape_meta = f"Ano/{current_year}"
    # --- FIM DA NOVA LÓGICA PARA O RODAPÉ ---
    # A lógica para o gráfico de barras de metas ao lado continua a mesma (anual)
    MESES_PT = { 1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro' }
    metas_ano_corrente = MetaFaturamento.objects.filter(user=request.user, vendedor__isnull=True, mes_ano__year=current_year)
    vendas_ano_corrente = Venda.objects.filter(user=request.user, data_venda__year=current_year).annotate(mes=TruncMonth('data_venda__date')).values('mes').annotate(total=Sum('valor_total_liquido')).order_by('mes')
    metas_por_mes = {m.mes_ano.month: float(m.valor_meta) for m in metas_ano_corrente}
    vendas_por_mes = {v['mes'].month: float(v['total']) for v in vendas_ano_corrente}
    labels_meses_meta = list(MESES_PT.values())
    metas_data_final = [metas_por_mes.get(i, 0) for i in range(1, 13)]
    faturamento_data_meta = [vendas_por_mes.get(i, 0) for i in range(1, 13)]
    # --- LÓGICA PARA A SETA E TEXTO DO VELOCÍMETRO ---
    velocimetro_rotacao = 0
    # velocimetro_status_texto = "Abaixo da Meta"
    # cor_status = "#dc3545" # Vermelho

    # A rotação vai de 0 (extrema esquerda) a 180 (extrema direita)
    # O cálculo converte o percentual (0-100+) para o ângulo (0-180)
    # Limitamos o percentual a 100 para o cálculo do ângulo, para a seta não passar do verde
    percentual_para_rotacao = min(float(atingimento_percentual), 100.0)
    velocimetro_rotacao = ((percentual_para_rotacao / 100.0) * 180) - 90

    # if atingimento_percentual >= 95:
    #     velocimetro_status_texto = "Meta Atingida"
    #     cor_status = "#28a745" # Verde
    # elif atingimento_percentual >= 70:
    #     velocimetro_status_texto = "Na Média da Meta"
    #     cor_status = "#ffc107" # Amarelo
        # --- CONTEXTO FINAL 100% COMPLETO ---

    

    context = {
        'faturamento_periodo': faturamento_periodo, 'faturamento_chart_json': faturamento_chart_json,
        'current_period_sales': current_period_sales, 'previous_period_sales': previous_period_sales,
        'sales_variation': sales_variation, 'period_text': period_text, 'prev_period_text': prev_period_text,
        'valor_vendido_hoje': valor_vendido_hoje,
        'valor_vendido_mes': valor_vendido_mes,
        'qtd_vendas_hoje': qtd_vendas_hoje,
        'qtd_vendas_mes': qtd_vendas_mes,
        'novos_clientes_hoje': novos_clientes_hoje,
        'novos_clientes_mes': novos_clientes_mes,
        'clientes_recorrentes_hoje': clientes_recorrentes_hoje,
        'clientes_recorrentes_mes': clientes_recorrentes_mes,
        # Variáveis antigas mantidas para outras seções
        'total_vendido': total_vendido, 'qtd_vendas': qtd_vendas,
        'ticket_medio': ticket_medio, 
        
        # NOVAS VARIÁVEIS PARA OS CARDS DE ANÁLISE DE CLIENTES
        'card_total_clientes': total_clientes_base,
        'card_clientes_periodo': clientes_no_periodo,
        'card_clientes_recorrentes': clientes_recorrentes,
        'card_cobertura_carteira': cobertura_carteira,
        'vendas_por_mes_json': vendas_por_mes_json,
        'vendas_por_estado_json': vendas_por_estado_json,
        'abc_data': abc_data, 'abc_chart_json': abc_chart_json,
        'novo_dash_total_vendido': total_vendido, 'novo_dash_qtd_vendas': qtd_vendas,
        'novo_dash_ticket_medio': ticket_medio, 'novo_dash_itens_vendidos': total_itens_vendidos,
        'indicadores_mensais_json': json.dumps(indicadores_data),
        'top_produtos_servicos_chart_json': top_produtos_servicos_chart_json,
        'top_clientes_chart_json': top_clientes_chart_json,
        'vendas_por_produto_data_json': vendas_por_produto_data_json,
        'vendas_por_servico_data_json': vendas_por_servico_data_json,
        'analise_cliente_json': analise_cliente_json,
        'top_3_vendedores': top_3_vendedores, 'valor_vendido_top_3': valor_vendido_top_3,
        'qtd_vendas_top_3': qtd_vendas_top_3, 'vendedores_data_table': vendedores_data_table,
        'vendas_periodo_json': vendas_periodo_json, 'resumo_valor_vendido': resumo_valor_vendido,
        'resumo_qtd_vendas': resumo_qtd_vendas, 'resumo_ticket_medio': resumo_ticket_medio,
        'resumo_ticket_medio': resumo_ticket_medio, 'resumo_cobertura_carteira': resumo_cobertura_carteira,
        'analise_vendedor_chart_json': analise_vendedor_chart_json,
        'vendedor_table_totals': vendedor_table_totals,
        'abc_summary': abc_summary,
        'top_cidades_json': top_cidades_json,
        'prazo_medio_recebimento': prazo_medio_recebimento,
        'melhor_mes_valor': melhor_mes_valor,
        'pior_mes_valor': pior_mes_valor,
        'melhor_mes_qtd': melhor_mes_qtd,
        'pior_mes_qtd': pior_mes_qtd,
        'melhor_mes_ticket': melhor_mes_ticket,
        'pior_mes_ticket': pior_mes_ticket,
        'melhor_mes_prazo': melhor_mes_prazo,
        'pior_mes_prazo': pior_mes_prazo,
        'vendas_por_cliente_json': vendas_por_cliente_json,
        'vendas_por_produto_servico_json': vendas_por_produto_servico_json,
        'dados_evolucao_cliente': dados_evolucao_cliente,
        'evolucao_clientes_chart_json': evolucao_clientes_chart_json,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        # --- ADICIONE ESTAS 3 LINHAS AO FINAL DO SEU CONTEXTO ---
        'labels_meses_meta_json': json.dumps(labels_meses_meta),
        'faturamento_data_meta_json': json.dumps(faturamento_data_meta),
        'metas_data_final_json': json.dumps(metas_data_final),
        # NOVAS Variáveis do Card Velocímetro
        'valor_acumulado': valor_acumulado,
        'valor_meta_atual': valor_meta_atual,
        'meta_periodo': meta_periodo,
        'diferenca_meta_valor': diferenca_meta_valor,
        'diferenca_meta_percentual': diferenca_meta_percentual,
        'atingimento_percentual': float(atingimento_percentual),
        'velocimetro_rotacao': str(velocimetro_rotacao).replace(',', '.'),
        'percentual_valor_top_3': percentual_valor_top_3,
        'texto_rodape_meta': texto_rodape_meta,
        
    }
    return render(request, 'accounts/faturamento_dashboard.html', context)




@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_cadastros_comercial')
def comercial_cadastros_view(request):
    # Inicializa os formulários antes de qualquer lógica
    produto_form = ProdutoServicoForm()
    vendedor_form = VendedorForm()

    if request.method == 'POST':
        # 1. Checa se a ação é de excluir produtos
        if 'delete_selected_products' in request.POST:
            product_ids = request.POST.getlist('product_ids')
            if product_ids:
                produtos_para_excluir = ProdutoServico.objects.filter(id__in=product_ids, user=request.user)
                excluidos_count, protegidos = 0, []
                for produto in produtos_para_excluir:
                    if produto.itemvenda_set.exists():
                        protegidos.append(produto.nome)
                    else:
                        produto.delete()
                        excluidos_count += 1
                if excluidos_count > 0:
                    messages.success(request, f'{excluidos_count} produto(s) foram excluídos com sucesso.')
                if protegidos:
                    messages.warning(request, f'Os seguintes produtos não puderam ser excluídos pois estão associados a vendas: {", ".join(protegidos)}.')
            return redirect('comercial_cadastros')

        # 2. Checa se a ação é de excluir vendedores
        elif 'delete_selected_sellers' in request.POST:
            seller_ids = request.POST.getlist('seller_ids')
            if seller_ids:
                Vendedor.objects.filter(id__in=seller_ids, user=request.user).delete()
                messages.success(request, f'{len(seller_ids)} vendedor(es) selecionado(s) foram excluídos.')
            return redirect('comercial_cadastros')
        
        # 5. Checa se a ação é de excluir clientes (NOVO)
        elif 'delete_selected_clients' in request.POST:
            client_ids = request.POST.getlist('client_ids')
            if client_ids:
                # Proteção: Não excluir se tiver vendas
                clientes_para_excluir = Cliente.objects.filter(id__in=client_ids, user=request.user)
                excluidos_count = 0
                protegidos = []
                for cli in clientes_para_excluir:
                    if cli.venda_set.exists(): # Verifica se tem vendas
                        protegidos.append(cli.nome)
                    else:
                        cli.delete()
                        excluidos_count += 1
                
                if excluidos_count > 0:
                    messages.success(request, f'{excluidos_count} cliente(s) excluídos.')
                if protegidos:
                    messages.warning(request, f'Clientes com vendas não podem ser excluídos: {", ".join(protegidos)}.')
            return redirect('comercial_cadastros')

        # 3. Checa se a ação é de criar um produto
        elif request.POST.get('form_type') == 'produto':
            # ADICIONAMOS user=request.user AQUI
            produto_form = ProdutoServicoForm(request.POST, user=request.user) 
            
            if produto_form.is_valid():
                produto = produto_form.save(commit=False)
                produto.user = request.user
                produto.save()
                messages.success(request, 'Produto/Serviço cadastrado com sucesso!')
                return redirect('comercial_cadastros')

        # 4. Checa se a ação é de criar um vendedor
        elif request.POST.get('form_type') == 'vendedor':
            vendedor_form = VendedorForm(request.POST) # Repopula o formulário com os dados enviados
            if vendedor_form.is_valid():
                vendedor = vendedor_form.save(commit=False)
                vendedor.user = request.user
                vendedor.save()
                messages.success(request, 'Vendedor cadastrado com sucesso!')
                return redirect('comercial_cadastros')
            # Se o formulário for inválido, a função continuará e renderizará a página com os erros

    # Lógica GET (executada em carregamento normal da página ou quando um formulário POST é inválido)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_product = request.GET.get('search_product', '')

    produtos_list = ProdutoServico.objects.filter(user=request.user).order_by('-created_at')
    if start_date:
        produtos_list = produtos_list.filter(created_at__date__gte=start_date)
    if end_date:
        produtos_list = produtos_list.filter(created_at__date__lte=end_date)
    if search_product:
        produtos_list = produtos_list.filter(Q(nome__icontains=search_product) | Q(codigo__icontains=search_product))
    
    produto_paginator = Paginator(produtos_list, 10)
    produto_page_number = request.GET.get('produto_page')
    produtos_page_obj = produto_paginator.get_page(produto_page_number)
    
    vendedores_list = Vendedor.objects.filter(user=request.user).order_by('-created_at')
    # O filtro de data não se aplica a vendedores se estiver buscando produtos
    if start_date and not search_product:
        vendedores_list = vendedores_list.filter(created_at__date__gte=start_date)
    if end_date and not search_product:
        vendedores_list = vendedores_list.filter(created_at__date__lte=end_date)

    vendedor_paginator = Paginator(vendedores_list, 10)
    vendedor_page_number = request.GET.get('vendedor_page')
    vendedores_page_obj = vendedor_paginator.get_page(vendedor_page_number)

    # --- LÓGICA DE CLIENTES (NOVO) ---
    search_client = request.GET.get('search_client', '')
    clientes_list = Cliente.objects.filter(user=request.user).order_by('nome')
    
    if search_client:
        clientes_list = clientes_list.filter(
            Q(nome__icontains=search_client) | 
            Q(cpf_cnpj__icontains=search_client) |
            Q(email__icontains=search_client)
        )

    cliente_paginator = Paginator(clientes_list, 10)
    cliente_page_number = request.GET.get('cliente_page')
    clientes_page_obj = cliente_paginator.get_page(cliente_page_number)
    # ---------------------------------
    
    context = {
        'produto_form': produto_form, 
        'vendedor_form': vendedor_form,
        'produtos': produtos_page_obj, 
        'vendedores': vendedores_page_obj,
        'clientes': clientes_page_obj,
        'start_date': start_date, 
        'end_date': end_date,
        'search_product': search_product,
        'search_client': search_client,
    }
    return render(request, 'accounts/comercial_cadastros.html', context)



@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_cadastros_comercial')
def editar_produto_view(request, pk):
    # Busca o produto garantindo que pertence ao usuário logado
    produto = get_object_or_404(ProdutoServico, pk=pk, user=request.user)
    
    if request.method == 'POST':
        # Lógica ao SALVAR (Envia o user=request.user)
        form = ProdutoServicoForm(request.POST, instance=produto, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Produto atualizado com sucesso!')
            return redirect('comercial_cadastros')
    else:
        # Lógica ao ABRIR A TELA (GET) - Aqui estava faltando o user=request.user provavelmente
        form = ProdutoServicoForm(instance=produto, user=request.user)
    
    context = {
        'form': form,
        'item_name': produto.nome
    }
    return render(request, 'accounts/editar_item.html', context)

@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_cadastros_comercial')
def editar_vendedor_view(request, pk):
    vendedor = get_object_or_404(Vendedor, pk=pk, user=request.user)
    if request.method == 'POST':
        form = VendedorForm(request.POST, instance=vendedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vendedor atualizado com sucesso!')
            return redirect('comercial_cadastros')
    else:
        form = VendedorForm(instance=vendedor)

    context = {
        'form': form,
        'item_name': vendedor.nome
    }
    return render(request, 'accounts/editar_item.html', context)


@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_vendas')
def vendas_view(request):
    # --- ADIÇÃO 1: Lógica GET para buscar dados de uma venda para edição ---
    if 'fetch_sale_data' in request.GET:
        sale_id = request.GET.get('fetch_sale_data')
        try:
            venda = Venda.objects.select_related(
                'cliente', 'vendedor'
            ).prefetch_related(
                'itens__produto' # Carrega itens e produtos juntos
            ).get(id=sale_id, user=request.user)

            # Prepara a lista de itens no formato que o JavaScript espera
            itens_list = []
            for item in venda.itens.all():
                itens_list.append({
                    'id': item.produto.id,
                    'nome': item.produto.nome,
                    'quantidade': float(item.quantidade),
                    'preco': float(item.preco_unitario),
                    'desconto': float(item.desconto_item),
                    'estoque': item.produto.estoque_atual if item.produto.tipo == 'PRODUTO' else None,
                    'tipo': item.produto.tipo
                })

            # Monta o dicionário com os dados
            sale_data = {
                'venda_id': venda.id,
                'cliente_id': venda.cliente.id,
                'vendedor_id': venda.vendedor.id if venda.vendedor else None,
                'cidade': venda.cidade,
                'estado': venda.estado,
                'itens': itens_list,
                'cliente_details': {
                    'cpf_cnpj': venda.cliente.cpf_cnpj or '',
                    'email': venda.cliente.email or '',
                    'telefone': venda.cliente.telefone or '',
                    'endereco': venda.cliente.endereco or ''
                }
            }
            # O return DEVE estar dentro do try
            return JsonResponse({'status': 'success', 'sale_data': sale_data}) 
            
        except Venda.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Venda não encontrada.'}, status=404)
        except Exception as e:
            # Este bloco é que está enviando o erro para você
            return JsonResponse({'status': 'error', 'message': f'Erro interno no Python: {str(e)}'}, status=500)
    # --- FIM DA ADIÇÃO 1 ---
    if request.method == 'POST':
        # --- ADIÇÃO 2: Lógica POST para exclusão individual ---
        if 'action_delete_sale' in request.POST:
            sale_id = request.POST.get('action_delete_sale')
            venda = get_object_or_404(Venda, id=sale_id, user=request.user)
            try:
                with transaction.atomic():
                    # Restaura o estoque
                    for item in venda.itens.all():
                        if item.produto.tipo == 'PRODUTO':
                            item.produto.estoque_atual += item.quantidade
                            item.produto.save()
                    # Exclui a venda (isso também excluirá ItemVenda e PagamentoVenda associados)
                    venda.delete()
                    # Tenta excluir a conta a receber associada (se existir)
                    # Ajuste 'name' se o nome da conta a receber for diferente
                    ReceivableAccount.objects.filter(user=request.user, name=f"Ref. Venda #{sale_id}").delete()
                    messages.success(request, f'Venda #{sale_id} excluída com sucesso e estoque restaurado.')
            except Exception as e:
                messages.error(request, f'Erro ao excluir a venda #{sale_id}: {str(e)}')
            return redirect('vendas')
        # --- FIM DA ADIÇÃO 2 ---
        # --- INÍCIO: LÓGICA DE EXCLUSÃO DE VENDAS ---
        if 'delete_selected_sales' in request.POST:
            sale_ids = request.POST.getlist('sale_ids')
            if sale_ids:
                vendas_para_excluir = Venda.objects.filter(id__in=sale_ids, user=request.user)
                
                with transaction.atomic():
                    # Restaura o estoque dos produtos
                    for venda in vendas_para_excluir:
                        for item in venda.itens.all():
                            if item.produto.tipo == 'PRODUTO':
                                item.produto.estoque_atual += item.quantidade
                                item.produto.save()
                    
                    # Exclui as vendas
                    count, _ = vendas_para_excluir.delete()
                    messages.success(request, f'{count} venda(s) selecionada(s) foram excluídas e o estoque foi restaurado.')
            
            return redirect('vendas')
        # --- FIM: LÓGICA DE EXCLUSÃO DE VENDAS ---

        # Lógica de criação de venda via AJAX
        try:
            data = json.loads(request.body)
            original_sale_id = data.get('editing_sale_id')
            # Em accounts/views.py, substitua o bloco with transaction.atomic(): dentro da vendas_view

            with transaction.atomic():
                # Coleta de todos os dados do request
                cliente_id = data.get('cliente')
                vendedor_id = data.get('vendedor')
                cidade = data.get('cidade')
                estado = data.get('estado')
                itens_venda = data.get('itens', [])
                pagamentos_venda = data.get('pagamentos', [])
                due_date_str = data.get('due_date')
                
                # Validação
                if not all([cliente_id, itens_venda, pagamentos_venda, due_date_str, cidade, estado]):
                    return JsonResponse({'status': 'error', 'message': 'Dados incompletos. Cliente, Itens, Pagamento, Data, Cidade e Estado são obrigatórios.'}, status=400)

                cliente = get_object_or_404(Cliente, id=cliente_id, user=request.user)
                vendedor = get_object_or_404(Vendedor, id=vendedor_id, user=request.user) if vendedor_id else None

                # --- INÍCIO DA NOVA LÓGICA DE EDIÇÃO (UPDATE) ---
                if original_sale_id:
                    # Se estamos editando, ATUALIZAMOS a venda
                    venda = get_object_or_404(Venda, id=original_sale_id, user=request.user)
                    
                    # 1. Restaura o estoque antigo
                    for item in venda.itens.all():
                        if item.produto.tipo == 'PRODUTO':
                            item.produto.estoque_atual += item.quantidade
                            item.produto.save()
                    
                    # 2. Exclui itens, pagamentos e contas a receber antigas
                    venda.itens.all().delete()
                    venda.pagamentos.all().delete()
                    # Graças ao Passo 1, agora podemos encontrar a conta a receber antiga pelo ID da venda
                    ReceivableAccount.objects.filter(
                        user=request.user,
                        description__startswith=f"Ref. Venda #{venda.id}:"
                    ).delete()
                    
                    # 3. Atualiza os dados da Venda existente
                    venda.cliente = cliente
                    venda.vendedor = vendedor
                    venda.cidade = cidade
                    venda.estado = estado
                    venda.valor_total_bruto = data.get('total_bruto', 0)
                    venda.desconto_geral = data.get('desconto_total', 0)
                    venda.valor_total_liquido = data.get('total_liquido', 0)
                    venda.status = 'FINALIZADA' # Atualiza o status
                    venda.data_venda = timezone.now() # Atualiza a data para a da finalização
                    venda.save()
                    
                    # Define a 'nova_venda' como a venda que acabamos de atualizar
                    nova_venda = venda
                    success_message = f'Venda #{nova_venda.id} (originada do orçamento) foi atualizada e finalizada com sucesso!'
                
                else:
                    # Se não estamos editando, CRIAMOS uma nova venda (lógica original)
                    nova_venda = Venda.objects.create(
                        user=request.user,
                        cliente=cliente,
                        vendedor=vendedor,
                        data_venda=timezone.now(),
                        cidade=cidade,
                        estado=estado,
                        valor_total_bruto=data.get('total_bruto', 0),
                        desconto_geral=data.get('desconto_total', 0),
                        valor_total_liquido=data.get('total_liquido', 0),
                        status='FINALIZADA'
                    )
                    success_message = f'Venda #{nova_venda.id} finalizada e enviada para Contas a Receber!'
                # --- FIM DA NOVA LÓGICA DE EDIÇÃO (UPDATE) ---

                # O restante da lógica para criar Itens, Pagamentos e Contas a Receber
                # agora se aplica tanto à Venda nova quanto à Venda atualizada.
                
                description_parts = []
                tipos_de_item = set()
                
                for item_data in itens_venda:
                    produto = get_object_or_404(ProdutoServico, id=item_data['id'], user=request.user)
                    quantidade = Decimal(item_data['quantidade'])
                    description_parts.append(f"{quantidade}x {produto.nome}")
                    tipos_de_item.add(produto.tipo)
                    
                    if produto.tipo == 'PRODUTO' and produto.estoque_atual < quantidade:
                        raise Exception(f'Estoque insuficiente para {produto.nome}.')
                        
                    ItemVenda.objects.create(
                        venda=nova_venda,
                        produto=produto,
                        quantidade=quantidade,
                        preco_unitario=item_data['preco'],
                        desconto_item=item_data.get('desconto', 0)
                    )
                    if produto.tipo == 'PRODUTO':
                        produto.estoque_atual -= quantidade
                        produto.save()
                        
                for pag_data in pagamentos_venda:
                    PagamentoVenda.objects.create(
                        venda=nova_venda,
                        forma_pagamento=pag_data['forma'],
                        valor=pag_data['valor'],
                        parcelas=pag_data.get('parcelas', 1)
                    )
                    
                category_name = "Venda"
                if len(tipos_de_item) > 1: category_name = "Venda de Produtos e Serviços"
                elif 'PRODUTO' in tipos_de_item: category_name = "Venda de Produtos"
                elif 'SERVICO' in tipos_de_item: category_name = "Venda de Serviços"

                venda_category, _ = Category.objects.get_or_create(name=category_name)
                # A descrição da conta a receber AGORA USA o ID da venda (seja ela nova ou atualizada)
                receivable_description = f"Ref. Venda #{nova_venda.id}: " + ", ".join(description_parts)
                
                # ▼▼▼ Lógica de criação de Contas a Receber (sem alterações) ▼▼▼
                parcelas = data.get('parcelas', 1)
                payment_method = pagamentos_venda[0]['forma'] if pagamentos_venda else 'DINHEIRO'

                if payment_method == 'BOLETO' and parcelas > 1:
                    valor_total = Decimal(str(data.get('total_liquido', '0.0')))
                    valor_parcela = Decimal(round(valor_total / parcelas, 2))
                    primeira_data_vencimento = parse_date(due_date_str)
                    total_alocado = Decimal('0.00')

                    for i in range(parcelas):
                        if i == parcelas - 1:
                            valor_da_parcela_atual = valor_total - total_alocado
                        else:
                            valor_da_parcela_atual = valor_parcela
                            total_alocado += valor_parcela
                        data_vencimento_parcela = primeira_data_vencimento + relativedelta(months=i)
                        description_parcela = f"Parcela {i+1}/{parcelas} - {receivable_description}" # Usa a descrição base

                        ReceivableAccount.objects.create(
                            user=request.user, name=cliente.nome, description=description_parcela,
                            due_date=data_vencimento_parcela, amount=valor_da_parcela_atual,
                            category=venda_category, dre_area='BRUTA',
                            payment_method=payment_method, occurrence='AVULSO', is_received=False
                        )
                else:
                    ReceivableAccount.objects.create(
                        user=request.user, name=cliente.nome, description=receivable_description,
                        due_date=due_date_str, amount=nova_venda.valor_total_liquido,
                        category=venda_category, dre_area='BRUTA',
                        payment_method=payment_method, occurrence='AVULSO', is_received=False
                    )
                # ▲▲▲ Fim da Lógica de Contas a Receber ▲▲▲
                
            # Mensagem de sucesso (agora usa a variável definida na lógica)
            messages.success(request, success_message)
            return JsonResponse({'status':'success','redirect_url':request.path})
        except (json.JSONDecodeError, AttributeError):
             messages.error(request, 'Ocorreu um erro inesperado.')
             return redirect('vendas')
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    # ▼▼▼ ADICIONE O BLOCO DE CÓDIGO DE EXPORTAÇÃO AQUI ▼▼▼
    # ▼▼▼ SUBSTITUA O BLOCO DE EXPORTAÇÃO ANTIGO POR ESTE NOVO BLOCO COMPLETO ▼▼▼
    if 'export_excel' in request.GET:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        vendas_list = Venda.objects.filter(user=request.user).exclude(status='SUBSTITUIDA').order_by('-data_venda')
        if start_date_str:
            vendas_list = vendas_list.filter(data_venda__date__gte=start_date_str)
        if end_date_str:
            vendas_list = vendas_list.filter(data_venda__date__lte=end_date_str)

        # --- NOVO: Calcular os totais antes de criar a planilha ---
        totals = vendas_list.aggregate(
            total_bruto=Sum('valor_total_bruto'),
            total_desconto=Sum('desconto_geral'),
            total_liquido=Sum('valor_total_liquido')
        )
        total_bruto_sum = totals.get('total_bruto') or 0
        total_desconto_sum = totals.get('total_desconto') or 0
        total_liquido_sum = totals.get('total_liquido') or 0

        # Lógica para criar o arquivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="vendas_realizadas.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendas Realizadas"

        

        # Cabeçalhos
        headers = ['#ID', 'Cliente', 'Vendedor', 'Data', 'Itens da Venda', 'Valor Bruto', 'Descontos', 'Valor Líquido', 'Status']
        ws.append(headers)

        # Aplica negrito ao cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Dados
        for venda in vendas_list:
            itens_str = "\n".join([
                f"{item.quantidade}x {item.produto.nome}" for item in venda.itens.all()
            ])
            
            ws.append([
                venda.id,
                venda.cliente.nome,
                venda.vendedor.nome if venda.vendedor else '-',
                timezone.localtime(venda.data_venda).strftime('%d/%m/%Y %H:%M'),
                itens_str,
                venda.valor_total_bruto,
                venda.desconto_geral,
                venda.valor_total_liquido,
                venda.get_status_display()
            ])
        
        # --- NOVO: Adicionar uma linha em branco e a linha de totais ---
        ws.append([]) # Linha em branco para espaçamento
        
        total_row_data = ['', '', '', '', 'TOTAIS:', total_bruto_sum, total_desconto_sum, total_liquido_sum, '']
        ws.append(total_row_data)

        # Aplica negrito à linha de totais
        last_row_index = ws.max_row
        for cell in ws[last_row_index]:
            cell.font = Font(bold=True)
            if cell.column_letter in ['E']: # Alinha o texto "TOTAIS:" à direita
                 cell.alignment = Alignment(horizontal='right')

        # --- CÓDIGO CORRIGIDO: Formatar CÉLULAS como moeda e ajustar largura ---
        currency_format = '"R$ "#,##0.00'

        # Itera da linha 2 (após o cabeçalho) até a última linha (incluindo os totais)
        # e aplica o formato de moeda em cada célula das colunas F, G e H.
        for row_index in range(2, ws.max_row + 1):
            for col_letter in ['F', 'G', 'H']:
                ws[f'{col_letter}{row_index}'].number_format = currency_format

        # Ajusta a largura das colunas (esta parte continua igual)
        ws.column_dimensions['B'].width = 35 # Cliente
        ws.column_dimensions['C'].width = 35 # Vendedor
        ws.column_dimensions['D'].width = 20 # Data
        ws.column_dimensions['E'].width = 45 # Itens da Venda
        ws.column_dimensions['F'].width = 18 # Valor Bruto
        ws.column_dimensions['G'].width = 18 # Descontos
        ws.column_dimensions['H'].width = 18 # Valor Líquido
        
        # Ajusta a altura da linha para o texto com quebra de linha
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 2): # Exclui a linha de total
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "\n" in cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        wb.save(response)
        return response
    # ▲▲▲ FIM DO BLOCO DE SUBSTITUIÇÃO ▲▲▲
    # --- INÍCIO DA LÓGICA GET CORRIGIDA E MELHORADA ---
    
    # Pega as datas do formulário
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Define um período padrão (últimos 30 dias) se nenhuma data for fornecida
    if not start_date_str and not end_date_str:
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
        # Converte para string para preencher o formulário
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')
    
    # Filtra a lista de vendas
    vendas_list = Venda.objects.filter(user=request.user).order_by('-data_venda')
    if start_date_str:
        vendas_list = vendas_list.filter(data_venda__date__gte=start_date_str)
    if end_date_str:
        vendas_list = vendas_list.filter(data_venda__date__lte=end_date_str)
    # ▼▼▼ ADICIONE O CÁLCULO DO TOTAL AQUI ▼▼▼
    total_valor_liquido_periodo = vendas_list.aggregate(
        total=Sum('valor_total_liquido')
    )['total'] or Decimal('0.00')
    # Lógica de paginação (sem alterações)
    paginator = Paginator(vendas_list, 10)
    page_number = request.GET.get('page')
    vendas_page_obj = paginator.get_page(page_number)

    # Lógica para carregar dados para o formulário de nova venda (sem alterações)
    clientes = Cliente.objects.filter(user=request.user)
    vendedores = Vendedor.objects.filter(user=request.user)
    produtos = ProdutoServico.objects.filter(user=request.user)
    produtos_json = json.dumps([{'id':p.id,'nome':p.nome,'codigo':p.codigo,'preco':float(p.preco_venda),'estoque':p.estoque_atual,'tipo':p.tipo} for p in produtos])

    # Contexto final para o template
    context = {
        'clientes': clientes,
        'vendedores': vendedores,
        'produtos_json': produtos_json,
        'payment_methods': PAYMENT_METHODS,
        'vendas': vendas_page_obj,
        'start_date': start_date_str, # Passa as strings para preencher o form
        'end_date': end_date_str,   # Passa as strings para preencher o form
        'total_valor_liquido_periodo': total_valor_liquido_periodo,
    }
    return render(request, 'accounts/vendas.html', context)
    # --- FIM DA LÓGICA GET CORRIGIDA E MELHORADA ---


# accounts/views.py

from django.http import JsonResponse
from .forms import ClienteForm # Certifique-se que está importado

def cadastrar_cliente_rapido(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST, user=request.user)
        if form.is_valid():
            # 1. Cria o objeto na memória, mas NÃO salva no banco ainda (commit=False)
            cliente = form.save(commit=False)
            
            # 2. Preenche manualmente quem é o dono desse cliente (o usuário logado)
            cliente.user = request.user 
            
            # 3. Agora sim, salva no banco de dados com o ID do usuário preenchido
            cliente.save()

            # --- O restante da função continua igual ---
            endereco_visual = f"{cliente.logradouro}, {cliente.numero} - {cliente.bairro}"
            if cliente.cidade:
                endereco_visual += f" - {cliente.cidade}/{cliente.uf}"

            return JsonResponse({
                'status': 'success',
                'cliente': {
                    'id': cliente.id,
                    'nome': cliente.nome,
                    'cpf_cnpj': cliente.cpf_cnpj or '',
                    'email': cliente.email or '',
                    'telefone': cliente.telefone or '',
                    'endereco': endereco_visual
                }
            })
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors})
    return JsonResponse({'status': 'error', 'message': 'Método inválido'})

# Em accounts/views.py

@login_required
def bpo_dashboard_view(request):
    """
    Dashboard principal para o BPO Admin selecionar qual cliente gerenciar.
    """
    # 1. SEGURANÇA: Garante que não está 'impersonando' ninguém agora
    if request.is_managing:
        return redirect('stop_managing')

    # 2. SEGURANÇA: Verifica se é realmente um BPO
    try:
        if request.user.subscription.user_type != 'BPO':
            return redirect('home')
    except Subscription.DoesNotExist:
        return redirect('home')

    # 3. BUSCA NO BANCO: Pega todos os clientes vinculados
    client_links = BPOClientLink.objects.filter(bpo_admin=request.user).select_related('client', 'client__subscription')

    # --- [NOVO BLOCO] CÁLCULO DE LIMITES ---
    current_count = client_links.count()
    limit = request.user.subscription.client_limit
    can_add = current_count < limit
    # ---------------------------------------

    # 4. PREPARA A LISTA PARA EXIBIR NA TABELA
    clients_data = []
    for link in client_links:
        # Pega o nome da empresa do perfil ou usa o username
        nome_exibicao = link.client.username
        if hasattr(link.client, 'company_profile'):
            nome_exibicao = link.client.company_profile.nome_empresa or link.client.username

        clients_data.append({
            'user_id': link.client.id,
            'username': nome_exibicao,
            'email': link.client.email,
            'subscription_status': link.client.subscription.status if hasattr(link.client, 'subscription') else 'N/A'
        })

    # 5. CONTEXTO FINAL (Com as novas variáveis)
    context = {
        'clients': clients_data,
        'is_bpo_dashboard': True, # Para o template saber onde está
        
        # Novas variáveis para o botão "+ Adicionar Cliente"
        'current_count': current_count,
        'limit': limit,
        'can_add': can_add,
    }
    return render(request, 'accounts/bpo_dashboard.html', context)

@login_required
def switch_to_client_view(request, client_id):
    """
    Troca a sessão para a do cliente selecionado.
    """
    # Garante que o usuário logado é um BPO Admin
    try:
        if request.user.subscription.user_type != 'BPO':
            messages.error(request, "Acesso não autorizado.")
            return redirect('home')
    except Subscription.DoesNotExist:
        messages.error(request, "Acesso não autorizado.")
        return redirect('home')

    # Verifica se o BPO Admin tem permissão para gerenciar este cliente
    try:
        link = BPOClientLink.objects.get(bpo_admin=request.user, client_id=client_id)

        # Define as variáveis de sessão que o middleware irá ler
        request.session['real_user_id'] = request.user.id
        request.session['managed_user_id'] = link.client.id

        messages.success(request, f"Você agora está gerenciando: {link.client.username}")
        # Redireciona para a 'home' (que agora será a home do cliente)
        return redirect('home')

    except BPOClientLink.DoesNotExist:
        messages.error(request, "Você não tem permissão para gerenciar este cliente.")
        return redirect('bpo_dashboard')
    except User.DoesNotExist:
        messages.error(request, "Cliente não encontrado.")
        return redirect('bpo_dashboard')


@login_required
def stop_managing_view(request):
    """
    Limpa a sessão de gerenciamento e retorna o BPO ao seu dashboard.
    """
    if 'real_user_id' in request.session:
        del request.session['real_user_id']
    if 'managed_user_id' in request.session:
        del request.session['managed_user_id']

    messages.info(request, "Você voltou para sua conta principal.")
    return redirect('bpo_dashboard')


def logout_view(request):
    logout(request)
    return redirect('login')



@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_metas_comerciais')
def metas_comerciais_view(request):
    form = MetaFaturamentoForm(user=request.user)
    
    if request.method == 'POST':
        post_data = request.POST.copy()
        mes_ano_str = post_data.get('mes_ano')
        if mes_ano_str:
            post_data['mes_ano'] = f"{mes_ano_str}-01"
        
        form = MetaFaturamentoForm(post_data, user=request.user)
        
        if form.is_valid():
            alvo_id = form.cleaned_data['alvo']
            mes_ano_completo = form.cleaned_data['mes_ano']
            valor_meta = form.cleaned_data['valor_meta']
            vendedor_instance = None
            if alvo_id != 'empresa':
                vendedor_instance = get_object_or_404(Vendedor, id=alvo_id, user=request.user)
            
            MetaFaturamento.objects.update_or_create(
                user=request.user,
                vendedor=vendedor_instance,
                mes_ano=mes_ano_completo,
                defaults={'valor_meta': valor_meta}
            )
            messages.success(request, 'Meta salva com sucesso!')
            ano_selecionado = request.GET.get('ano', timezone.now().year)
            return redirect(f"{request.path}?ano={ano_selecionado}")

    # --- INÍCIO DAS CORREÇÕES NA LÓGICA GET ---
    
    # 1. Lógica do Filtro de Ano (CORRIGIDO)
    # --- INÍCIO DA CORREÇÃO DA LÓGICA GET ---
    
    # 1. Lógica do Filtro de Ano (VERSÃO CORRIGIDA)
    # Primeiro, buscamos os objetos de data do banco de dados.
    anos_queryset = MetaFaturamento.objects.filter(user=request.user).dates('mes_ano', 'year', order='DESC')
    # AGORA, A CORREÇÃO: convertemos os objetos de data em uma lista de números inteiros.
    anos_disponiveis = [data.year for data in anos_queryset]
    
    # Adiciona o ano corrente à lista se ele ainda não existir
    ano_corrente = timezone.now().year
    if ano_corrente not in anos_disponiveis:
        anos_disponiveis.insert(0, ano_corrente)
        
    # Pega o ano selecionado e garante que ele seja um número inteiro
    ano_selecionado_str = request.GET.get('ano', str(ano_corrente))
    ano_selecionado = int(ano_selecionado_str)
    
    # 2. Filtra as metas pelo ano selecionado
    metas = MetaFaturamento.objects.filter(user=request.user, mes_ano__year=ano_selecionado)

    # 3. Calcula o valor acumulado
    total_anual_empresa = metas.filter(vendedor__isnull=True).aggregate(
        total=Sum('valor_meta')
    )['total'] or Decimal('0.00')

    # 4. Agrupa metas e calcula percentuais com meses em Português (CORRIGIDO)
    
    # Dicionário para traduzir os meses
    MESES_PT = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
        7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }

    metas_agrupadas = defaultdict(lambda: {'empresa': None, 'vendedores': []})
    
    for meta in metas.order_by('mes_ano'):
        # Formata o mês manualmente usando o dicionário
        mes_nome = MESES_PT[meta.mes_ano.month]
        mes_str = f"{mes_nome}/{meta.mes_ano.year}"
        
        if meta.vendedor:
            metas_agrupadas[mes_str]['vendedores'].append(meta)
        else:
            metas_agrupadas[mes_str]['empresa'] = meta

    for mes, dados_meta in metas_agrupadas.items():
        meta_empresa = dados_meta.get('empresa')
        if meta_empresa and meta_empresa.valor_meta > 0:
            for meta_vendedor in dados_meta['vendedores']:
                percentual = (meta_vendedor.valor_meta / meta_empresa.valor_meta) * 100
                meta_vendedor.percentual = f"{percentual:.2f}%"
        else:
            for meta_vendedor in dados_meta['vendedores']:
                meta_vendedor.percentual = "N/A"
    # --- FIM DAS CORREÇÕES ---

    # --- INÍCIO DA LÓGICA PARA O NOVO GRÁFICO ANUAL ---
    
    # 1. Busca as metas da empresa para o ano selecionado
    metas_anuais_query = MetaFaturamento.objects.filter(
        user=request.user,
        vendedor__isnull=True,
        mes_ano__year=ano_selecionado
    )
    
    # 2. Cria um dicionário para mapear mês -> valor da meta
    metas_por_mes = {meta.mes_ano.month: float(meta.valor_meta) for meta in metas_anuais_query}
    
    # 3. Prepara as listas para o Chart.js, garantindo 12 meses
    labels_grafico_anual = list(MESES_PT.values()) # Usa o dicionário de meses em português
    data_grafico_anual = []
    for i in range(1, 13):
        # Adiciona o valor da meta do mês, ou 0 se não houver meta cadastrada
        data_grafico_anual.append(metas_por_mes.get(i, 0))

    # --- FIM DA LÓGICA PARA O NOVO GRÁFICO ANUAL ---
            
    context = {
        'form': form,
        'metas_agrupadas': dict(metas_agrupadas),
        'anos_disponiveis': anos_disponiveis,
        'ano_selecionado': ano_selecionado,
        'total_anual_empresa': total_anual_empresa,
        'labels_grafico_anual_json': json.dumps(labels_grafico_anual),
        'data_grafico_anual_json': json.dumps(data_grafico_anual),
    }
    return render(request, 'accounts/metas_comerciais.html', context)



def search_cities(request):
    query = request.GET.get('term', '')
    cidades = []
    if len(query) > 2: # Começa a buscar a partir de 3 caracteres
        cidades_qs = Cidade.objects.filter(nome__istartswith=query).select_related('estado')[:10] # Limita a 10 resultados
        for cidade in cidades_qs:
            cidades.append({
                'label': f"{cidade.nome}, {cidade.estado.uf}", # Texto que aparece na sugestão
                'value': cidade.nome, # Valor que vai para o input de cidade
                'estado': cidade.estado.uf # Valor que vai para o input de estado
            })
    return JsonResponse(cidades, safe=False)



@login_required
def chat_room_view(request, room_name):
    return render(request, 'accounts/chat.html', {
        'room_name': room_name
    })




@owner_required
@login_required
def assinatura_view(request):
    subscription, created = Subscription.objects.get_or_create(user=request.user)
    context = {
        'subscription': subscription
    }
    return render(request, 'accounts/assinatura.html', context)    



def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            # 1. Salva o formulário e cria o novo usuário
            user = form.save()
            
            # 2. CRIA AUTOMATICAMENTE A ASSINATURA 'EXPIRADA' PARA O NOVO USUÁRIO
            
            
            # 3. Loga o usuário automaticamente após o cadastro
            login(request, user)
            
            messages.success(request, 'Cadastro realizado com sucesso! Bem-vindo(a).')
            return redirect('assinatura')
    else:
        form = CustomUserCreationForm()
        
    return render(request, 'accounts/register.html', {'form': form})


@login_required
@check_employee_permission('can_access_painel_financeiro')
def gerar_laudo_financeiro(request):
    # --- 1. Definição das Datas (Estrutura Mantida) ---
    period = request.GET.get('period', '30')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if period == 'custom' and start_date_str and end_date_str:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
    else:
        end_date = datetime.today().date()
        days = int(period) if period in ['90', '180'] else 30
        start_date = end_date - timedelta(days=days)

    # Datas do período anterior para comparação (Month-over-Month ou Period-over-Period)
    duration = end_date - start_date
    prev_end_date = start_date - timedelta(days=1)
    prev_start_date = prev_end_date - duration

    # --- 2. Coleta de Dados (Estrutura Mantida - Regime de Caixa) ---
    
    # PERÍODO ATUAL
    entradas = ReceivableAccount.objects.filter(
        user=request.user, 
        is_received=True, 
        due_date__range=[start_date, end_date] 
    ).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    saidas = PayableAccount.objects.filter(
        user=request.user, 
        is_paid=True, 
        due_date__range=[start_date, end_date]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    geracao_caixa = entradas - saidas

    # PERÍODO ANTERIOR
    entradas_ant = ReceivableAccount.objects.filter(
        user=request.user, 
        is_received=True, 
        due_date__range=[prev_start_date, prev_end_date]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    saidas_ant = PayableAccount.objects.filter(
        user=request.user, 
        is_paid=True, 
        due_date__range=[prev_start_date, prev_end_date]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    geracao_caixa_ant = entradas_ant - saidas_ant

    # --- 3. Cálculos de Variação e KPI's (Expandido) ---
    
    # Variações Percentuais
    var_entradas = 0
    if entradas_ant > 0:
        var_entradas = ((entradas - entradas_ant) / entradas_ant) * 100
    
    var_saidas = 0
    if saidas_ant > 0:
        var_saidas = ((saidas - saidas_ant) / saidas_ant) * 100

    # Variação do Resultado Líquido
    var_caixa = 0
    if abs(geracao_caixa_ant) > 0:
        var_caixa = ((geracao_caixa - geracao_caixa_ant) / abs(geracao_caixa_ant)) * 100

    # Margem de Caixa (Eficiência de conversão de receita em caixa livre)
    margem_caixa = 0
    if entradas > 0:
        margem_caixa = (geracao_caixa / entradas) * 100

    # Gap de Crescimento (Receita vs Despesa) - Positivo indica ganho de alavancagem
    gap_crescimento = var_entradas - var_saidas

    # --- 4. Geração do HTML do Laudo (Estilo Executivo/Diretoria) ---
    
    periodo_str = f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
    laudo_html = f"<h2>Relatório Executivo Financeiro<br><small style='font-size: 0.6em; color: #666;'>Período de Análise: {periodo_str}</small></h2><hr>"
    
    # Seção 1: Sumário Executivo
    laudo_html += "<h3><strong>1. Sumário Executivo de Caixa</strong></h3>"
    
    saldo_class = "text-success" if geracao_caixa >= 0 else "text-danger"
    sinal_saldo = "SUPERÁVIT" if geracao_caixa >= 0 else "DÉFICIT"
    cor_margem = "green" if margem_caixa >= 10 else ("orange" if margem_caixa > 0 else "red")
    
    laudo_html += f"""
    <p>O período encerrou com um <strong>{sinal_saldo} OPERACIONAL</strong> de <span class="{saldo_class}" style="font-size: 1.2em; font-weight: bold;">R$ {geracao_caixa:,.2f}</span>.</p>
    <ul style="list-style-type: none; padding: 0;">
        <li><strong>Entradas Totais (Recebimentos):</strong> R$ {entradas:,.2f}</li>
        <li><strong>Saídas Totais (Pagamentos):</strong> R$ {saidas:,.2f}</li>
        <li><strong>Margem de Caixa:</strong> <span style="color: {cor_margem}; font-weight: bold;">{margem_caixa:.1f}%</span> <small>(Percentual da receita que sobrou em caixa)</small></li>
    </ul>
    """

    # Seção 2: Análise de Tendência e Eficiência (Cruzamento de Dados)
    laudo_html += "<h4>📊 2. Análise de Tendência e Eficiência</h4>"
    
    # Análise das Entradas
    if var_entradas > 5:
        analise_ent = f"As entradas apresentaram uma <strong>expansão sólida de {var_entradas:.1f}%</strong> frente ao período anterior, indicando aquecimento nas vendas ou melhora na inadimplência."
    elif var_entradas >= -5:
        analise_ent = f"As entradas mantiveram-se <strong>estáveis ({var_entradas:.1f}%)</strong>, sugerindo manutenção do patamar de faturamento."
    else:
        analise_ent = f"Houve uma <strong>retração de {var_entradas:.1f}%</strong> nas entradas, o que exige investigação sobre sazonalidade ou perda de performance comercial."

    laudo_html += f"<p>{analise_ent}</p>"

    # Análise Cruzada (Receita x Despesa) - O ponto chave para a diretoria
    if var_entradas > var_saidas:
        laudo_html += f"""
        <p>✅ <strong>Ganho de Alavancagem:</strong> Positivamente, as receitas cresceram acima das despesas (Gap de {gap_crescimento:.1f} p.p.). Isso demonstra diluição de custos fixos e aumento da eficiência operacional no período.</p>
        """
    elif var_saidas > var_entradas:
        laudo_html += f"""
        <p>⚠️ <strong>Atenção à Eficiência:</strong> As despesas cresceram em ritmo acelerado ({var_saidas:.1f}%), superando a variação das receitas ({var_entradas:.1f}%). É crucial auditar os custos variáveis e fixos para evitar erosão da margem.</p>
        """
    
    # Seção 3: Diagnóstico Estratégico
    laudo_html += "<h4>🎯 3. Diagnóstico e Plano de Ação</h4>"

    if geracao_caixa < 0:
        laudo_html += """
        <p><strong>Situação: <span style='color:red'>CONSUMO DE CAIXA (BURN RATE).</span></strong></p>
        <p>A operação não foi capaz de se autofinanciar neste período. Dependência de capital de terceiros ou reservas.</p>
        <p><strong>Plano Recomendado:</strong>
        1. Suspender investimentos não essenciais imediatamente.<br>
        2. Renegociar prazos com fornecedores ABC (Curva A).<br>
        3. Realizar ação comercial de 'Liquidez Imediata' para antecipar recebíveis.</p>
        """
    elif margem_caixa < 10:
        laudo_html += f"""
        <p><strong>Situação: <span style='color:orange'>EQUILÍBRIO TENSO (Margem Baixa).</span></strong></p>
        <p>A operação é sustentável, mas vulnerável a imprevistos. A margem de {margem_caixa:.1f}% deixa pouco espaço para reinvestimento.</p>
        <p><strong>Plano Recomendado:</strong>
        1. Focar em produtos/serviços de maior margem de contribuição.<br>
        2. Revisar contratos recorrentes em busca de saving de 5-10%.<br>
        3. Evitar novas dívidas de curto prazo.</p>
        """
    else:
        laudo_html += f"""
        <p><strong>Situação: <span style='color:green'>SOLIDEZ FINANCEIRA (Alta Liquidez).</span></strong></p>
        <p>Excelente performance com geração de caixa de {margem_caixa:.1f}%. A empresa demonstra capacidade de investimento sem comprometer o fluxo.</p>
        <p><strong>Plano Recomendado:</strong>
        1. Constituir ou reforçar reserva de emergência (mínimo 3 meses de custo fixo).<br>
        2. Avaliar antecipação de pagamentos com desconto junto a fornecedores.<br>
        3. Planejar investimentos estratégicos para expansão (CAPEX).</p>
        """

    # Rodapé Técnico
    laudo_html += "<hr><p style='font-size: 0.8em; color: #888;'><em>Relatório gerado via Inteligência de Dados Financlass. Base de cálculo: Movimentações financeiras efetivamente liquidadas (Regime de Caixa).</em></p>"

    return JsonResponse({'laudo_html': laudo_html})



@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_precificacao')
def precificacao_view(request):
    # --- Lógica POST (sem alterações) ---
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            produto_id = data.get('produto_id')
            if not produto_id:
                return JsonResponse({'status': 'error', 'message': 'Produto não selecionado.'}, status=400)
            produto = ProdutoServico.objects.get(id=produto_id, user=request.user)
            Precificacao.objects.create(
                user=request.user,
                produto=produto,
                preco_custo=Decimal(data.get('preco_custo', 0)),
                perc_despesas_fixas=Decimal(data.get('perc_despesas_fixas', 0)),
                perc_comissao=Decimal(data.get('perc_comissao', 0)),
                perc_impostos=Decimal(data.get('perc_impostos', 0)),
                perc_lucro=Decimal(data.get('perc_lucro', 0)),
                preco_venda_sugerido=Decimal(data.get('preco_venda_sugerido', 0)),
                is_price_updated=False
            )
            return JsonResponse({'status': 'success', 'message': 'Precificação salva com sucesso!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # --- Lógica GET (ATUALIZADA) ---
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=90)

    # ... (cálculos de sugestões, sem alterações) ...
    receita_bruta = ReceivableAccount.objects.filter(
        user=request.user, due_date__range=[start_date, end_date]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    impostos_sobre_receita = PayableAccount.objects.filter(
        user=request.user, dre_area='DEDUCAO', due_date__range=[start_date, end_date]
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    receita_liquida = receita_bruta - impostos_sobre_receita
    despesas_fixas_totais = PayableAccount.objects.filter(
        user=request.user, cost_type='FIXO', due_date__range=[start_date, end_date]
    ).exclude(dre_area='NAO_CONSTAR').aggregate(total=Sum('amount'))['total'] or Decimal('0')
    suggested_fixed_cost_perc = (despesas_fixas_totais / receita_liquida * 100) if receita_liquida > 0 else Decimal('0')
    suggested_tax_perc = (impostos_sobre_receita / receita_bruta * 100) if receita_bruta > 0 else Decimal('0')
    
    produtos = ProdutoServico.objects.filter(user=request.user).order_by('nome')
    vendedores = Vendedor.objects.filter(user=request.user).order_by('nome')
    
    # --- INÍCIO DAS NOVAS LÓGICAS DE HISTÓRICO ---
    
    # 1. Busca o termo da pesquisa na URL
    search_query = request.GET.get('q', '')
    
    # 2. Busca o histórico e usa select_related('produto') para CORRIGIR o nome do produto
    historico_list = Precificacao.objects.filter(user=request.user).select_related('produto')
    
    # 3. Se houver uma busca, filtra a lista pelo nome do produto
    if search_query:
        historico_list = historico_list.filter(produto__nome__icontains=search_query)

    # 4. Aplica a paginação na lista (filtrada ou não)
    paginator = Paginator(historico_list, 10) # 10 itens por página
    page_number = request.GET.get('page')
    historico_page_obj = paginator.get_page(page_number)
    
    # --- FIM DAS NOVAS LÓGICAS ---

    produtos_json = json.dumps([
        {'id': p.id, 'nome': p.nome, 'preco_custo': float(p.preco_custo), 'preco_venda_atual': float(p.preco_venda)} 
        for p in produtos
    ])
    vendedores_json = json.dumps([
        {'id': v.id, 'nome': v.nome, 'comissao': float(v.comissao_percentual)}
        for v in vendedores
    ])

    context = {
        'produtos': produtos,
        'vendedores': vendedores,
        'produtos_json': produtos_json,
        'vendedores_json': vendedores_json,
        'suggested_fixed_cost_perc': f"{suggested_fixed_cost_perc:.2f}".replace(',', '.'),
        'suggested_tax_perc': f"{suggested_tax_perc:.2f}".replace(',', '.'),
        
        # --- VARIÁVEIS ATUALIZADAS PARA O TEMPLATE ---
        'historico_page_obj': historico_page_obj,
        'search_query': search_query,
    }
    return render(request, 'accounts/precificacao.html', context)




@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_orcamento_anual')
def orcamento_anual_view(request):
    # Lógica do formulário POST (continua a mesma)
    if request.method == 'POST':
        form_data = request.POST.copy()
        if 'mes_ano' in form_data and form_data['mes_ano']:
            form_data['mes_ano'] = f"{form_data['mes_ano']}-01"
        
        form = OrcamentoForm(form_data, user=request.user)
        if form.is_valid():
            category = form.cleaned_data['category']
            mes_ano = form.cleaned_data['mes_ano']
            valor_orcado = form.cleaned_data['valor_orcado']

            Orcamento.objects.update_or_create(
                user=request.user, category=category, mes_ano=mes_ano,
                defaults={'valor_orcado': valor_orcado}
            )
            messages.success(request, f"Orçamento para '{category.name}' em {mes_ano.strftime('%m/%Y')} salvo com sucesso!")
            return redirect(f"{request.path}?ano={mes_ano.year}")
        else:
            messages.error(request, f"Erro ao salvar o orçamento: {form.errors.as_text()}")

    # --- LÓGICA GET ATUALIZADA PARA SEPARAR RECEITAS E DESPESAS ---
    
    # Filtro de ano (continua o mesmo)
    anos_disponiveis = Orcamento.objects.filter(user=request.user).dates('mes_ano', 'year', order='DESC')
    anos_list = [d.year for d in anos_disponiveis]
    ano_corrente = datetime.now().year
    if ano_corrente not in anos_list:
        anos_list.insert(0, ano_corrente)
    
    ano_selecionado = int(request.GET.get('ano', str(ano_corrente)))

    # --- LÓGICA DE DESPESAS ---
    orcamentos_despesa = Orcamento.objects.filter(user=request.user, mes_ano__year=ano_selecionado)
    realizado_despesa = PayableAccount.objects.filter(
        user=request.user, is_paid=True, due_date__year=ano_selecionado
    ).annotate(month=TruncMonth('due_date')).values('category__name', 'month').annotate(total=Sum('amount'))

    expense_budget_data = defaultdict(lambda: {
        'orcado_mensal': [Decimal('0')] * 12, 'realizado_mensal': [Decimal('0')] * 12
    })

    # Preenche despesas realizadas
    for gasto in realizado_despesa:
        if gasto['category__name']:
            mes_index = gasto['month'].month - 1
            expense_budget_data[gasto['category__name']]['realizado_mensal'][mes_index] += gasto['total']

    # --- LÓGICA DE RECEITAS ---
    orcamentos_receita = Orcamento.objects.filter(user=request.user, mes_ano__year=ano_selecionado)
    realizado_receita = ReceivableAccount.objects.filter(
        user=request.user, is_received=True, due_date__year=ano_selecionado
    ).annotate(month=TruncMonth('due_date')).values('category__name', 'month').annotate(total=Sum('amount'))

    revenue_budget_data = defaultdict(lambda: {
        'orcado_mensal': [Decimal('0')] * 12, 'realizado_mensal': [Decimal('0')] * 12
    })
    
    # Preenche receitas realizadas
    for receita in realizado_receita:
        if receita['category__name']:
            mes_index = receita['month'].month - 1
            revenue_budget_data[receita['category__name']]['realizado_mensal'][mes_index] += receita['total']

    # --- DISTRIBUI OS ORÇAMENTOS PARA AS TABELAS CORRETAS ---
    # Pega IDs de categorias usadas em receitas para saber onde alocar o orçamento
    revenue_category_ids = set(ReceivableAccount.objects.filter(user=request.user).values_list('category_id', flat=True))

    for orcamento in orcamentos_despesa:
        if orcamento.category:
            mes_index = orcamento.mes_ano.month - 1
            # Se a categoria do orçamento for uma categoria de receita, coloque na tabela de receitas
            if orcamento.category.id in revenue_category_ids:
                revenue_budget_data[orcamento.category.name]['orcado_mensal'][mes_index] += orcamento.valor_orcado
            # Senão, coloque na tabela de despesas
            else:
                expense_budget_data[orcamento.category.name]['orcado_mensal'][mes_index] += orcamento.valor_orcado

    # Garante que todas as categorias com lançamentos (mesmo sem orçamento) apareçam
    all_expense_categories = Category.objects.filter(payableaccount__user=request.user).distinct()
    for cat in all_expense_categories:
        expense_budget_data[cat.name]
        
    all_revenue_categories = Category.objects.filter(receivableaccount__user=request.user).distinct()
    for cat in all_revenue_categories:
        revenue_budget_data[cat.name]

    # --- CÁLCULO DE TOTAIS (SEPARADAMENTE) ---
    def calcular_totais(data_dict):
        totais = {'orcado': [Decimal('0')] * 12, 'realizado': [Decimal('0')] * 12}
        for data in data_dict.values():
            data['total_orcado_ano'] = sum(data['orcado_mensal'])
            data['total_realizado_ano'] = sum(data['realizado_mensal'])
            for i in range(12):
                totais['orcado'][i] += data['orcado_mensal'][i]
                totais['realizado'][i] += data['realizado_mensal'][i]
        totais['total_orcado_ano'] = sum(totais['orcado'])
        totais['total_realizado_ano'] = sum(totais['realizado'])
        return totais

    totais_despesas = calcular_totais(expense_budget_data)
    totais_receitas = calcular_totais(revenue_budget_data)
    
    form = OrcamentoForm(user=request.user)

    context = {
        'form': form,
        'expense_budget_data': dict(sorted(expense_budget_data.items())),
        'revenue_budget_data': dict(sorted(revenue_budget_data.items())),
        'totais_despesas': totais_despesas,
        'totais_receitas': totais_receitas,
        'ano_selecionado': ano_selecionado,
        'anos_disponiveis': anos_list,
        'meses': ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez'],
    }
    return render(request, 'accounts/orcamento_anual.html', context)



# Adicione esta nova view ao seu arquivo
@login_required
def salvar_e_atualizar_precificacao_view(request):
    if request.method == 'POST':
        try:
            # Garante que as duas operações (salvar histórico e atualizar produto)
            # aconteçam juntas. Se uma falhar, a outra é desfeita.
            with transaction.atomic():
                data = json.loads(request.body)
                produto_id = data.get('produto_id')

                if not produto_id:
                    return JsonResponse({'status': 'error', 'message': 'Produto não selecionado.'}, status=400)

                # --- Ação 1: Encontrar o produto ---
                produto = get_object_or_404(ProdutoServico, id=produto_id, user=request.user)

                # --- Ação 2: Salvar o registro no histórico de precificação ---
                Precificacao.objects.create(
                    user=request.user,
                    produto=produto,
                    preco_custo=Decimal(data.get('preco_custo', 0)),
                    perc_despesas_fixas=Decimal(data.get('perc_despesas_fixas', 0)),
                    perc_comissao=Decimal(data.get('perc_comissao', 0)),
                    perc_impostos=Decimal(data.get('perc_impostos', 0)),
                    perc_lucro=Decimal(data.get('perc_lucro', 0)),
                    preco_venda_sugerido=Decimal(data.get('preco_venda_sugerido', 0)),
                    is_price_updated=True 
                    
                )

                # --- Ação 3: Atualizar o preço de venda do produto no cadastro ---
                novo_preco = Decimal(data.get('preco_venda_sugerido', 0))
                if novo_preco > 0:
                    produto.preco_venda = novo_preco
                    produto.save()

                return JsonResponse({'status': 'success', 'message': 'Precificação salva e preço do produto atualizado com sucesso!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # Se não for POST, redireciona para a página principal da precificação
    return redirect('precificacao')




@login_required
@check_employee_permission('can_access_painel_vendas')
def gerar_laudo_comercial(request):
    # 1. PEGAR OS FILTROS DE PERÍODO (código similar ao da view de faturamento)
    period = request.GET.get('period', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.now().date()
    vendas_filtradas = Venda.objects.filter(user=request.user)

    if period == 'custom' and start_date_str and end_date_str:
        start_date = parse_date(start_date_str)
        end_date = parse_date(end_date_str)
        vendas_filtradas = vendas_filtradas.filter(data_venda__date__range=[start_date, end_date])
    elif period.isdigit():
        days = int(period)
        end_date = today
        start_date = end_date - timedelta(days=days)
        vendas_filtradas = vendas_filtradas.filter(data_venda__date__range=[start_date, end_date])
    else: # 'all'
        try:
            start_date = Venda.objects.filter(user=request.user).earliest('data_venda').data_venda.date()
        except Venda.DoesNotExist:
            start_date = today
        end_date = today

    # 2. CALCULAR MÉTRICAS COMERCIAIS
    total_vendido = vendas_filtradas.aggregate(total=Sum('valor_total_liquido'))['total'] or Decimal('0')
    qtd_vendas = vendas_filtradas.count()
    ticket_medio = total_vendido / qtd_vendas if qtd_vendas > 0 else Decimal('0')
    
    top_produtos = ItemVenda.objects.filter(venda__in=vendas_filtradas).values('produto__nome').annotate(valor=Sum(F('quantidade') * F('preco_unitario'))).order_by('-valor')[:3]
    top_clientes = vendas_filtradas.values('cliente__nome').annotate(valor_total=Sum('valor_total_liquido')).order_by('-valor_total')[:3]
    top_vendedores = vendas_filtradas.filter(vendedor__isnull=False).values('vendedor__nome').annotate(valor_vendas=Sum('valor_total_liquido')).order_by('-valor_vendas')[:3]

    # 3. GERAR O TEXTO DO LAUDO
    periodo_str = f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
    laudo_html = f"<h2>Laudo Comercial<br><small>Período {periodo_str}</small></h2><hr>"
    laudo_html += "<h3><strong>Análise Sintética e Recomendações</strong></h3>"

    # Seção 1: Desempenho Geral
    laudo_html += "<h4>📈 1. Desempenho Geral de Vendas</h4>"
    laudo_html += f"<p><strong>Situação:</strong> No período analisado, o faturamento total foi de R$ {total_vendido:,.2f}, realizado através de {qtd_vendas} vendas, resultando em um ticket médio de R$ {ticket_medio:,.2f}.</p>"
    if ticket_medio > 500: # Exemplo de condição
        laudo_html += "<p><strong>Recomendação:</strong> ✅ Seu ticket médio é robusto. Continue focando em estratégias de up-sell e cross-sell para mantê-lo ou aumentá-lo.</p>"
    else:
        laudo_html += "<p><strong>Recomendação:</strong> 💡 O ticket médio pode ser otimizado. Avalie a criação de kits de produtos, ofertas 'compre junto' ou programas de fidelidade para incentivar compras de maior valor.</p>"

    # Seção 2: Análise de Produtos
    produtos_str = ", ".join([p['produto__nome'] for p in top_produtos]) or "N/A"
    laudo_html += "<h4>📦 2. Análise de Produtos e Serviços</h4>"
    laudo_html += f"<p><strong>Destaques:</strong> Seus produtos/serviços de maior faturamento foram: {produtos_str}.</p>"
    laudo_html += "<p><strong>Recomendação:</strong> 🚀 Destaque seus produtos campeões em campanhas de marketing. Para os demais, considere promoções ou revise a estratégia de precificação para aumentar a saída.</p>"

    # Seção 3: Análise de Clientes
    clientes_str = ", ".join([c['cliente__nome'] for c in top_clientes]) or "N/A"
    laudo_html += "<h4>👥 3. Análise de Clientes</h4>"
    laudo_html += f"<p><strong>Destaques:</strong> Seus principais clientes por faturamento foram: {clientes_str}.</p>"
    laudo_html += "<p><strong>Recomendação:</strong> 🎯 Fortaleça o relacionamento com seus clientes-chave. Um contato pós-venda ou uma oferta exclusiva pode garantir a recorrência e a fidelidade.</p>"

    # Seção 4: Análise de Vendedores
    if top_vendedores:
        vendedores_str = ", ".join([v['vendedor__nome'] for v in top_vendedores]) or "N/A"
        laudo_html += "<h4>🏆 4. Análise de Vendedores</h4>"
        laudo_html += f"<p><strong>Destaques:</strong> Seus vendedores com melhor desempenho no período foram: {vendedores_str}.</p>"
        laudo_html += "<p><strong>Recomendação:</strong> 👏 Reconheça o bom desempenho dos seus top vendedores. Analise suas estratégias e compartilhe as melhores práticas com o restante da equipe para nivelar o time por cima.</p>"

    # Conclusão
    laudo_html += "<h4><strong>Conclusão e Próximos Passos</strong></h4>"
    laudo_html += "<p>A análise comercial indica uma base sólida de produtos e clientes. O foco agora deve ser em aumentar o ticket médio e replicar as estratégias de sucesso dos principais vendedores para toda a equipe. Monitore a performance dos produtos de menor saída e crie ações para impulsioná-los.</p>"

    return JsonResponse({'laudo_html': laudo_html})



import logging # Garanta que logging está importado
logger = logging.getLogger(__name__)

@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_orcamentos_venda')
def orcamentos_venda_view(request):

    # --- INÍCIO: LÓGICA AJAX GET PARA BUSCAR DETALHES (NOVO) ---
    if 'fetch_orcamento_data' in request.GET:
        orcamento_id = request.GET.get('fetch_orcamento_data')
        try:
            orcamento = OrcamentoVenda.objects.select_related(
                'cliente', 'vendedor'
            ).prefetch_related(
                'itens__produto'
            ).get(pk=orcamento_id, user=request.user)

            itens_list = []
            for item in orcamento.itens.all():
                itens_list.append({
                    'id': item.produto.id,
                    'nome': item.produto.nome,
                    'quantidade': float(item.quantidade),
                    'preco': float(item.preco_unitario)
                })

            orcamento_data = {
                'id': orcamento.id,
                'cliente_id': orcamento.cliente.id,
                'vendedor_id': orcamento.vendedor.id if orcamento.vendedor else None,
                'data_validade': orcamento.data_validade.strftime('%Y-%m-%d'),
                'observacoes': orcamento.observacoes or '',
                'valor_total': float(orcamento.valor_total),
                'itens': itens_list
            }
            return JsonResponse({'status': 'success', 'orcamento': orcamento_data})

        except OrcamentoVenda.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Orçamento não encontrado.'}, status=404)
        except Exception as e:
            logger.error(f"Erro ao buscar detalhes do orçamento {orcamento_id} via GET: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Erro interno: {str(e)}'}, status=500)
    # --- FIM: LÓGICA AJAX GET ---

    # --- LÓGICA POST (MODIFICADA para Atualizar/Criar) ---
    if request.method == 'POST':
        # (Lógica de action_delete e action_change_status continua igual)
        if 'action_delete' in request.POST:
            orcamento_id = request.POST.get('action_delete')
            orcamento = get_object_or_404(OrcamentoVenda, id=orcamento_id, user=request.user)
            orcamento.delete()
            messages.success(request, f'Orçamento #{orcamento.id} foi excluído com sucesso.')
            return redirect('orcamentos_venda')
        elif 'action_change_status' in request.POST:
            combined_value = request.POST.get('action_change_status')
            novo_status, orcamento_id = combined_value.split('-')
            orcamento = get_object_or_404(OrcamentoVenda, id=orcamento_id, user=request.user)
            orcamento.status = novo_status
            orcamento.save()
            messages.success(request, f'Status do Orçamento #{orcamento.id} alterado para "{orcamento.get_status_display()}".')
            return redirect('orcamentos_venda')
        
        # LÓGICA PARA SALVAR (CRIAR OU ATUALIZAR) VIA AJAX POST
        try:
            data = json.loads(request.body)
            orcamento_id = data.get('orcamento_id') # Pega o ID enviado pelo JS

            with transaction.atomic():
                cliente = get_object_or_404(Cliente, id=data.get('cliente_id'), user=request.user)
                vendedor = get_object_or_404(Vendedor, id=data.get('vendedor_id'), user=request.user) if data.get('vendedor_id') else None
                
                if orcamento_id: # Se tem ID, ATUALIZA
                    orcamento = get_object_or_404(OrcamentoVenda, id=orcamento_id, user=request.user)
                    orcamento.cliente = cliente
                    orcamento.vendedor = vendedor
                    orcamento.data_validade = data.get('data_validade')
                    orcamento.valor_total = Decimal(data.get('valor_total', '0'))
                    orcamento.observacoes = data.get('observacoes', '')
                    # Considerar voltar o status para PENDENTE ao editar?
                    # orcamento.status = 'PENDENTE' 
                    orcamento.save() 
                    orcamento.itens.all().delete() # Remove itens antigos
                    success_message = f'Orçamento #{orcamento.id} atualizado com sucesso!'
                else: # Se não tem ID, CRIA
                    orcamento = OrcamentoVenda.objects.create(
                        user=request.user,
                        cliente=cliente,
                        vendedor=vendedor,
                        data_validade=data.get('data_validade'),
                        valor_total=Decimal(data.get('valor_total', '0')),
                        observacoes=data.get('observacoes', '')
                        # Status PENDENTE é o default
                    )
                    success_message = f'Orçamento #{orcamento.id} criado com sucesso!'

                # (Re)Cria os itens
                for item_data in data.get('itens', []):
                    produto = get_object_or_404(ProdutoServico, id=item_data['id'], user=request.user)
                    ItemOrcamento.objects.create(
                        orcamento=orcamento,
                        produto=produto,
                        quantidade=Decimal(item_data['quantidade']),
                        preco_unitario=Decimal(item_data['preco'])
                    )
                
                messages.success(request, success_message) 
                return JsonResponse({'status': 'success', 'redirect_url': request.path}) # AJAX redireciona
        except Exception as e:
            logger.error(f"Erro ao salvar/atualizar orçamento via POST AJAX: {str(e)}") 
            return JsonResponse({'status': 'error', 'message': f'Erro interno: {str(e)}'}, status=400)

    # --- LÓGICA GET (Listagem e Filtros - Sem alterações) ---
    status_filter = request.GET.get('status', '')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    orcamentos_list = OrcamentoVenda.objects.filter(user=request.user).select_related('cliente')
    
    if status_filter:
        orcamentos_list = orcamentos_list.filter(status=status_filter)
    if start_date:
        orcamentos_list = orcamentos_list.filter(data_criacao__date__gte=start_date)
    if end_date:
        orcamentos_list = orcamentos_list.filter(data_criacao__date__lte=parse_date(end_date))

    # ▼▼▼ INÍCIO DA LÓGICA DE EXPORTAÇÃO (ORÇAMENTOS) ▼▼▼
    
    # 1. Exportação Excel
    if 'export_excel' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="orcamentos_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orçamentos"

        # Cabeçalhos
        headers = ['ID', 'Cliente', 'Vendedor', 'Data Criação', 'Validade', 'Itens', 'Observações', 'Valor Total', 'Status']
        ws.append(headers)

        # Estilo do cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Dados
        for orc in orcamentos_list:
            # Formata a lista de itens em uma string única
            itens_str = " | ".join([f"{item.quantidade}x {item.produto.nome}" for item in orc.itens.all()])
            
            # Formata datas
            data_criacao = orc.data_criacao.strftime('%d/%m/%Y') if orc.data_criacao else '-'
            data_validade = orc.data_validade.strftime('%d/%m/%Y') if orc.data_validade else '-'

            ws.append([
                orc.id,
                orc.cliente.nome,
                orc.vendedor.nome if orc.vendedor else '-',
                data_criacao,
                data_validade,
                itens_str,
                orc.observacoes or '',
                orc.valor_total,
                orc.get_status_display()
            ])

        # Formatação de Moeda na coluna "Valor Total" (Coluna H -> índice 8)
        currency_format = '"R$ "#,##0.00'
        for row in range(2, ws.max_row + 1):
            ws[f'H{row}'].number_format = currency_format

        # Ajuste de largura das colunas
        ws.column_dimensions['B'].width = 30 # Cliente
        ws.column_dimensions['F'].width = 50 # Itens
        ws.column_dimensions['G'].width = 30 # Observações
        
        wb.save(response)
        return response

    # 2. Exportação PDF
    if 'export_pdf' in request.GET:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        elements.append(Paragraph(f"Relatório de Orçamentos", styles['Title']))
        elements.append(Spacer(1, 12))

        # Cabeçalho da Tabela
        data = [['ID', 'Cliente', 'Data', 'Validade', 'Valor Total', 'Status']]

        # Dados da Tabela
        for orc in orcamentos_list:
            data_criacao = orc.data_criacao.strftime('%d/%m/%Y') if orc.data_criacao else '-'
            data_validade = orc.data_validade.strftime('%d/%m/%Y') if orc.data_validade else '-'
            valor_fmt = f"R$ {orc.valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            data.append([
                str(orc.id),
                orc.cliente.nome[:25], 
                data_criacao,
                data_validade,
                valor_fmt,
                orc.get_status_display()
            ])

        # Estilo da Tabela
        table = Table(data, colWidths=[1.5*cm, 8*cm, 3*cm, 3*cm, 4*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')
    # ▲▲▲ FIM DA LÓGICA DE EXPORTAÇÃO ▲▲▲
        
    paginator = Paginator(orcamentos_list, 10)
    page_number = request.GET.get('page')
    orcamentos_page_obj = paginator.get_page(page_number)
    
    # Dados para o formulário de criação
    clientes = Cliente.objects.filter(user=request.user).order_by('nome')
    vendedores = Vendedor.objects.filter(user=request.user).order_by('nome')
    produtos = ProdutoServico.objects.filter(user=request.user)
    produtos_json = json.dumps([{'id': p.id, 'nome': p.nome, 'codigo': p.codigo, 'preco': float(p.preco_venda)} for p in produtos])
    clientes_json = json.dumps([{'id': c.id, 'nome': c.nome, 'cpf_cnpj': c.cpf_cnpj or '', 'email': c.email or '', 'telefone': c.telefone or ''} for c in clientes])

    context = {
        'orcamentos': orcamentos_page_obj,
        'clientes': clientes,
        'vendedores': vendedores,
        'produtos_json': produtos_json,
        'clientes_json': clientes_json, 
        'status_choices': OrcamentoVenda.STATUS_CHOICES,
        'status_filter': status_filter,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'accounts/orcamentos_venda.html', context)

# --- NOVA VIEW PARA CONVERTER ORÇAMENTO EM VENDA ---
@login_required
@module_access_required('commercial')
@check_employee_permission('can_access_orcamentos_venda')
def converter_orcamento_venda_view(request, pk):
    orcamento = get_object_or_404(OrcamentoVenda, pk=pk, user=request.user)
    
    if orcamento.status != 'ACEITO':
        messages.error(request, "Apenas orçamentos com status 'Aceito' podem ser convertidos em venda.")
        return redirect('orcamentos_venda')
        
    try:
        with transaction.atomic():
            # 1. Cria a Venda
            nova_venda = Venda.objects.create(
                user=request.user,
                cliente=orcamento.cliente,
                vendedor=orcamento.vendedor,
                data_venda=timezone.now(),
                valor_total_liquido=orcamento.valor_total,
                status='EM_ANDAMENTO'
                # Outros campos como cidade/estado podem ser adicionados aqui se necessário
            )

            description_parts = []
            # 2. Cria os Itens da Venda e ajusta o estoque
            for item_orcamento in orcamento.itens.all():
                ItemVenda.objects.create(
                    venda=nova_venda,
                    produto=item_orcamento.produto,
                    quantidade=item_orcamento.quantidade,
                    preco_unitario=item_orcamento.preco_unitario
                )
                if item_orcamento.produto.tipo == 'PRODUTO':
                    item_orcamento.produto.estoque_atual -= item_orcamento.quantidade
                    item_orcamento.produto.save()
                description_parts.append(f"{item_orcamento.quantidade}x {item_orcamento.produto.nome}")

            # 3. Cria a Conta a Receber
            venda_category, _ = Category.objects.get_or_create(name="Venda via Orçamento")
            # --- Bloco Corrigido ---
            # Cria a Conta a Receber com uma descrição rastreável
            ReceivableAccount.objects.create(
                user=request.user,
                name=f"Venda do Orçamento #{orcamento.id}",
                # A LINHA MAIS IMPORTANTE:
                description=f"Ref. Venda #{nova_venda.id}: " + ", ".join(description_parts),
                due_date=orcamento.data_validade, # Usa a data de validade como vencimento
                amount=nova_venda.valor_total_liquido,
                category=venda_category,
                dre_area='BRUTA',
                is_received=False
            )
            # --- Fim do Bloco Corrigido ---
            
            # 4. Atualiza o status do orçamento para evitar dupla conversão (opcional)
            # orcamento.status = 'CONVERTIDO' # Você pode adicionar um novo status se quiser
            # orcamento.save()
            # --- ADICIONE AS DUAS LINHAS ABAIXO ---
            orcamento.status = 'CONVERTIDO' # Muda o status
            orcamento.save()                # Salva a alteração
            # --- FIM DAS LINHAS ADICIONADAS ---
            messages.success(request, f'Orçamento #{orcamento.id} convertido na Venda #{nova_venda.id} com sucesso!')
            return redirect('vendas')

    except Exception as e:
        messages.error(request, f'Ocorreu um erro ao converter o orçamento: {str(e)}')
        return redirect('orcamentos_venda')    
    


import logging

logger = logging.getLogger(__name__)


# Em accounts/views.py

# ▼▼▼ SUBSTITUA sua função existente por este bloco de código ▼▼▼
@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_contratos')
def gerenciamento_contratos_view(request):
    edit_id = request.GET.get('edit')
    instance = None
    if edit_id:
        instance = get_object_or_404(Contract, id=edit_id, user=request.user)

    # Inicializa o form aqui - usado no GET ou se o POST for inválido
    form = ContractForm(instance=instance, user=request.user)

    if request.method == 'POST':
        redirect_url = request.META.get('HTTP_REFERER', 'gerenciamento_contratos')

        # --- Tratamento das Ações (Excluir, Anexar) ---
        # NENHUMA ALTERAÇÃO necessária nestes blocos de ação
        if 'action_delete_contract' in request.POST:
            contract_id = request.POST.get('action_delete_contract')
            contract = get_object_or_404(Contract, id=contract_id, user=request.user)
            contract.delete()
            messages.success(request, f'Contrato "{contract.title}" excluído com sucesso.')
            return redirect(redirect_url)

        elif 'delete_selected_contracts' in request.POST:
            contract_ids = request.POST.getlist('contract_ids')
            if contract_ids:
                count, _ = Contract.objects.filter(id__in=contract_ids, user=request.user).delete()
                messages.success(request, f'{count} contrato(s) selecionado(s) foram excluídos.')
            return redirect(redirect_url)

        elif 'action_attach' in request.POST and 'file' in request.FILES:
            contract_id = request.POST.get('attach_contract_id')
            contract = get_object_or_404(Contract, id=contract_id, user=request.user)
            contract.document = request.FILES['file']
            contract.save()
            messages.success(request, 'Documento anexado ao contrato com sucesso.')
            return redirect(redirect_url)

        # --- Processamento do Formulário Principal (Criar/Editar) ---
        else:
            # Determina a instância para edição baseado nos dados do POST
            post_instance_id = request.POST.get('contract_id')
            if post_instance_id:
                instance = get_object_or_404(Contract, id=post_instance_id, user=request.user)
            else:
                instance = None # Garante que instance é None para criação

            # **SUBSTITUIR**: Instancia o form COM os dados do POST
            form = ContractForm(request.POST, request.FILES, instance=instance, user=request.user)

            if form.is_valid(): # Se o formulário FOR válido
                form.instance.user = request.user
                form.save()
                messages.success(request, 'Contrato salvo com sucesso.')
                return redirect('gerenciamento_contratos') # Redireciona após salvar com sucesso
            else: # Se o formulário FOR INVÁLIDO
                # **SUBSTITUIR**: Loga o erro e define a mensagem, MAS NÃO RODA a lógica GET aqui
                logger.error(f"Erro no formulário de Contrato: {form.errors.as_json()}")
                messages.error(request, f'Erro no formulário. Verifique os dados.')
                # A variável 'form' inválida será usada na renderização abaixo

    # --- Lógica GET (Roda na requisição GET inicial OU após um POST inválido) ---

    # **ADICIONAR**: Lê todos os filtros da requisição GET
    filter_status = request.GET.get('status', 'all')
    start_date_str = request.GET.get('start_date') # Usar sufixo _str
    end_date_str = request.GET.get('end_date')     # Usar sufixo _str
    search_term = request.GET.get('search_term', '') # <<< ADICIONAR termo de busca

    # REMOVER: Nenhuma filtragem de data padrão é necessária aqui

    # Query base
    contracts_query = Contract.objects.filter(user=request.user).select_related('client').order_by('-start_date')

    # **ADICIONAR**: Aplica o filtro de busca PRIMEIRO se search_term existir
    if search_term:
        contracts_query = contracts_query.filter(
            Q(title__icontains=search_term) |        # Busca no título OU
            Q(client__nome__icontains=search_term) # Busca no nome do cliente
        )

    # Aplica outros filtros SOMENTE se eles existirem na URL
    if filter_status != 'all' and filter_status:
        contracts_query = contracts_query.filter(status=filter_status)
    if start_date_str: # Verifica a versão _str
        try: # Adiciona try-except para parse robusto da data
             start_date_obj = parse_date(start_date_str)
             if start_date_obj:
                 contracts_query = contracts_query.filter(start_date__gte=start_date_obj)
        except ValueError:
             messages.warning(request, "Formato de data inicial inválido.")
    if end_date_str: # Verifica a versão _str
        try: # Adiciona try-except para parse robusto da data
             end_date_obj = parse_date(end_date_str)
             if end_date_obj:
                  contracts_query = contracts_query.filter(end_date__lte=end_date_obj) # Usa end_date para verificação <=
        except ValueError:
             messages.warning(request, "Formato de data final inválido.")

    # Calcula o total
    total_value = contracts_query.aggregate(total=Sum('value'))['total'] or Decimal('0.00')

    # Paginação
    paginator = Paginator(contracts_query, 10)
    page_number = request.GET.get('page')
    contracts_page_obj = paginator.get_page(page_number)

    # Prepara o contexto
    context = {
        'form': form, # Este é ou o form limpo do GET ou o form inválido do POST
        'contracts': contracts_page_obj,
        'filter_status': filter_status,
        'start_date': start_date_str if start_date_str else '', # Passa a string de volta para o template
        'end_date': end_date_str if end_date_str else '',       # Passa a string de volta para o template
        'search_term': search_term, # <<< ADICIONAR termo de busca ao contexto
        'total_value': total_value,
        'status_choices': Contract.STATUS_CHOICES,
    }
    return render(request, 'accounts/gerenciamento_contratos.html', context)


CLIENT_ID = os.environ.get('CONTA_AZUL_CLIENT_ID')
CLIENT_SECRET = os.environ.get('CONTA_AZUL_CLIENT_SECRET')
REDIRECT_URI = os.environ.get('CONTA_AZUL_REDIRECT_URI')
# Endereços da API Conta Azul (CORRIGIDOS)
AUTHORIZATION_URL = 'https://auth.contaazul.com/oauth2/authorize' # <-- MUDADO AQUI
TOKEN_URL = 'https://auth.contaazul.com/oauth2/token'         # <-- MUDADO AQUI

@login_required
def contaazul_auth_redirect(request):
    """Redireciona para o Conta Azul para autorização."""
    # Escopo corrigido (com espaços)
    scope = "openid profile aws.cognito.signin.user.admin" 
    
    oauth = OAuth2Session(CLIENT_ID, redirect_uri=REDIRECT_URI, scope=scope)
    
    # --- ALTERAÇÃO AQUI ---
    # Adicionamos kwargs={'prompt': 'login consent'} para forçar o login e a escolha
    authorization_url, state = oauth.authorization_url(
        AUTHORIZATION_URL,
        prompt='login consent' 
    )
    # ----------------------
    
    request.session['oauth_state'] = state
    return redirect(authorization_url)

@login_required
def contaazul_callback(request):
    """Recebe o callback do Conta Azul após autorização."""
    # Garante que CLIENT_ID e CLIENT_SECRET foram carregados
    if not CLIENT_ID or not CLIENT_SECRET:
        messages.error(request, "Credenciais do Conta Azul não configuradas no servidor.")
        return redirect('home')

    try:
        oauth = OAuth2Session(
            CLIENT_ID,
            redirect_uri=REDIRECT_URI, # Usará a URL do Render aqui
            state=request.session.get('oauth_state')
        )
        # Monta a URL completa de resposta (incluindo https)
        # Use request.scheme para obter https
        full_response_url = request.scheme + '://' + request.get_host() + request.get_full_path()

        token = oauth.fetch_token(
            TOKEN_URL,
            # Usa a URL completa na resposta
            authorization_response=full_response_url,
            # Inclui client_id e client_secret no corpo para alguns fluxos
            # client_secret=CLIENT_SECRET # Descomente se fetch_token reclamar
            # Pode precisar passar explicitamente:
             auth=(CLIENT_ID, CLIENT_SECRET) # Autenticação Basic para o token endpoint
        )

        expires_in = token.get('expires_in', 3600) # Default 1 hora
        expires_at_time = timezone.now() + timedelta(seconds=expires_in)

        creds, created = ContaAzulCredentials.objects.update_or_create(
            user=request.user,
            defaults={
                'access_token': token['access_token'],
                'refresh_token': token.get('refresh_token'), # refresh_token pode não vir sempre
                'expires_at': expires_at_time
            }
        )
        messages.success(request, 'Conta Azul conectada com sucesso!')
        return redirect('home')

    except Exception as e:
        # Log detalhado do erro no servidor para depuração
        import traceback
        print(f"Erro no callback Conta Azul: {e}")
        print(traceback.format_exc())
        messages.error(request, f'Erro ao conectar com Conta Azul. Verifique as configurações e tente novamente. Detalhe: {e}')
        return redirect('home')
    


@login_required
@owner_required
@subscription_required  # Garante que o dono tenha uma assinatura ativa
def manage_users_view(request):
    owner = request.user
    try:
        # Usa request.active_subscription (do decorador)
        subscription = request.active_subscription
        employee_limit = subscription.employee_limit
    except (Subscription.DoesNotExist, AttributeError):
        messages.error(request, "Assinatura principal não encontrada.")
        return redirect('home')

    current_employees_links = CompanyUserLink.objects.filter(owner=owner)
    current_employee_count = current_employees_links.count()
    can_add_more = current_employee_count < employee_limit

    password_form = PasswordChangeForm(owner)
    employee_form = EmployeeCreationForm()

    if request.method == 'POST':
        if 'change_password' in request.POST:
            password_form = PasswordChangeForm(owner, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Sua senha foi alterada com sucesso!')
                return redirect('manage_users')
            else:
                messages.error(request, 'Erro ao alterar a senha. Verifique os campos.')

        elif 'add_employee' in request.POST:
            if not can_add_more:
                messages.error(request, f"Limite de {employee_limit} funcionários adicionais atingido.")
                return redirect('manage_users')

            employee_form = EmployeeCreationForm(request.POST)
            if employee_form.is_valid():
                try:
                    new_employee = User.objects.create_user(
                        username=employee_form.cleaned_data['email'],
                        email=employee_form.cleaned_data['email'],
                        password=employee_form.cleaned_data['password'],
                        first_name=employee_form.cleaned_data['first_name']
                    )
                    # Pega os dados de permissão do formulário
                    CompanyUserLink.objects.create(
                        owner=owner, 
                        employee=new_employee,
                        can_access_home = employee_form.cleaned_data.get('can_access_home', True),

                        # Permissões Módulo Financeiro
                        can_access_contas_pagar = employee_form.cleaned_data.get('can_access_contas_pagar', False),
                        can_access_contas_receber = employee_form.cleaned_data.get('can_access_contas_receber', False),
                        can_access_tarefas = employee_form.cleaned_data.get('can_access_tarefas', False),
                        can_access_orcamento_anual = employee_form.cleaned_data.get('can_access_orcamento_anual', False),
                        can_access_painel_financeiro = employee_form.cleaned_data.get('can_access_painel_financeiro', False),
                        can_access_fornecedores = employee_form.cleaned_data.get('can_access_fornecedores', False),
                        can_access_clientes_financeiro = employee_form.cleaned_data.get('can_access_clientes_financeiro', False),

                        # Permissões Módulo Comercial
                        can_access_painel_vendas = employee_form.cleaned_data.get('can_access_painel_vendas', False),
                        can_access_notas_fiscais = employee_form.cleaned_data.get('can_access_notas_fiscais', False),
                        can_access_orcamentos_venda = employee_form.cleaned_data.get('can_access_orcamentos_venda', False),
                        can_access_contratos = employee_form.cleaned_data.get('can_access_contratos', False),
                        can_access_cadastros_comercial = employee_form.cleaned_data.get('can_access_cadastros_comercial', False),
                        can_access_vendas = employee_form.cleaned_data.get('can_access_vendas', False),
                        can_access_metas_comerciais = employee_form.cleaned_data.get('can_access_metas_comerciais', False),
                        can_access_precificacao = employee_form.cleaned_data.get('can_access_precificacao', False),
                        can_access_pdv = employee_form.cleaned_data.get('can_access_pdv', False)
                    )
                    messages.success(request, f"Funcionário '{new_employee.first_name}' adicionado com sucesso!")
                    # ▲▲▲ FIM DA SUBSTITUIÇÃO ▲▲▲
                except Exception as e:
                    messages.error(request, f"Erro ao criar funcionário: {e}")

                return redirect('manage_users')
            else:
                messages.error(request, 'Erro ao adicionar funcionário. Verifique os campos.')

        elif 'delete_employee' in request.POST:
            employee_link_id = request.POST.get('employee_link_id')
            link_to_delete = get_object_or_404(CompanyUserLink, id=employee_link_id, owner=owner)
            employee_user = link_to_delete.employee
            try:
                link_to_delete.delete()
                employee_user.delete()
                messages.success(request, f"Funcionário '{employee_user.first_name}' excluído com sucesso.")
            except Exception as e:
                messages.error(request, f"Erro ao excluir funcionário: {e}")
            return redirect('manage_users')

    context = {
        'password_form': password_form,
        'employee_form': employee_form,
        'current_employees': current_employees_links.select_related('employee'),
        'employee_limit': employee_limit,
        'current_employee_count': current_employee_count,
        'can_add_more': can_add_more,
    }
    return render(request, 'accounts/manage_users.html', context)


@login_required
def smart_redirect_view(request):
    """
    Redireciona o usuário para a página correta após o login,
    usando 'request.company_link' do middleware.
    """
    link = getattr(request, 'company_link', None)

    # Se for o dono da conta (link é None) ou superuser, vai para a home.
    if link is None or request.user.is_superuser:
        return redirect('home')

    # --- ORDEM DE PRIORIDADE DE REDIRECIONAMENTO (para Funcionário) ---

    if link.can_access_home:
        return redirect('home')

    # Prioridade Financeira
    if link.can_access_painel_financeiro:
        return redirect('dashboards')
    if link.can_access_contas_pagar:
        return redirect('contas_pagar')
    if link.can_access_contas_receber:
        return redirect('contas_receber')

    # Prioridade Comercial
    if link.can_access_painel_vendas:
        return redirect('faturamento_dashboard')
    if link.can_access_vendas:
        return redirect('vendas')

    # Fallback (copie todas as suas permissões)
    if link.can_access_tarefas:
        return redirect('tarefas_home')
    if link.can_access_orcamento_anual:
        return redirect('orcamento_anual')
    if link.can_access_fornecedores:
        return redirect('fornecedores')
    if link.can_access_clientes_financeiro:
        return redirect('clientes')
    if link.can_access_notas_fiscais:
        return redirect('lista_notas_fiscais')
    if link.can_access_orcamentos_venda:
        return redirect('orcamentos_venda')
    if link.can_access_contratos:
        return redirect('gerenciamento_contratos')
    if link.can_access_cadastros_comercial:
        return redirect('comercial_cadastros')
    if link.can_access_metas_comerciais:
        return redirect('metas_comerciais')
    if link.can_access_precificacao:
        return redirect('precificacao')

    # Se o funcionário não tiver NENHUMA permissão
    messages.warning(request, 'Sua conta não possui permissão para acessar nenhuma página. Fale com o administrador.')
    return redirect('logout')



@login_required
@subscription_required
@module_access_required('commercial')
@check_employee_permission('can_access_cadastros_comercial')
def editar_cliente_view(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk, user=request.user)
    
    if request.method == 'POST':
        # Passamos o user no __init__ para a validação de CPF funcionar
        form = ClienteForm(request.POST, instance=cliente, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente atualizado com sucesso!')
            return redirect('comercial_cadastros')
    else:
        form = ClienteForm(instance=cliente, user=request.user)

    context = {
        'form': form,
        'item_name': cliente.nome,
        'is_cliente_edit': True # Flag opcional para o template saber
    }
    # Vamos reutilizar o template genérico de edição
    return render(request, 'accounts/editar_item.html', context)

# MANTENHA AS IMPORTAÇÕES EXISTENTES E ADICIONE AS NOVAS NO TOPO:
from .utils_exports import (
    gerar_pdf_generic, gerar_excel_generic, 
    gerar_excel_orcamento, gerar_pdf_orcamento,
    gerar_excel_dre, gerar_pdf_dre,                 # <--- Novo
    gerar_excel_fluxo_caixa, gerar_pdf_fluxo_caixa  # <--- Novo
)
from collections import defaultdict, OrderedDict
from django.db.models.functions import TruncMonth

@login_required
@subscription_required
def relatorios_view(request):
    if request.method == 'POST' or request.GET.get('gerar'):
        tipo = request.GET.get('tipo') or request.POST.get('tipo')
        formato = request.GET.get('formato') or request.POST.get('formato')
        start_date_str = request.GET.get('start_date') or request.POST.get('start_date')
        end_date_str = request.GET.get('end_date') or request.POST.get('end_date')
        
        # Define datas padrão se não vierem
        if not start_date_str or not end_date_str:
            hj = datetime.now().date()
            start_date = hj.replace(day=1)
            end_date = hj
        else:
            start_date = parse_date(start_date_str)
            end_date = parse_date(end_date_str)

        periodo_label = f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"

        # --- DRE GERENCIAL ---
        if tipo == 'dre':
            # Copia a lógica de cálculo da DRE (Regime de Caixa do Dashboard)
            receita_bruta = ReceivableAccount.objects.filter(user=request.user, is_received=True, dre_area='BRUTA', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
            impostos = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DEDUCAO', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
            custos = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='CUSTOS', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
            despesas_operacionais = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='OPERACIONAL', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
            depreciacao = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DEPRECIACAO', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
            nao_operacionais = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='NAO_OPERACIONAL', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
            tributacao = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='TRIBUTACAO', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
            distribuicao_lucro = PayableAccount.objects.filter(user=request.user, is_paid=True, dre_area='DISTRIBUICAO', due_date__range=[start_date, end_date]).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

            # Cálculos derivados
            receita_liquida = receita_bruta - impostos
            lucro_bruto = receita_liquida - custos
            ebitda = lucro_bruto - despesas_operacionais
            ebit = ebitda - depreciacao
            lair = ebit - nao_operacionais
            lucro_liquido = lair - tributacao
            resultado_final = lucro_liquido - distribuicao_lucro

            dados_dre = {
                'receita_bruta': receita_bruta, 'impostos': impostos, 'receita_liquida': receita_liquida,
                'custos': custos, 'lucro_bruto': lucro_bruto, 'despesas_operacionais': despesas_operacionais,
                'ebitda': ebitda, 'depreciacao': depreciacao, 'ebit': ebit, 'nao_operacionais': nao_operacionais,
                'lair': lair, 'tributacao': tributacao, 'lucro_liquido': lucro_liquido,
                'distribuicao_lucro': distribuicao_lucro, 'resultado_final': resultado_final
            }

            if formato == 'pdf': return gerar_pdf_dre(dados_dre, periodo_label)
            elif formato == 'excel': return gerar_excel_dre(dados_dre, periodo_label)

        # --- FLUXO DE CAIXA ---
        elif tipo == 'fluxo_caixa':
            # Agrupa por mês dentro do período selecionado
            entradas_qs = ReceivableAccount.objects.filter(
                user=request.user, is_received=True, due_date__range=[start_date, end_date]
            ).exclude(dre_area='NAO_CONSTAR').annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount')).order_by('month')

            saidas_qs = PayableAccount.objects.filter(
                user=request.user, is_paid=True, due_date__range=[start_date, end_date]
            ).exclude(dre_area='NAO_CONSTAR').annotate(month=TruncMonth('due_date')).values('month').annotate(total=Sum('amount')).order_by('month')

            # Organiza os dados em dicionários por mês
            entradas_dict = {e['month'].strftime('%b/%Y'): float(e['total']) for e in entradas_qs}
            saidas_dict = {s['month'].strftime('%b/%Y'): float(s['total']) for s in saidas_qs}

            # Cria lista unificada de todos os meses que têm dados
            todos_meses = sorted(list(set(list(entradas_dict.keys()) + list(saidas_dict.keys()))), 
                                 key=lambda x: datetime.strptime(x, '%b/%Y'))

            labels = []
            lista_entradas = []
            lista_saidas = []
            lista_saldos = []

            # Se não houver dados, cria pelo menos uma entrada vazia
            if not todos_meses:
                labels = ["Sem dados"]
                lista_entradas = [0]
                lista_saidas = [0]
                lista_saldos = [0]
            else:
                for mes in todos_meses:
                    ent = entradas_dict.get(mes, 0)
                    sai = saidas_dict.get(mes, 0)
                    labels.append(mes)
                    lista_entradas.append(ent)
                    lista_saidas.append(sai)
                    lista_saldos.append(ent - sai)

            dados_fc = {
                'labels': labels,
                'entradas': lista_entradas,
                'saidas': lista_saidas,
                'geracao_caixa': lista_saldos
            }

            if formato == 'pdf': return gerar_pdf_fluxo_caixa(dados_fc, periodo_label)
            elif formato == 'excel': return gerar_excel_fluxo_caixa(dados_fc, periodo_label)

        # --- ORÇAMENTO ANUAL ---
        elif tipo == 'orcamento':
            # (Mantém a lógica que já criamos para Orçamento...)
            try:
                ano = start_date.year
            except:
                ano = datetime.now().year
            
            # ... (Copie/Mantenha a lógica de coleta de dados do orçamento aqui) ...
            # Para economizar espaço na resposta, estou assumindo que o código anterior 
            # do orçamento permanece aqui.
            
            # 1. Busca dados do Banco
            orcamentos = Orcamento.objects.filter(user=request.user, mes_ano__year=ano)
            realizado_pagar = PayableAccount.objects.filter(user=request.user, is_paid=True, due_date__year=ano).annotate(month=TruncMonth('due_date')).values('category__name', 'month').annotate(total=Sum('amount'))
            realizado_receber = ReceivableAccount.objects.filter(user=request.user, is_received=True, due_date__year=ano).annotate(month=TruncMonth('due_date')).values('category__name', 'month').annotate(total=Sum('amount'))

            revenue_data = defaultdict(lambda: {'orcado_mensal': [Decimal(0)]*12, 'realizado_mensal': [Decimal(0)]*12})
            for o in orcamentos:
                if o.category and o.category.category_type == 'RECEIVABLE': revenue_data[o.category.name]['orcado_mensal'][o.mes_ano.month - 1] += o.valor_orcado
            for r in realizado_receber:
                if r['category__name']: revenue_data[r['category__name']]['realizado_mensal'][r['month'].month - 1] += r['total']

            expense_data = defaultdict(lambda: {'orcado_mensal': [Decimal(0)]*12, 'realizado_mensal': [Decimal(0)]*12})
            for o in orcamentos:
                if o.category and o.category.category_type == 'PAYABLE': expense_data[o.category.name]['orcado_mensal'][o.mes_ano.month - 1] += o.valor_orcado
            for p in realizado_pagar:
                if p['category__name']: expense_data[p['category__name']]['realizado_mensal'][p['month'].month - 1] += p['total']

            for d in revenue_data.values(): d['total_orcado_ano'] = sum(d['orcado_mensal']); d['total_realizado_ano'] = sum(d['realizado_mensal'])
            for d in expense_data.values(): d['total_orcado_ano'] = sum(d['orcado_mensal']); d['total_realizado_ano'] = sum(d['realizado_mensal'])

            dados_consolidados = {'receitas': dict(revenue_data), 'despesas': dict(expense_data)}

            if formato == 'pdf': return gerar_pdf_orcamento(dados_consolidados, ano)
            elif formato == 'excel': return gerar_excel_orcamento(dados_consolidados, ano)


        # --- LÓGICA PADRÃO (Contas a Pagar/Receber) ---
        status = request.GET.get('status') or request.POST.get('status')
        bank_id = request.GET.get('bank') or request.POST.get('bank')

        queryset = None
        if tipo == 'pagar':
            queryset = PayableAccount.objects.filter(user=request.user).order_by('due_date')
            if status == 'aberto': queryset = queryset.filter(is_paid=False)
            elif status == 'pago': queryset = queryset.filter(is_paid=True)
        else: # receber
            queryset = ReceivableAccount.objects.filter(user=request.user).order_by('due_date')
            if status == 'aberto': queryset = queryset.filter(is_received=False)
            elif status == 'recebido': queryset = queryset.filter(is_received=True)

        if start_date: queryset = queryset.filter(due_date__gte=start_date)
        if end_date: queryset = queryset.filter(due_date__lte=end_date)
        if bank_id: queryset = queryset.filter(bank_account_id=bank_id)

        if formato == 'pdf': return gerar_pdf_generic(queryset, tipo)
        elif formato == 'excel': return gerar_excel_generic(queryset, tipo)

    user_banks = BankAccount.objects.filter(user=request.user)
    return render(request, 'accounts/relatorios.html', {'user_banks': user_banks})



@login_required
@subscription_required
def configurar_inter_view(request):
    # Tenta pegar a instância existente ou cria uma nova
    try:
        instance = request.user.inter_creds
    except InterCredentials.DoesNotExist:
        instance = None

    if request.method == 'POST':
        form = InterCredentialsForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            creds = form.save(commit=False)
            creds.user = request.user
            creds.save()
            messages.success(request, 'Credenciais do Banco Inter salvas com sucesso!')
            return redirect('importar_ofx')
        else:
            messages.error(request, 'Erro ao salvar. Verifique os arquivos.')
    else:
        form = InterCredentialsForm(instance=instance)

    return render(request, 'accounts/configurar_inter.html', {'form': form})




# accounts/views.py

# ... imports existentes ...


@login_required
@subscription_required
def configurar_mercadopago_view(request):
    # Tenta pegar a instância existente ou define como None para criar
    try:
        instance = request.user.mercadopago_creds
    except MercadoPagoCredentials.DoesNotExist:
        instance = None

    if request.method == 'POST':
        form = MercadoPagoCredentialsForm(request.POST, instance=instance)
        if form.is_valid():
            creds = form.save(commit=False)
            creds.user = request.user
            creds.save()
            messages.success(request, 'Credenciais do Mercado Pago salvas com sucesso!')
            # Redireciona para a tela de importação, igual ao fluxo do Banco Inter
            return redirect('importar_ofx')
        else:
            messages.error(request, 'Erro ao salvar. Verifique os dados inseridos.')
    else:
        form = MercadoPagoCredentialsForm(instance=instance)

    return render(request, 'accounts/configurar_mercadopago.html', {'form': form})




@login_required
@subscription_required
def configurar_asaas_view(request):
    try:
        instance = request.user.asaas_creds
    except AsaasCredentials.DoesNotExist:
        instance = None

    if request.method == 'POST':
        form = AsaasCredentialsForm(request.POST, instance=instance)
        if form.is_valid():
            creds = form.save(commit=False)
            creds.user = request.user
            creds.save()
            messages.success(request, 'Credenciais do Asaas salvas com sucesso!')
            return redirect('importar_ofx')
        else:
            messages.error(request, 'Erro ao salvar. Verifique a chave inserida.')
    else:
        form = AsaasCredentialsForm(instance=instance)

    return render(request, 'accounts/configurar_asaas.html', {'form': form})


@login_required
def create_checkout_session(request, plan_type):
    """
    Cria uma sessão de Checkout no Stripe para o plano escolhido
    plan_type: 'financeiro' ou 'completo'
    """
    stripe.api_key = settings.STRIPE_SECRET_KEY

    # 1. Identifica qual ID de preço usar
    price_id = settings.STRIPE_PRICE_IDS.get(plan_type)
    
    if not price_id:
        messages.error(request, "Plano inválido selecionado.")
        return redirect('assinatura')

    try:
        # 2. Busca ou Cria o Cliente no Stripe
        # Primeiro checamos se o usuário já tem um ID stripe salvo
        subscription = getattr(request.user, 'subscription', None)
        stripe_customer_id = subscription.stripe_customer_id if subscription else None

        if not stripe_customer_id:
            # Cria um novo cliente no Stripe com o email do usuário
            customer = stripe.Customer.create(
                email=request.user.email,
                name=request.user.get_full_name() or request.user.username,
                metadata={'user_id': request.user.id} # Importante para o Webhook achar o usuário depois
            )
            stripe_customer_id = customer.id
            
            # Salva o ID no banco para não criar duplicado depois
            if subscription:
                subscription.stripe_customer_id = stripe_customer_id
                subscription.save()

        # 3. Cria a Sessão de Checkout
        checkout_session = stripe.checkout.Session.create(
            customer=stripe_customer_id,
            payment_method_types=['card'], # Aceitar cartão (pode adicionar 'boleto' se ativar no painel)
            line_items=[
                {
                    'price': price_id,
                    'quantity': 1,
                },
            ],
            mode='subscription', # Modo assinatura (recorrente)
            # --- ADICIONE ESTE BLOCO ---
            subscription_data={
                'trial_period_days': 7, # O Stripe espera 7 dias para cobrar
            },
            # ---------------------------
            # URLs para onde o cliente volta após pagar ou cancelar
            success_url=settings.DOMAIN_URL + '/assinatura/sucesso/',
            cancel_url=settings.DOMAIN_URL + '/assinatura/cancelado/',
            metadata={
                'user_id': request.user.id,
                'plan_type': plan_type  # 'financeiro' ou 'completo'
            }
        )

        # 4. Redireciona o usuário para o link seguro do Stripe
        return redirect(checkout_session.url, code=303)

    except Exception as e:
        messages.error(request, f"Erro ao conectar com Stripe: {str(e)}")
        return redirect('assinatura')

# Views simples de retorno (só para não dar erro 404)
@login_required
def assinatura_sucesso(request):
    messages.success(request, "Pagamento recebido! Sua assinatura será ativada em instantes.")
    return redirect('home')

@login_required
def assinatura_cancelado(request):
    messages.warning(request, "A assinatura não foi completada. Você não foi cobrado.")
    return redirect('assinatura')



# No final de accounts/views.py

from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
import stripe
# Certifique-se que 'timedelta' e 'timezone' estão importados lá no topo,
# mas por segurança, vou usar imports diretos aqui dentro para evitar conflitos:
from django.utils import timezone
from datetime import timedelta 

@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    event = None

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, settings.STRIPE_WEBHOOK_SECRET
        )
    except ValueError as e:
        # Payload inválido
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        # Assinatura inválida
        return HttpResponse(status=400)

    # --- LÓGICA DO EVENTO ---
    if event['type'] == 'checkout.session.completed':
        session = event['data']['object']
        
        # Recupera metadata
        user_id = session.get('metadata', {}).get('user_id')
        plan_type = session.get('metadata', {}).get('plan_type')
        stripe_subscription_id = session.get('subscription')

        print(f"🔔 WEBHOOK RECEBIDO: User ID: {user_id}, Plano: {plan_type}")

        if user_id and plan_type:
            try:
                user = User.objects.get(id=user_id)
                subscription = Subscription.objects.get(user=user)

                # Atualiza dados básicos
                subscription.status = 'active'
                subscription.stripe_subscription_id = stripe_subscription_id
                
                # CORREÇÃO AQUI: Usando timedelta direto
                subscription.valid_until = timezone.now().date() + timedelta(days=30)

                # Lógica de Permissão
                if plan_type == 'financeiro':
                    subscription.has_financial_module = True
                    subscription.has_commercial_module = False
                    subscription.employee_limit = 2
                
                elif plan_type == 'completo':
                    subscription.has_financial_module = True
                    subscription.has_commercial_module = True
                    subscription.employee_limit = 5
                
                subscription.save()
                print(f"✅ SUCESSO TOTAL: Assinatura de {user.username} atualizada para {plan_type}!")

            except User.DoesNotExist:
                print("❌ ERRO: Usuário não encontrado no banco.")
            except Subscription.DoesNotExist:
                print("❌ ERRO: Assinatura (Subscription) não existe para este usuário.")
            except Exception as e:
                # Esse print vai nos salvar se houver outro erro!
                print(f"❌ ERRO CRÍTICO NO CÓDIGO: {str(e)}")
                return HttpResponse(status=500)

    return HttpResponse(status=200)




# accounts/views.py

from .forms import BPOAddClientForm # Importe o formulário novo

@login_required
def bpo_add_client_view(request):
    # 1. Verifica se é BPO
    try:
        sub = request.user.subscription
        if sub.user_type != 'BPO':
            messages.error(request, "Acesso negado. Apenas contas BPO podem adicionar clientes.")
            return redirect('home')
    except Subscription.DoesNotExist:
        return redirect('home')

    # 2. Verifica o Limite (AQUI ESTÁ A MÁGICA)
    current_clients_count = BPOClientLink.objects.filter(bpo_admin=request.user).count()
    limit = sub.client_limit

    if current_clients_count >= limit:
        messages.error(request, f"Você atingiu seu limite de {limit} clientes. Faça um upgrade para adicionar mais.")
        return redirect('bpo_dashboard')

    if request.method == 'POST':
        form = BPOAddClientForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                # A. Cria o Usuário do Cliente
                new_client_user = User.objects.create_user(
                    username=form.cleaned_data['email'],
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password'],
                    first_name=form.cleaned_data['first_name']
                )

                # B. Atualiza a Assinatura do Cliente (que já foi criada pelo signal post_save)
                # Usamos get_or_create por segurança, caso o signal falhe ou seja removido
                client_sub, created = Subscription.objects.get_or_create(user=new_client_user)
                
                client_sub.status = 'active'
                client_sub.user_type = 'CLIENTE'
                client_sub.valid_until = datetime.today().date() + timedelta(days=3650) # Validade longa
                client_sub.has_financial_module = True
                client_sub.has_commercial_module = sub.has_commercial_module # Herda do BPO
                client_sub.save()

                # C. Cria o Vínculo (Link)
                BPOClientLink.objects.create(
                    bpo_admin=request.user,
                    client=new_client_user
                )

                messages.success(request, f"Cliente '{new_client_user.first_name}' criado e vinculado com sucesso!")
                return redirect('bpo_dashboard')
    else:
        form = BPOAddClientForm()

    return render(request, 'accounts/bpo_add_client.html', {'form': form, 'limit': limit, 'current': current_clients_count})




@login_required
@subscription_required
def configurar_omie_view(request):
    # Tenta pegar a instância existente ou cria uma nova
    try:
        instance = request.user.omie_creds
    except OmieCredentials.DoesNotExist:
        instance = None

    if request.method == 'POST':
        form = OmieCredentialsForm(request.POST, instance=instance)
        if form.is_valid():
            creds = form.save(commit=False)
            creds.user = request.user
            creds.save()
            messages.success(request, 'Credenciais Omie salvas com sucesso!')
            # Redireciona para a tela de importação, mantendo o fluxo das outras integrações
            return redirect('importar_ofx')
        else:
            messages.error(request, 'Erro ao salvar. Verifique as chaves inseridas.')
    else:
        form = OmieCredentialsForm(instance=instance)

    return render(request, 'accounts/configurar_omie.html', {'form': form})



@login_required
@subscription_required
def configurar_nibo_view(request):
    try:
        instance = request.user.nibo_creds
    except NiboCredentials.DoesNotExist:
        instance = None

    if request.method == 'POST':
        form = NiboCredentialsForm(request.POST, instance=instance)
        if form.is_valid():
            creds = form.save(commit=False)
            creds.user = request.user
            creds.save()
            messages.success(request, 'Credenciais Nibo salvas com sucesso!')
            return redirect('importar_ofx')
        else:
            messages.error(request, 'Erro ao salvar. Verifique o token inserido.')
    else:
        form = NiboCredentialsForm(instance=instance)

    return render(request, 'accounts/configurar_nibo.html', {'form': form})

@login_required
@subscription_required
def configurar_tiny_view(request):
    try:
        instance = request.user.tiny_creds
    except TinyCredentials.DoesNotExist:
        instance = None

    if request.method == 'POST':
        form = TinyCredentialsForm(request.POST, instance=instance)
        if form.is_valid():
            creds = form.save(commit=False)
            creds.user = request.user
            creds.save()
            messages.success(request, 'Credenciais do Tiny salvas com sucesso!')
            return redirect('importar_ofx')
        else:
            messages.error(request, 'Erro ao salvar. Verifique o token inserido.')
    else:
        form = TinyCredentialsForm(instance=instance)

    return render(request, 'accounts/configurar_tiny.html', {'form': form})

def landing_page(request):
    if request.user.is_authenticated:
        return redirect('smart_redirect') # Se já logado, vai pro sistema
    return render(request, 'landing_page.html')




# 1. Certifique-se que este import está lá no TOPO do views.py, junto com os outros imports do Django
from django.views.decorators.http import require_POST

@login_required
@require_POST
def cancelar_assinatura_manual(request):
    try:
        subscription = Subscription.objects.get(user=request.user)
        
        if subscription.stripe_subscription_id:
            stripe.api_key = settings.STRIPE_SECRET_KEY
            
            # 1. Avisa ao Stripe: "Não cobre mais, mas deixe rodar até o fim"
            stripe.Subscription.modify(
                subscription.stripe_subscription_id,
                cancel_at_period_end=True
            )
            
            # --- CORREÇÃO AQUI ---
            # NÃO mudamos o status para 'canceled' agora.
            # O usuário pagou, ele tem direito de usar até o fim.
            # subscription.status = 'canceled'  <-- REMOVA OU COMENTE ESSA LINHA
            
            # Apenas salvamos se tivermos alterado algum outro campo (opcional aqui)
            # subscription.save() 
            
            messages.success(request, f"Renovação automática cancelada. Seu acesso continua garantido até {subscription.valid_until.strftime('%d/%m/%Y')}.")
        else:
            # Se for assinatura manual (sem Stripe), aí sim cancelamos na hora ou definimos a lógica que você preferir
            subscription.status = 'canceled'
            subscription.save()
            messages.warning(request, "Assinatura manual cancelada.")
            
    except Exception as e:
        messages.error(request, f"Erro ao cancelar: {str(e)}")
        
    return redirect('assinatura') 



