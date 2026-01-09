

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from accounts.models import PayableAccount, ReceivableAccount
import calendar # <--- Adicione isso junto com os outros imports
# O login_required vem direto do Django
from django.contrib.auth.decorators import login_required 

# Os outros vêm do seu arquivo de decorators customizados
from accounts.decorators import subscription_required, check_employee_permission, module_access_required

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.template.loader import get_template # Certifique-se de ter este import também
# ... mantenha seus outros imports (login_required, etc) ...

# --- FUNÇÃO AUXILIAR (Extrai a lógica para ser usada na Tela, Excel e PDF) ---
def _processar_dados_fluxo(request, ano_atual, mes_atual, view_mode, status_view):
    # 1. Definir Colunas
    if view_mode == 'diario':
        _, num_dias = calendar.monthrange(ano_atual, mes_atual)
        colunas_ids = range(1, num_dias + 1)
        colunas_labels = [str(d) for d in colunas_ids]
    else:
        colunas_ids = range(1, 13)
        colunas_labels = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

    # 2. Função interna de processamento
    def processar_modelo(modelo, is_field):
        campo_data = 'payment_date' if status_view == 'realizado' else 'due_date'
        
        filtro_base = {
            'user': request.user,
            f'{campo_data}__year': ano_atual
        }
        
        if status_view == 'realizado':
            filtro_base[f'{is_field}'] = True

        if view_mode == 'diario':
            filtro_base[f'{campo_data}__month'] = mes_atual

        nomes = modelo.objects.filter(**filtro_base).values_list('name', flat=True).distinct().order_by('name')
        
        dados = []
        totais_coluna = {c: 0 for c in colunas_ids}
        total_geral_ano = 0

        for nome in nomes:
            valores_periodo = []
            total_linha = 0
            
            for col in colunas_ids:
                filtro_celula = filtro_base.copy()
                filtro_celula['name'] = nome
                
                if view_mode == 'diario':
                    filtro_celula[f'{campo_data}__day'] = col
                else:
                    filtro_celula[f'{campo_data}__month'] = col

                valor = modelo.objects.filter(**filtro_celula).aggregate(Sum('amount'))['amount__sum'] or 0
                
                valores_periodo.append(valor)
                total_linha += valor
                totais_coluna[col] += valor
            
            dados.append({'nome': nome, 'valores': valores_periodo, 'total': total_linha})
            total_geral_ano += total_linha
            
        return dados, list(totais_coluna.values()), total_geral_ano

    # 3. Executa
    dados_receita, totais_receita_col, total_receitas_ano = processar_modelo(ReceivableAccount, 'is_received')
    dados_despesa, totais_despesa_col, total_despesas_ano = processar_modelo(PayableAccount, 'is_paid')

    # 4. Saldo
    saldo_mensal = []
    saldo_acumulado = 0
    for i, _ in enumerate(colunas_ids):
        saldo = totais_receita_col[i] - totais_despesa_col[i]
        saldo_mensal.append(saldo)
        saldo_acumulado += saldo

    return {
        'colunas_labels': colunas_labels,
        'dados_receita': dados_receita,
        'totais_receita_col': totais_receita_col,
        'total_receitas_ano': total_receitas_ano,
        'dados_despesa': dados_despesa,
        'totais_despesa_col': totais_despesa_col,
        'total_despesas_ano': total_despesas_ano,
        'saldo_mensal': saldo_mensal,
        'saldo_acumulado_ano': saldo_acumulado,
    }

# --- VIEW PRINCIPAL (TELA) ---
@login_required
@subscription_required
@module_access_required('financial')
@check_employee_permission('can_access_fluxo_caixa')
def fluxo_caixa_analitico(request):
    ano_get = request.GET.get('ano', str(timezone.now().year))
    try:
        ano_atual = int(str(ano_get).replace('.', ''))
    except ValueError:
        ano_atual = timezone.now().year

    view_mode = request.GET.get('view', 'mensal')
    status_view = request.GET.get('status_view', 'realizado')
    mes_atual = int(request.GET.get('mes', timezone.now().month))

    # Usa a função auxiliar para pegar os dados
    contexto_dados = _processar_dados_fluxo(request, ano_atual, mes_atual, view_mode, status_view)

    context = {
        'ano': ano_atual,
        'mes': mes_atual,
        'view_mode': view_mode,
        'status_view': status_view,
        **contexto_dados # Desempacota os dados retornados
    }
    
    return render(request, 'relatorios/fluxo_analitico.html', context)

# --- VIEW EXPORTAR EXCEL ---
@login_required
def exportar_fluxo_excel(request):
    # Pega os mesmos parâmetros da tela
    ano_get = request.GET.get('ano', str(timezone.now().year))
    try:
        ano_atual = int(str(ano_get).replace('.', ''))
    except:
        ano_atual = timezone.now().year
    
    view_mode = request.GET.get('view', 'mensal')
    status_view = request.GET.get('status_view', 'realizado')
    mes_atual = int(request.GET.get('mes', timezone.now().month))

    # Busca os dados
    data = _processar_dados_fluxo(request, ano_atual, mes_atual, view_mode, status_view)

    # Cria o Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fluxo {ano_atual}"

    # Estilos
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    green_font = Font(color="006400", bold=True)
    red_font = Font(color="8B0000", bold=True)
    blue_font = Font(color="00008B", bold=True)
    
    # Cabeçalho
    headers = ["Nome / Descrição"] + data['colunas_labels'] + ["Total"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill

    # -- RECEITAS --
    ws.append(["RECEITAS"])
    ws[ws.max_row][0].font = green_font

    for item in data['dados_receita']:
        row = [item['nome']] + item['valores'] + [item['total']]
        ws.append(row)
    
    # Total Receitas
    row_total_rec = ["TOTAL RECEITAS"] + data['totais_receita_col'] + [data['total_receitas_ano']]
    ws.append(row_total_rec)
    for cell in ws[ws.max_row]: cell.font = green_font

    ws.append([]) # Linha vazia

    # -- DESPESAS --
    ws.append(["DESPESAS"])
    ws[ws.max_row][0].font = red_font

    for item in data['dados_despesa']:
        row = [item['nome']] + item['valores'] + [item['total']]
        ws.append(row)
    
    # Total Despesas
    row_total_desp = ["TOTAL DESPESAS"] + data['totais_despesa_col'] + [data['total_despesas_ano']]
    ws.append(row_total_desp)
    for cell in ws[ws.max_row]: cell.font = red_font

    ws.append([])

    # -- SALDO --
    row_saldo = ["SALDO LÍQUIDO"] + data['saldo_mensal'] + [data['saldo_acumulado_ano']]
    ws.append(row_saldo)
    for cell in ws[ws.max_row]: cell.font = blue_font

    # Resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Fluxo_Analitico_{ano_atual}.xlsx'
    wb.save(response)
    return response

# --- VIEW EXPORTAR PDF (Versão xhtml2pdf - Compatível com Windows) ---
@login_required
def exportar_fluxo_pdf(request):
    # Mesma lógica de parâmetros
    ano_get = request.GET.get('ano', str(timezone.now().year))
    try:
        ano_atual = int(str(ano_get).replace('.', ''))
    except:
        ano_atual = timezone.now().year
    
    view_mode = request.GET.get('view', 'mensal')
    status_view = request.GET.get('status_view', 'realizado')
    mes_atual = int(request.GET.get('mes', timezone.now().month))

    # Chama o cérebro para pegar os dados
    data = _processar_dados_fluxo(request, ano_atual, mes_atual, view_mode, status_view)
    
    context = {
        'ano': ano_atual,
        'mes': mes_atual,
        'view_mode': view_mode,
        'status_view': status_view,
        'is_pdf': True,
        **data
    }

    # Renderiza o HTML
    template_path = 'relatorios/fluxo_analitico_pdf.html'
    template = get_template(template_path)
    html = template.render(context)

    # Cria a resposta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=Fluxo_Analitico_{ano_atual}.pdf'

    # Gera o PDF usando xhtml2pdf (pisa)
    pisa_status = pisa.CreatePDF(
       html, dest=response
    )

    if pisa_status.err:
       return HttpResponse('Ocorreu um erro ao gerar o PDF <pre>' + html + '</pre>')
       
    return response